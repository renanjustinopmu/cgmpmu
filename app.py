# app.py - Sistema completo de registro de horas (Flask + SQLite)
# Salve este arquivo como app.py e rode:
# python -m venv venv
# source venv/bin/activate  (ou venv\Scripts\activate no Windows)
# pip install flask
# python app.py

from flask import Flask, render_template_string, request, redirect, session, send_file, Response, jsonify
import io
import csv
import json
from datetime import date
import os
import psycopg2
import psycopg2.extras
from psycopg2 import IntegrityError

from datetime import datetime

import requests

def parse_data_excel(valor):
    if not valor:
        return None

    if isinstance(valor, datetime):
        return valor

    try:
        return datetime.strptime(str(valor).strip(), "%d/%m/%Y %H:%M:%S")
    except ValueError:
        try:
            return datetime.strptime(str(valor).strip(), "%d/%m/%Y")
        except ValueError:
            return None


def data_padrao_2026():
    hoje = date.today()
    if hoje.year == 2026:
        return hoje.isoformat()
    return "2026-01-01"

from datetime import datetime, date

def fmt(d):
    if not d:
        return ""

    # Se vier como date ou datetime (PostgreSQL / Supabase)
    if isinstance(d, (date, datetime)):
        return d.strftime("%d/%m/%Y")

    # Se vier como string yyyy-mm-dd
    if isinstance(d, str):
        try:
            return datetime.strptime(d[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
        except:
            return d

    return str(d)

def fmt_br(valor):
    try:
        return f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return "0,00"

def parse_hora(h):
    if not h:
        return None
    h = h.strip()
    if len(h) == 5:        # HH:MM
        return datetime.strptime(h, "%H:%M")
    else:                 # HH:MM:SS
        return datetime.strptime(h, "%H:%M:%S")

from datetime import datetime, date

from datetime import datetime, date

def calcular_prazo(dt_inicio, dt_previsao_fim, dt_conclusao=None):

    if not dt_inicio or not dt_previsao_fim:
        return "", ""

    try:
        inicio = datetime.strptime(str(dt_inicio)[:10], "%Y-%m-%d").date()
        fim = datetime.strptime(str(dt_previsao_fim)[:10], "%Y-%m-%d").date()

        prazo_total = (fim - inicio).days

        if prazo_total <= 0:
            return "", ""

        # ==================================================
        # 🔵 1) SE JÁ CONCLUÍDO
        # ==================================================
        if dt_conclusao:
            conclusao = datetime.strptime(str(dt_conclusao)[:10], "%Y-%m-%d").date()

            if conclusao <= fim:
                dias_antes = (fim - conclusao).days
                return (
                    f"{prazo_total} dias",
                    f"<span style='color:#16a34a;font-weight:bold;'>Concluído {dias_antes} dias antes ✔</span>"
                )
            else:
                atraso = (conclusao - fim).days
                return (
                    f"{prazo_total} dias",
                    f"<span style='color:#dc2626;font-weight:bold;'>Atraso de {atraso} dias ⚠</span>"
                )

        # ==================================================
        # 🔵 2) EM ANDAMENTO
        # ==================================================
        hoje = date.today()

        # Se ainda não iniciou
        if hoje < inicio:
            restante = prazo_total
        else:
            restante = (fim - hoje).days

        # 🔴 Já estourou
        if restante < 0:
            return (
                f"{prazo_total} dias",
                f"<span style='color:#dc2626;font-weight:bold;'>Atrasado {-restante} dias ⚑</span>"
            )

        # 🔹 Percentual do prazo restante
        percentual_restante = restante / prazo_total
        percentual_restante = max(0, min(1, percentual_restante))

        # 🔹 Regra de cores baseada no prazo restante
        if percentual_restante >= 0.7:
            bandeira = "<span style='color:#16a34a;font-size:18px;'>⚑</span>"
        elif percentual_restante >= 0.3:
            bandeira = "<span style='color:#facc15;font-size:18px;'>⚑</span>"
        else:
            bandeira = "<span style='color:#dc2626;font-size:18px;'>⚑</span>"

        return f"{prazo_total} dias", f"{restante} dias {bandeira}"

    except Exception as e:
        print("Erro calcular_prazo:", e)
        return "", ""

app = Flask(__name__)
app.secret_key = 'troque_esta_chave'

# -------------------------
# Helpers
# -------------------------
DATABASE_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql://postgres.dppcedgeybwhcochkjqv:cgmagdt2025@aws-1-sa-east-1.pooler.supabase.com:5432/postgres"
)

def get_db():
    return psycopg2.connect(
        DATABASE_URL,
        cursor_factory=psycopg2.extras.RealDictCursor
    )

def init_db():
    con = get_db()
    cur = con.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS colaboradores (
        id SERIAL PRIMARY KEY,
        nome TEXT,
        login TEXT UNIQUE,
        senha TEXT,
        perfil TEXT
    );
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS projeto_paint (
        id SERIAL PRIMARY KEY,
        classificacao TEXT,
        item_paint TEXT UNIQUE,
        tipo_atividade TEXT,
        objeto TEXT,
        objetivo_geral TEXT,
        dt_ini DATE,
        dt_fim DATE,
        hh_atual INTEGER DEFAULT 0
    );
    """)

        # OS
    cur.execute('''
CREATE TABLE IF NOT EXISTS os (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    codigo TEXT UNIQUE,
    item_paint TEXT,
    resumo TEXT,
    unidade TEXT,
    supervisao TEXT,
    coordenacao TEXT,
    equipe TEXT,
    observacao TEXT,
    status TEXT,
    plan INTEGER DEFAULT 0,
    exec INTEGER DEFAULT 0,
    rp INTEGER DEFAULT 0,
    rf INTEGER DEFAULT 0,
    dt_conclusao TEXT
)
''')

    cur.execute("""
    CREATE TABLE IF NOT EXISTS delegacoes (
        id SERIAL PRIMARY KEY,
        requisicoes TEXT,
        os_codigo TEXT,
        colaborador_id INTEGER REFERENCES colaboradores(id),
        data_inicio DATE,
        status TEXT DEFAULT 'Em Andamento',
        grau TEXT,
        data_fim DATE,
        criterio TEXT
    );
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS horas (
        id SERIAL PRIMARY KEY,
        colaborador_id INTEGER REFERENCES colaboradores(id),
        data DATE,
        item_paint TEXT,
        os_codigo TEXT,
        atividade TEXT,
        duracao TEXT,
        hora_inicio TIME,
        hora_fim TIME,
        duracao_minutos INTEGER,
        delegacao_id INTEGER REFERENCES delegacoes(id),
        observacoes TEXT
    );
    """)
    
    cur.execute("""
    CREATE TABLE IF NOT EXISTS atendimentos (
        id SERIAL PRIMARY KEY,
    
        -- vínculo com o lançamento de horas
        hora_id INTEGER REFERENCES horas(id) ON DELETE CASCADE,
        colaborador_id INTEGER REFERENCES colaboradores(id),
    
        -- identificação da OS
        os_codigo TEXT,
        os_resumo TEXT,
    
        -- dados do atendimento
        responsaveis_consultoria TEXT,   -- "Nome A, Nome B"
        macro TEXT,
        diretoria TEXT,
        atividade TEXT,
        data_consultoria DATE,
        assunto TEXT,
        participantes_externos TEXT,
        entidades TEXT,                  -- "CM, SEGOV, SMGAS"
        meio_contato TEXT,               -- Presencial | Email | Telefone
        observacao TEXT,
    
        -- dados derivados da hora
        duracao_minutos INTEGER,
        data_lancamento DATE,
    
        created_at TIMESTAMP DEFAULT NOW()
    );
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS consultorias (
        id SERIAL PRIMARY KEY,
    
        -- vínculo com o lançamento de horas
        hora_id INTEGER REFERENCES horas(id) ON DELETE CASCADE,
        colaborador_id INTEGER REFERENCES colaboradores(id),
    
        -- identificação da OS
        os_codigo TEXT,
        os_resumo TEXT,
    
        -- dados da consultoria
        responsaveis TEXT,   -- "Nome A, Nome B"
        data_consul DATE,
        assunto TEXT,
        secretarias TEXT,                  -- "CM, SEGOV, SMGAS"
        meio TEXT,               -- Presencial | Email | Telefone
        tipo TEXT,
        palavras_dhave TEXT,
        num_oficio TEXT,
        observacao TEXT,
    
        -- dados derivados da hora
        duracao_minutos INTEGER,
        data_lancamento DATE,
    
        created_at TIMESTAMP DEFAULT NOW()
    );

        CREATE TABLE IF NOT EXISTS requisicoes (
    id SERIAL PRIMARY KEY,

    -- chave de negócio
    chave TEXT UNIQUE,               -- ex: 1000/2026/CGM
    requisicao_num TEXT,             -- 1000/2026
    sigla TEXT,                      -- CGM

    -- dados importados do XLSX
    secretaria TEXT,
    tipo_documento TEXT,
    valor_requisicao NUMERIC(15,2),
    nome_solicitante TEXT,
    data_criacao TIMESTAMP,
    status_atual TEXT,
    data_tramitacao TIMESTAMP,
    natureza_despesa TEXT,
    item_despesa TEXT,
    nome_fornecedor TEXT,
    edital TEXT,
    contrato TEXT,
    data_medicao TEXT,
    data_liquidacao TEXT,
    empenho TEXT,
    ficha_despesa TEXT,

    -- controle de carga
    data_corte DATE,

    -- CAMPOS DE ANÁLISE (preenchidos depois)
    tipo TEXT CHECK (tipo IN ('CONTRATAÇÃO','LIQUIDAÇÃO','ADITAMENTO')),
    status_analise TEXT CHECK (status_analise IN ('ANDAMENTO','ANALISANDO','ANALISADO')),
    criterio TEXT CHECK (criterio IN ('MATERIALIDADE','RELEVÂNCIA','RISCO','ENGENHARIA')),

    servidor_id INTEGER REFERENCES colaboradores(id),

    nota TEXT CHECK (nota IN ('SIM','NÃO')),
    num_nota TEXT,
    oficio TEXT,
    monitoramento TEXT CHECK (monitoramento IN ('SIM','NÃO')),
    monitoramento_resposta TEXT,

    observacoes TEXT,
    valor_posterior NUMERIC(15,2),

    created_at TIMESTAMP DEFAULT NOW()
);
    
    """)

    con.commit()
    con.close()

# -------------------------
# Seed inicial
# -------------------------
COLABS = [
    "Alexandra Kátia", "Ana Paula Vilela", "Anatole Reis", "Aurélio Feitosa", "Caprice Cardoso", "Carlos Augusto",
    "Fernanda Lima", "Grazielle Carrijo", "Laianne Fogaça", "Marcelo Marques", "Maria Cristina", "Mariana Mota",
    "Mariana Cavanha", "Michelle Terêncio", "Paula Renata", "Paulo Sérgio", "Priscilla da Silva", "Syria Galvão",
    "Thamy Ponciano", "Tiago Pinheiro"
]

ITEMS_PAINT = [
    'O-1', 'O-2', 'O-3', 'O-4', 'O-5', 'O-6', 'O-7',
    'P-1', 'P-2', 'P-3', 'P-4', 'P-5', 'P-6', 'P-7', 'P-8', 'P-10', 'P-11', 'P-12', 'P-13', 'P-14', 'P-15', 'P-16',
    'P-17', 'P-18', 'P-19', 'P-20', 'P-21', 'P-22', 'P-23', 'P-24', 'P-25', 'P-26', 'P-27', 'P-28', 'P-29', 'P-30',
    'P-31', 'P-32', 'P-33', 'P-34', 'P-35', 'P-36', 'P-37', 'P-38', 'P-39', 'P-40', 'P-41', 'P-42', 'P-43', 'P-44',
    'P-45', 'P-46', 'P-47', 'P-48'
]


# --------- Seed compatível com Flask 3 ---------

def executar_seed():
    init_db()
    con = get_db()
    cur = con.cursor()

    # seed colaboradores
    for c in COLABS:
        login = c.lower().replace(' ', '.')
        try:
            cur.execute('INSERT INTO colaboradores (nome,login,senha,perfil) VALUES (%s,%s,%s,%s)',
                        (c, login, '123', 'comum'))
        except IntegrityError:
            con.rollback()

    # admin
    try:
        cur.execute('INSERT INTO colaboradores (nome,login,senha,perfil) VALUES (%s,%s,%s,%s)',
                    ('Renan Justino', 'renan.justino', '123', 'admin'))
    except IntegrityError:
        con.rollback()

    # projeto_paint base
    for it in ITEMS_PAINT:
        try:
            cur.execute(
                'INSERT INTO projeto_paint (classificacao, item_paint, tipo_atividade, objeto, objetivo_geral, dt_ini, dt_fim, hh_atual) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)',
                ('Obrigatório', it, '', '', '', None, None, 0))
        except IntegrityError:
            con.rollback()

    con.commit()
    con.close()


# -------------------------
# Templates base (usamos render_template_string para facilitar entrega única)
# -------------------------
BASE = """
<!doctype html>
<html>
<head>
<style>

    /* ---------------------- TEMA GLOBAL ---------------------- */
    body {
        background: #e9f1fb;
        color: #1d2a3a;
        font-family: "Segoe UI", sans-serif;
        margin: 0;
        padding: 20px 40px;   /* 👈 cria margem interna lateral */
    }

    h1, h2, h3, h4 {
        color: #1e4f9c;
        margin-bottom: 10px;
    }

    header {
    display:flex;
    justify-content:space-between;
    align-items:center;
}

    .small {
        font-size:0.9em;
        color:#555;
    }

    a { 
        color: #1e4f9c; 
        text-decoration: none; 
    }
    a:hover { 
        color: #3b7ae0; 
    }

    /* ---------------------- CONTAINER ---------------------- */
    .container {
        max-width: 1200px;
        margin: auto;
        padding: 20px;
    }

    /* ---------------------- BOTÕES ---------------------- */
    .btn {
        background: #1e74d9;
        color: white !important;
        padding: 8px 16px;
        border-radius: 6px;
        border: none;
        font-size: 14px;
        cursor: pointer;
        display: inline-block;
        text-align: center;
        transition: 0.2s;
    }

    .btn:hover {
        background: #3a8af0;
    }

    .btn-danger {
        background: #d9534f !important;
    }

    .btn-danger:hover {
        background: #e6736f !important;
    }

    /* ---------------------- INPUTS / SELECTS / TEXTAREA ---------------------- */
    input, select, textarea {
        background: #ffffff;
        border: 1px solid #c3d4ea;
        color: #1d2a3a;
        padding: 8px;
        border-radius: 6px;
        width: 100%;
        margin-top: 4px;
        margin-bottom: 12px;
        font-size: 14px;
    }

    input:focus, select:focus, textarea:focus {
        border-color: #1e74d9;
        outline: none;
        box-shadow: 0 0 0 3px rgba(30, 116, 217, 0.28);
    }

    textarea {
        min-height: 100px;
        resize: vertical;
    }

    /* ---------------------- CARD ---------------------- */
    .card {
        background: #ffffff;
        border: 1px solid #d3e0f0;
        padding: 20px;
        border-radius: 12px;
        margin-bottom: 25px;
        box-shadow: 0 2px 8px rgba(0, 0, 0, 0.06);

        /* texto corrigido */
        color: #1d2a3a !important;
    }

    /* ---------------------- TABELAS ---------------------- */
    table {
        border-collapse: collapse;
        margin-top: 15px;
        background: #ffffff;
        border-radius: 12px;
        overflow: hidden;
        font-size: 14px;
        box-shadow: 0 2px 6px rgba(0,0,0,0.05);
    }

    th {
        background: #dce8fb;
        padding: 10px;
        text-align: left;
        color: #1e4f9c;
        font-weight: 600;
        border-bottom: 1px solid #c7d5ee;
    }

    td {
        padding: 9px 12px;
        border-top: 1px solid #e1e8f2;
        color: #1d2a3a;
    }

    tr:nth-child(even) {
        background: #f4f8ff;
    }

    tr:hover {
        background: #e9f2ff;
    }

    /* ---------------------- TOPBAR ---------------------- */
    .topbar {
        background: #1e74d9;
        padding: 15px 25px;
        color: white !important;
        font-size: 18px;
        font-weight: bold;
        display: flex;
        align-items: center;
        column-gap: 20px;
        flex-wrap: wrap;
    }

    .topbar a {
        color: white !important;
        font-weight: 500;
        opacity: 0.95;
    }

    .topbar a:hover {
        opacity: 1;
    }

    /* ---------------------- MENU LINKS (corrigido) ---------------------- */
.menu-links a {
    margin-right: 16px;
    padding: 7px 12px;
    border-radius: 6px;

    /* Fundo mais sólido e claro */
    background: #ffffff;

    /* Texto azul escuro para boa leitura */
    color: #1e4f9c !important;

    font-size: 14px;
    font-weight: 600;
    border: 1px solid #c5d7f2;

    transition: 0.2s;
}

.menu-links a:hover {
    background: #d9e8ff;        /* azul claro no hover */
    color: #124a99 !important;  /* azul mais forte */
    border-color: #a8c6f2;
}
/* ---------------------- coluna requisicoes largura ---------------------- */
    .col-requisicoes {
        max-width: 320px;
        white-space: normal;
        word-break: break-word;
    }

/* ---------------------- MENU AGRUPADO ---------------------- */

.navbar {
    display:flex;
    gap:14px;
    flex-wrap:wrap;
    margin-top:10px;
}

.menu-group {
    position:relative;
}

.menu-btn {
    padding:8px 14px;
    border-radius:8px;
    font-weight:600;
    border:none;
    cursor:pointer;
    font-size:14px;
}

.dropdown {
    display:none;
    position:absolute;
    top:38px;
    left:0;
    background:white;
    border:1px solid #d8e2f0;
    border-radius:10px;
    box-shadow:0 6px 14px rgba(0,0,0,0.08);
    min-width:180px;
    z-index:100;
}

.dropdown a {
    display:block;
    padding:9px 12px;
    color:#1d2a3a !important;
    font-size:14px;
}

.dropdown a:hover {
    background:#f2f6ff;
}

.dropdown {
    display:none;
    position:absolute;
    top:38px;
    left:0;
    background:white;
    border:1px solid #d8e2f0;
    border-radius:10px;
    box-shadow:0 6px 14px rgba(0,0,0,0.08);
    min-width:180px;
    z-index:100;
}

.menu-group.active .dropdown {
    display:block;
}

/* cores por grupo */

.menu1 { background:#e8f1ff; color:#1e4f9c; }
.menu2 { background:#e8fff1; color:#1c7a48; }
.menu3 { background:#fff3e8; color:#b76a1c; }
.menu4 { background:#f0e8ff; color:#6a3fb2; }
.menu5 { background:#ffe8e8; color:#a83232; }
.menu6 { background:#e9f8ff; color:#1c7aa6; }
.menu7 { background:#eeeeee; color:#444; }
    
/* ---------------------- CSS MINHAS DELEGACOES ---------------------- */
.btn {
    padding:6px 10px;
    border:none;
    cursor:pointer;
    border-radius:4px;
    font-weight:bold;
}

.all { background:#ccc; }
.andamento { background:#ffb3b3; }
.analisando { background:#ffe699; }
.analisado { background:#b6f2c2; }

tr.andamento { background:#ffe5e5; }
tr.analisando { background:#fff6d6; }
tr.analisado { background:#e6ffed; }

</style>


  <meta charset='utf-8'>
  <title>Sistema de Horas</title>
  <link rel="icon" type="image/png" href="https://i.ibb.co/M5ZcxYj6/favicon.png">

</head>
<body>
<header>
  <div>
    <h2>Sistema de Registro de Horas</h2>
    {% if user %}
      <div class='small'>Logado como: <strong>{{user}}</strong> ({{perfil}})</div>
    {% endif %}
  </div>
  <div>
    {% if user %}
      <div class="navbar">
        
        <!-- MENU -->
        <div class="menu-group">
        <button class="menu-btn menu1">🏠 Menu</button>
        <div class="dropdown">
        <a href="/menu">🏠 Menu</a>
        <a href="/lancar">⏱ Lançar Horas</a>
        <a href="/avisos">📢 Avisos</a> 
        </div>
        </div>
        
        <!-- RELATÓRIOS -->
        <div class="menu-group">
        <button class="menu-btn menu2">📊 Relatórios</button>
        <div class="dropdown">
        <a href="/relatorios">📊 Relatórios</a>
        </div>
        </div>
        
        <!-- REQUISIÇÕES -->
        <div class="menu-group">
        <button class="menu-btn menu3">📄 Requisições/N.A</button>
        <div class="dropdown">
        {% if perfil == 'admin' %}
        <a href="/requisicoes">📄 Requisições</a>
        <a href="/requisicoes/importar">📥 Import</a>
        {% else %}
        <a href="/minhas_delegacoes">📄 Requisições</a>
        {% endif %}
        
        <a href="/notas-auditoria">🧾 Notas Auditoria</a>
        </div>
        </div>
        
        <!-- ATENDIMENTOS -->
        <div class="menu-group">
        <button class="menu-btn menu4">🧑‍💻 Atendimentos</button>
        <div class="dropdown">
        
        {% if perfil == 'admin' %}
        <a href="/atendimentos">🧑‍💻 Atendimentos</a>
        {% else %}
        <a href="/atendimentos">🧑‍💻 Meus Atendimentos</a>
        {% endif %}
        
        {% if perfil == 'admin' %}
        <a href="/consultorias">💬 Consultorias</a>
        {% else %}
        <a href="/consultorias">💬 Minhas Consultorias</a>
        {% endif %}
        
        </div>
        </div>
        
        <!-- GERENCIAL -->
        {% if perfil=='admin' %}
        <div class="menu-group">
        <button class="menu-btn menu5">⚙ Gerencial</button>
        <div class="dropdown">
        <a href="/paint">🎨 PAINT</a>
        <a href="/os">🧾 O.S - Cadastro</a>
        <a href="/os/gestao">🗂️ O.S - Gestão</a>
        <a href="/os/rh">👨‍👩‍👧‍👦 O.S - RH</a>
        <a href="/colaboradores">👥 Colaboradores</a>
        </div>
        </div>
        {% endif %}
        
        <!-- PAINEL -->
        <div class="menu-group">
        <button class="menu-btn menu6">📈 Painel</button>
        <div class="dropdown">
        <a href="/admin_projetos">📂 Projetos</a>
        <a href="/painel_requisicoes">📊 Painel</a>
        
        {% if perfil == 'admin' %}
        <a href="/visao">📉 Visão/h</a>
        {% endif %}
        
        </div>
        </div>
        
        <!-- SAIR -->
        <div class="menu-group">
        <button class="menu-btn menu7">🚪 Conta</button>
        <div class="dropdown">
        <a href="/logout">🚪 Sair</a>
        </div>
        </div>
        
        </div>
    {% endif %}
  </div>
</header>
<hr>
<div>
  {% block content %}{% endblock %}
</div>
<script>

document.querySelectorAll(".menu-btn").forEach(btn => {

    btn.addEventListener("click", function(e){

        let group = this.parentElement

        document.querySelectorAll(".menu-group").forEach(g=>{
            if(g !== group){
                g.classList.remove("active")
            }
        })

        group.classList.toggle("active")

        e.stopPropagation()
    })

})

document.addEventListener("click", function(){
    document.querySelectorAll(".menu-group").forEach(g=>{
        g.classList.remove("active")
    })
})

</script>
</body>
</html>
"""

@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        login = request.form.get('login')
        senha = request.form.get('senha')

        con = get_db()
        cur = con.cursor()
        cur.execute(
            'SELECT id,nome,perfil FROM colaboradores WHERE login=%s AND senha=%s',
            (login, senha)
        )
        row = cur.fetchone()
        con.close()

        if row:
            session['user'] = row['nome']
            session['perfil'] = row['perfil']
            session['user_id'] = row['id']
            return redirect('/menu')
        else:
            conteudo = f"""
                {HEADER_LOGIN}
                <h3 style="text-align:center;">Login</h3>
                <p style="color:red;text-align:center;">Login inválido</p>
                {LOGIN_FORM}
            """

            return render_template_string(
                BASE.replace('{% block content %}{% endblock %}', conteudo),
                user=None,
                perfil=None
            )

    conteudo = f"""
        {HEADER_LOGIN}
        <h3 style="text-align:center;">Login</h3>
        {LOGIN_FORM}
    """

    return render_template_string(
        BASE.replace('{% block content %}{% endblock %}', conteudo),
        user=None,
        perfil=None
    )

LOGIN_FORM = """
<form method="post" style="max-width:380px;margin:0 auto;">
    <div>
        <label>Login</label>
        <input name="login" required>
    </div>

    <div>
        <label>Senha</label>
        <input type="password" name="senha" required>
    </div>

    <div style="margin-top:15px;">
        <button class="btn" style="width:100%;">Entrar</button>
    </div>
</form>
"""

HEADER_LOGIN = """
<div style="width:100%; text-align:center; margin-bottom:30px;">
    <img
        src="https://i.ibb.co/gFv5XWJp/8-Controladoria-geral.png"
        alt="Controladoria-Geral do Município"
        style="max-height:90px; margin-bottom:15px;"
    >

    <h1 style="
        font-size:28px;
        font-weight:800;
        color:#000;
        margin:0;
        letter-spacing:1px;
    ">
        Controladoria-Geral<br>
    </h1>

    <div style="
        width:100%;
        height:6px;
        background:#000;
        margin-top:15px;
    "></div>
</div>
"""

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/')

@app.route('/menu')
def menu():
    if 'user' not in session:
        return redirect('/')

    status_options = [
        "Não Iniciado",
        "Em Andamento",
        "Pausado",
        "Aguardando Servidor",
        "Concluído"
    ]

    def badge_class(s):
        return {
            "Não Iniciado": "st-nao",
            "Em Andamento": "st-and",
            "Pausado": "st-pausado",
            "Aguardando Servidor": "st-aguard",
            "Concluído": "st-ok"
        }.get(s, "")

    user = session['user']

    con = get_db()
    cur = con.cursor()

    cur.execute("""
        SELECT *
        FROM os
        WHERE 
        (
            equipe ILIKE %s OR
            coordenacao ILIKE %s OR
            supervisao ILIKE %s
        )
        AND (status IS NULL OR status <> 'Concluido')
        ORDER BY codigo
    """, (f"%{user}%", f"%{user}%", f"%{user}%"))

    oss = cur.fetchall()

    cur.execute("""
        SELECT os_codigo, status, observacao
        FROM os_status_user
        WHERE colaborador = %s
    """, (user,))

    status_map = {
        r['os_codigo']: r for r in cur.fetchall()
    }

    con.close()

    html = f"""
    <style>
    .top-actions {{
        display:flex;
        justify-content:center;
        gap:20px;
        margin:20px 0;
    }}
    
    .btn-big {{
        border:2px solid #2c5aa0;
        padding:10px 20px;
        border-radius:10px;
        text-decoration:none;
        font-weight:bold;
        color:#2c5aa0;
    }}
    
    .linha {{
        display:grid;
        grid-template-columns: 1fr 2fr 1fr 3fr 70px 70px; /* ↑ espaço */
        align-items:center;
        border:2px solid #2c5aa0;
        border-radius:10px;
        padding:10px;
        margin-bottom:10px;
        gap:15px; /* ↑ espaçamento geral */
    }}
    
    .linha input {{
        width:100%;
        border:none;
        background:transparent;
        border-bottom:1px solid #ccc;
    }}
    
    .icon-btn {{
        width:42px;
        height:42px;
        min-width:42px;   /* evita encolher */
        min-height:42px;
        border-radius:50%;
        display:flex;
        align-items:center;
        justify-content:center;
        text-decoration:none;
        font-size:18px;   /* ↑ ícone maior */
    }}
    
    .clock {{    
    background:#dfe9f8;   /* mais forte */
    color:#2c5aa0;        /* deixa o ícone mais visível */
    }}
    
    .info {{ background:#2c5aa0; color:white; }}
    
    .linha a.icon-btn {{
    justify-self:center;
}}

    .status-box {{
        display:flex;
        align-items:center;
        gap:5px;
    }}
    
    .badge {{
        padding:4px 8px;
        border-radius:8px;
        font-size:12px;
        color:white;
        font-weight:bold;
    }}
    
    .st-nao {{ background:#7f8c8d; }}
    .st-and {{ background:#3498db; }}
    .st-pausado {{ background:#e67e22; }}
    .st-aguard {{ background:#9b59b6; }}
    .st-ok {{ background:#27ae60; }}
    
    select.status {{
        border:none;
        background:transparent;
    }}
    </style>

    <div class="top-actions">
        <a class="btn-big" href="/lancar">⏱ LANÇAR HORAS</a>
        <a class="btn-big" href="/relatorios">📊 RELATÓRIO</a>
    </div>

    <h2>Minhas Ordens de Serviço</h2>
    """

    for r in oss:
        st = status_map.get(r['codigo'], {})

        html += f"""
        <div class="linha" data-os="{r['codigo']}">
        
            <div><b>{r['codigo']}</b></div>
        
            <div>{r['resumo']}</div>
        
            <div class="status-box">
                <span class="badge {badge_class(st.get('status'))}">
                    {st.get('status') or '-'}
                </span>
        
                <select class="status">
                    <option value=""></option>
        """
        for opt in status_options:
            sel = "selected" if st.get("status") == opt else ""
            html += f"<option {sel}>{opt}</option>"
        
        html += f"""
                </select>
            </div>
        
            <div>
                <input class="obs"
                       value="{st.get('observacao','') or ''}"
                       placeholder="Observação">
            </div>
        
            <a class="icon-btn clock"
               href="/lancar?os={r['codigo']}">⏱</a>
        
            <a class="icon-btn info"
               href="/os/view/{r['id']}">i</a>
        
        </div>
        """

    html += """
    <script>
    document.querySelectorAll(".linha input").forEach(inp=>{
        inp.addEventListener("change", function(){

            let row = this.closest(".linha")
            let os = row.dataset.os
            let status = row.querySelector(".status").value
            let obs = row.querySelector(".obs").value

            fetch("/menu/os_status", {
                method:"POST",
                headers:{"Content-Type":"application/x-www-form-urlencoded"},
                body:
                    "os="+encodeURIComponent(os)+
                    "&status="+encodeURIComponent(status)+
                    "&obs="+encodeURIComponent(obs)
            })
        })
    })
    </script>
    <script>
document.querySelectorAll(".linha").forEach(row=>{

    const select = row.querySelector(".status")
    const obs = row.querySelector(".obs")
    const badge = row.querySelector(".badge")

    function salvar(){
        fetch("/menu/os_status", {
            method:"POST",
            headers:{"Content-Type":"application/x-www-form-urlencoded"},
            body:
                "os="+encodeURIComponent(row.dataset.os)+
                "&status="+encodeURIComponent(select.value)+
                "&obs="+encodeURIComponent(obs.value)
        })

        badge.innerText = select.value

        badge.className = "badge " + ({
            "Não Iniciado":"st-nao",
            "Em Andamento":"st-and",
            "Pausado":"st-pausado",
            "Aguardando Servidor":"st-aguard",
            "Concluído":"st-ok"
        }[select.value] || "")
    }

    select.addEventListener("change", salvar)
    obs.addEventListener("change", salvar)

})
</script>
    """

    return render_template_string(
        BASE.replace('{% block content %}{% endblock %}', html),
        user=session['user'],
        perfil=session['perfil']
    )

@app.route('/menu/os_status', methods=['POST'])
def salvar_status_os():
    if 'user' not in session:
        return '', 403

    os_codigo = request.form.get('os')
    status = request.form.get('status')
    obs = request.form.get('obs')
    user = session['user']

    con = get_db()
    cur = con.cursor()

    cur.execute("""
        INSERT INTO os_status_user (os_codigo, colaborador, status, observacao)
        VALUES (%s,%s,%s,%s)
        ON CONFLICT (os_codigo, colaborador)
        DO UPDATE SET
            status = EXCLUDED.status,
            observacao = EXCLUDED.observacao,
            updated_at = now()
    """, (os_codigo, user, status, obs))

    con.commit()
    con.close()

    return '', 204

@app.route('/avisos', methods=['GET','POST'])
def avisos():
    if 'user' not in session:
        return redirect('/')

    from datetime import timedelta

    def fmt_data(dt):
        if not dt:
            return ''
        from datetime import datetime, timedelta
    
        if isinstance(dt, str):
            try:
                dt = datetime.fromisoformat(dt)
            except:
                return dt  # fallback
    
        return (dt - timedelta(hours=3)).strftime('%d/%m/%Y %H:%M')
        
    user = session['user']
    con = get_db()
    cur = con.cursor()

    if request.method == 'POST':
        msg = request.form.get('mensagem')
        parent = request.form.get('parent_id') or None

        delete_id = request.form.get('delete_id')

        if delete_id:
            cur.execute("""
                DELETE FROM avisos_posts
                WHERE id = %s AND colaborador = %s
            """, (delete_id, user))
            con.commit()
            return redirect('/avisos')

        cur.execute("""
            INSERT INTO avisos_posts (colaborador, mensagem, parent_id)
            VALUES (%s,%s,%s)
        """, (user, msg, parent))

        con.commit()

    cur.execute("""
        SELECT *
        FROM avisos_posts
        ORDER BY created_at DESC
        LIMIT 100
    """)

    posts = cur.fetchall()
    con.close()

    # organizar threads
    tree = {}
    for p in posts:
        tree.setdefault(p['parent_id'], []).append(p)

    def render(parent=None, nivel=0):
        html = ""
        for p in tree.get(parent, []):
            html += f"""
            <div style="
                margin-left:{nivel*20}px;
                border:1px solid #ccc;
                border-radius:10px;
                padding:10px;
                margin-bottom:10px;
                background:white;
            ">
                <div style="font-size:12px;color:#666">
                    <b>{p['colaborador']}</b> - {fmt_data(p['created_at'])}
                </div>

                <div>{p['mensagem']}</div>

                <div style="display:flex; gap:10px; margin-top:5px;">
                <button onclick="responder({p['id']})">Responder</button>
            
                <form method="post" style="display:inline;">
                    <input type="hidden" name="delete_id" value="{p['id']}">
                    <button style="background:#e74c3c;color:white;border:none;border-radius:5px;cursor:pointer;">
                        🗑
                    </button>
                </form>
            </div>

                <form method="post" id="resp_{p['id']}" style="display:none;margin-top:5px;">
                    <input type="hidden" name="parent_id" value="{p['id']}">
                    <textarea name="mensagem" style="width:100%;height:60px"></textarea>
                    <button>Enviar</button>
                </form>
            """

            html += render(p['id'], nivel+1)
            html += "</div>"

        return html

    html = """
    <h2>📢 Avisos</h2>

    <form method="post">
        <textarea name="mensagem" style="width:100%;height:80px"></textarea>
        <br><br>
        <button>Postar</button>
    </form>
    <br>
    """

    html += render()

    html += """
    <script>
    function responder(id){
        let f = document.getElementById("resp_"+id)
        f.style.display = (f.style.display=="none") ? "block" : "none"
    }
    </script>
    """

    return render_template_string(
        BASE.replace('{% block content %}{% endblock %}', html),
        user=session['user'],
        perfil=session['perfil']
    )

# -------------------------
# Colaboradores (admin)
# -------------------------
@app.route('/colaboradores', methods=['GET', 'POST'])
def colaboradores():
    if 'user' not in session:
        return redirect('/')
    if session['perfil'] != 'admin':
        return 'Acesso negado'

    ADMIN_MASTER_ID = 21
    admin_master = session.get("user_id") == ADMIN_MASTER_ID

    con = get_db()
    cur = con.cursor()
    acao = request.form.get('acao')

    # -------------------------
    # BLOQUEIO DE AÇÕES
    # -------------------------
    if acao and not admin_master:
        con.close()
        return "Ação não permitida para este usuário"

    # -------------------------
    # NOVO COLABORADOR
    # -------------------------
    if acao == 'novo':
        cur.execute("""
            INSERT INTO colaboradores (nome, login, senha, perfil)
            VALUES (%s,%s,%s,%s)
        """, (
            request.form['nome'],
            request.form['login'],
            request.form['senha'],
            request.form['perfil']
        ))
        con.commit()

    # -------------------------
    # EDITAR COLABORADOR
    # -------------------------
    elif acao == 'editar':
        cid = request.form['id']

        cur.execute("SELECT nome FROM colaboradores WHERE id=%s", (cid,))
        antigo = cur.fetchone()
        if not antigo:
            con.close()
            return "Colaborador não encontrado"

        cur.execute("""
            UPDATE colaboradores
            SET nome=%s, login=%s, perfil=%s
            WHERE id=%s
        """, (
            request.form['nome'],
            request.form['login'],
            request.form['perfil'],
            cid
        ))

        if request.form.get('senha'):
            cur.execute(
                "UPDATE colaboradores SET senha=%s WHERE id=%s",
                (request.form['senha'], cid)
            )

        if antigo['nome'] != request.form['nome']:
            cur.execute("""
                UPDATE atendimentos
                SET responsaveis_consultoria =
                    REPLACE(responsaveis_consultoria, %s, %s)
                WHERE responsaveis_consultoria ILIKE %s
            """, (
                antigo['nome'],
                request.form['nome'],
                f"%{antigo['nome']}%"
            ))

        con.commit()

    # -------------------------
    # EXCLUIR COLABORADOR
    # -------------------------
    elif acao == 'excluir':
        cid = request.form['id']

        cur.execute(
            "SELECT COUNT(*) AS qtd FROM horas WHERE colaborador_id=%s",
            (cid,)
        )
        if cur.fetchone()['qtd'] > 0:
            con.close()
            return "Não é possível excluir colaborador com horas lançadas"

        cur.execute("DELETE FROM delegacoes WHERE colaborador_id=%s", (cid,))
        cur.execute("DELETE FROM atendimentos WHERE colaborador_id=%s", (cid,))
        cur.execute("DELETE FROM colaboradores WHERE id=%s", (cid,))
        con.commit()

    # -------------------------
    # LISTAGEM
    # -------------------------
    cur.execute("""
        SELECT
            c.id,
            c.nome,
            c.login,
            c.perfil,
            COALESCE(
                ARRAY_AGG(DISTINCT h.item_paint)
                FILTER (WHERE h.item_paint IS NOT NULL),
                '{}'
            ) AS projetos,
            COALESCE(SUM(h.duracao_minutos), 0) AS total_minutos
        FROM colaboradores c
        LEFT JOIN horas h ON h.colaborador_id = c.id
        GROUP BY c.id
        ORDER BY c.nome
    """)
    colaboradores = cur.fetchall()
    con.close()

    # -------------------------
    # HTML
    # -------------------------
    html = "<h3>Colaboradores</h3>"

    # ➕ NOVO (somente admin 21)
    if admin_master:
        html += """
        <details style="margin-bottom:15px;">
            <summary style="cursor:pointer;">➕ Novo Colaborador</summary>
            <form method="post">
                <input type="hidden" name="acao" value="novo">
                <input name="nome" placeholder="Nome" required>
                <input name="login" placeholder="Login" required>
                <input name="senha" type="password" placeholder="Senha" required>
                <select name="perfil">
                    <option value="comum">comum</option>
                    <option value="admin">admin</option>
                </select>
                <button class="btn">Cadastrar</button>
            </form>
        </details>
        """

    html += """
    <table>
        <tr>
            <th>Nome</th>
            <th>Login</th>
            <th>Perfil</th>
            <th>Projetos PAINT</th>
            <th>Total Horas</th>
            <th>Ações</th>
        </tr>
    """

    for c in colaboradores:
        hh = c['total_minutos'] // 60
        mm = c['total_minutos'] % 60
        total = f"{hh:02d}:{mm:02d}"
        projetos = ", ".join(c['projetos']) if c['projetos'] else "-"

        html += f"""
        <tr>
            <td><a href="/colaborador/{c['id']}">{c['nome']}</a></td>
            <td>{c['login']}</td>
            <td>{c['perfil']}</td>
            <td>{projetos}</td>
            <td>{total}</td>
            <td>
        """

        # ✏️ / 🗑 somente admin 21
        if admin_master:
            html += f"""
                <details style="display:inline-block;">
                    <summary style="cursor:pointer;">✏️</summary>
                    <form method="post">
                        <input type="hidden" name="acao" value="editar">
                        <input type="hidden" name="id" value="{c['id']}">
                        <input name="nome" value="{c['nome']}" required>
                        <input name="login" value="{c['login']}" required>
                        <input name="senha" type="password" placeholder="Nova senha">
                        <select name="perfil">
                            <option value="comum" {'selected' if c['perfil']=='comum' else ''}>comum</option>
                            <option value="admin" {'selected' if c['perfil']=='admin' else ''}>admin</option>
                        </select>
                        <button class="btn">Salvar</button>
                    </form>
                </details>

                <form method="post" style="display:inline;"
                      onsubmit="return confirm('Excluir colaborador?');">
                    <input type="hidden" name="acao" value="excluir">
                    <input type="hidden" name="id" value="{c['id']}">
                    <button class="btn" style="background:#c00;">🗑</button>
                </form>
            """
        else:
            html += "—"

        html += "</td></tr>"

    html += "</table>"

    return render_template_string(
        BASE.replace('{% block content %}{% endblock %}', html),
        user=session['user'],
        perfil=session['perfil']
    )

# -------------------------
# Detalhes do Colaborador (admin)
# -------------------------
@app.route('/colaborador/<int:cid>')
def colaborador_detalhes(cid):
    if 'user' not in session:
        return redirect('/')
    if session['perfil'] != 'admin':
        return 'Acesso negado'

    con = get_db()
    cur = con.cursor()

    # -------------------------
    # Nome do colaborador
    # -------------------------
    cur.execute("SELECT nome FROM colaboradores WHERE id=%s", (cid,))
    col = cur.fetchone()
    if not col:
        con.close()
        return "Colaborador não encontrado"

    nome = col['nome']

    # -------------------------
    # Total por OS (agrupado)
    # -------------------------
    cur.execute("""
        SELECT
            os_codigo,
            item_paint,
            SUM(duracao_minutos) AS minutos
        FROM horas
        WHERE colaborador_id = %s
        GROUP BY os_codigo, item_paint
        ORDER BY item_paint, os_codigo
    """, (cid,))
    por_os = cur.fetchall()

    # -------------------------
    # Total por Item PAINT
    # -------------------------
    cur.execute("""
        SELECT
            item_paint,
            SUM(duracao_minutos) AS minutos
        FROM horas
        WHERE colaborador_id = %s
        GROUP BY item_paint
        ORDER BY item_paint
    """, (cid,))
    por_paint = cur.fetchall()

    con.close()

    # -------------------------
    # Helper
    # -------------------------
    def minutos_para_hhmm(minutos):
        minutos = int(minutos or 0)
        return f"{minutos//60:02d}:{minutos%60:02d}"

    # -------------------------
    # HTML
    # -------------------------
    html = f"<h3>Horas de {nome}</h3>"
    html += "<div style='display:flex; gap:40px; align-items:flex-start;'>"

    # ---- Tabela PAINT
    html += """
    <div style='width:45%;'>
        <h4>Total por Item PAINT</h4>
        <table>
            <tr><th>Item</th><th>Total (HH:MM)</th></tr>
    """
    for r in por_paint:
        html += f"""
        <tr>
            <td>{r['item_paint'] or '-'}</td>
            <td>{minutos_para_hhmm(r['minutos'])}</td>
        </tr>
        """
    html += "</table></div>"

    # ---- Tabela OS
    html += """
    <div style='width:50%;'>
        <h4>Total por O.S. (agrupado por Item PAINT)</h4>
        <table>
            <tr><th>OS</th><th>Item PAINT</th><th>Total (HH:MM)</th></tr>
    """
    for r in por_os:
        html += f"""
        <tr>
            <td>{r['os_codigo'] or '-'}</td>
            <td>{r['item_paint'] or '-'}</td>
            <td>{minutos_para_hhmm(r['minutos'])}</td>
        </tr>
        """
    html += "</table></div>"

    html += "</div>"
    html += "<br><a class='btn' href='/colaboradores'>Voltar</a>"

    return render_template_string(
        BASE.replace("{% block content %}{% endblock %}", html),
        user=session['user'],
        perfil=session['perfil']
    )

from datetime import datetime

# -------------------------
# Editar Projeto PAINT
# -------------------------
@app.route('/projeto/edit/<int:id>', methods=['GET', 'POST'])
def editar_projeto(id):
    if 'user' not in session:
        return redirect('/')
    if session['perfil'] != 'admin':
        return 'Acesso negado'

    con = get_db()
    cur = con.cursor()

    cur.execute("SELECT * FROM projeto_paint WHERE id = %s", (id,))
    projeto = cur.fetchone()

    if not projeto:
        con.close()
        return "Projeto não encontrado", 404

    item_antigo = projeto["item_paint"]

    if request.method == 'POST':
        classificacao = request.form.get('classificacao')
        item_novo = request.form.get('item_paint')
        tipo = request.form.get('tipo_atividade')
        objeto = request.form.get('objeto')
        objetivo = request.form.get('objetivo')
        dt_ini = request.form.get('dt_ini') or None
        dt_fim = request.form.get('dt_fim') or None
        hh_atual = request.form.get('hh_atual') or 0

        try:
            # ---- atualiza projeto ----
            cur.execute("""
                UPDATE projeto_paint
                SET classificacao=%s, item_paint=%s, tipo_atividade=%s, objeto=%s,
                    objetivo_geral=%s, dt_ini=%s, dt_fim=%s, hh_atual=%s
                WHERE id=%s
            """, (classificacao, item_novo, tipo, objeto,
                  objetivo, dt_ini, dt_fim, hh_atual, id))

            # ---- CASCADE MANUAL ----
            if item_antigo != item_novo:
                cur.execute(
                    "UPDATE os SET item_paint=%s WHERE item_paint=%s",
                    (item_novo, item_antigo)
                )
                cur.execute(
                    "UPDATE horas SET item_paint=%s WHERE item_paint=%s",
                    (item_novo, item_antigo)
                )

            con.commit()

        except Exception as e:
            con.rollback()
            con.close()
            return f"Erro ao atualizar projeto: {e}"

        con.close()
        return redirect('/paint')

    con.close()

    # ----- FORMULÁRIO HTML PRÉ-PREENCHIDO -----

    html = f"""
    <h3>Editar Projeto PAINT – {projeto['item_paint']}</h3>

    <form method='post'>
      <div>Classificação:
        <select name='classificacao'>
            <option value='Prioritário'  {"selected" if projeto["classificacao"] == "Prioritário" else ""}>Prioritário</option>
            <option value='Obrigatório'  {"selected" if projeto["classificacao"] == "Obrigatório" else ""}>Obrigatório</option>
            <option value='Complementar' {"selected" if projeto["classificacao"] == "Complementar" else ""}>Complementar</option>
            <option value='Novo'         {"selected" if projeto["classificacao"] == "Novo" else ""}>Novo</option>
        </select>
      </div>

      <div>Item PAINT:
        <input name='item_paint' value='{projeto["item_paint"]}' required>
      </div>

      <div>Tipo de Atividade:
        <input name='tipo_atividade' value='{projeto["tipo_atividade"] or ""}'>
      </div>

      <div>Objeto:
        <textarea name='objeto' rows='3'>{projeto["objeto"] or ""}</textarea>
      </div>

      <div>Objetivo Geral:
        <textarea name='objetivo' rows='3'>{projeto["objetivo_geral"] or ""}</textarea>
      </div>

      <div>Data Inicial:
        <input type='date' name='dt_ini' value='{projeto["dt_ini"] or ""}'>
      </div>

      <div>Data Final:
        <input type='date' name='dt_fim' value='{projeto["dt_fim"] or ""}'>
      </div>

      <div>HH Atual:
        <input type='number' name='hh_atual' value='{projeto["hh_atual"] or 0}'>
      </div>

      <button class='btn'>Salvar alterações</button>
      <a class='btn' href='/paint'>Voltar</a>
    </form>
    """

    return render_template_string(
        BASE.replace("{% block content %}{% endblock %}", html),
        user=session['user'], perfil=session['perfil']
    )

# -------------------------
# Projetos PAINT - list / add
# -------------------------
@app.route('/paint', methods=['GET', 'POST'])
def paint():
    if 'user' not in session:
        return redirect('/')
    if session['perfil'] != 'admin':
        return 'Acesso negado'

    con = get_db()
    cur = con.cursor()

    # ---------------- SALVAR NOVO PROJETO ----------------
    if request.method == 'POST':
        classificacao = request.form.get('classificacao')
        item = request.form.get('item_paint')
        tipo = request.form.get('tipo_atividade')
        objeto = request.form.get('objeto')
        objetivo = request.form.get('objetivo')
        dt_ini = request.form.get('dt_ini') or None
        dt_fim = request.form.get('dt_fim') or None

        try:
            cur.execute("""
                INSERT INTO projeto_paint
                (classificacao, item_paint, tipo_atividade, objeto, objetivo_geral, dt_ini, dt_fim)
                VALUES (%s,%s,%s,%s,%s,%s,%s)
            """, (classificacao, item, tipo, objeto, objetivo, dt_ini, dt_fim))
            con.commit()
        except Exception:
            con.rollback()

    # ---------------- BUSCAR PROJETOS ----------------
    cur.execute("""
        SELECT *
        FROM projeto_paint
        ORDER BY item_paint
    """)
    projetos = cur.fetchall()

    # ---------------- SOMATÓRIO DE HORAS (UMA QUERY) ----------------
    cur.execute("""
        SELECT
            item_paint,
            SUM(
                (SPLIT_PART(duracao, ':', 1)::int * 60) +
                 SPLIT_PART(duracao, ':', 2)::int
            ) AS minutos
        FROM horas
        GROUP BY item_paint
    """)
    horas_rows = cur.fetchall()

    # mapa item_paint -> minutos
    mapa_horas = {
        r["item_paint"]: r["minutos"] or 0
        for r in horas_rows
    }

    con.close()

    # ---------------- HTML ----------------
    html = "<h3>Cadastrar Projeto PAINT</h3>"
    html += """
    <form method='post'>
      <div>Classificação:
        <select name='classificacao'>
          <option value="Obrigatório">Obrigatório</option>
          <option value="Prioritário">Prioritário</option>
          <option value="Complementar">Complementar</option>
          <option value="Novo">Novo</option>
        </select>
      </div>
      <div>Item PAINT (ex: O-1): <input name='item_paint' required></div>
      <div>Tipo de Atividade: <input name='tipo_atividade'></div>
      <div>Objeto: <input name='objeto'></div>
      <div>Objetivo Geral: <input name='objetivo'></div>
      <div>Data Inicial: <input type='date' name='dt_ini'></div>
      <div>Data Final: <input type='date' name='dt_fim'></div>
      <div><button class='btn'>Adicionar</button></div>
    </form>
    """

    # ---------------- IMPORTAÇÃO EM LOTE ----------------
    html += """
    <h3>Importar múltiplos Projetos PAINT</h3>
    <form method='post' action='/paint/import'>
        <p>Cole os dados abaixo (uma linha por projeto, separando colunas por TAB ou ;):</p>
        <textarea name='bulk_data' rows='15' style='width:100%'></textarea>
        <div><button class='btn'>Importar Projetos</button></div>
    </form>
    """

    # ---------------- LISTAGEM ----------------
    html += "<h4>Projetos cadastrados</h4>"
    html += """
    <input type="text" id="searchPaint" placeholder="Pesquisar projetos..."
           style="width:100%; padding:6px; margin:8px 0;">

    <script>
    document.getElementById("searchPaint").addEventListener("keyup", function() {
        let filter = this.value.toLowerCase();
        let rows = document.querySelectorAll("#tabelaPaint tbody tr");
        rows.forEach(row => {
            row.style.display = row.innerText.toLowerCase().includes(filter) ? "" : "none";
        });
    });
    </script>
    """

    html += """
    <div style="margin-bottom:10px;">
        <a class='btn btn-danger' href='/projeto/delete_all'
           onclick="return confirm('Deseja realmente excluir TODOS os Projetos PAINT?');">
           Excluir todos
        </a>
    </div>
    """

    html += """
    <table id="tabelaPaint">
      <tr>
        <th>Item</th>
        <th>Classif.</th>
        <th>Tipo</th>
        <th>Dt. Início</th>
        <th>Dt. Fim</th>
        <th>HH Atual</th>
        <th>HH Executada</th>
        <th>% Executado</th>
        <th>Ações</th>
      </tr>
    """

    for p in projetos:
        minutos = mapa_horas.get(p["item_paint"], 0)

        hh = minutos // 60
        mm = minutos % 60
        soma = f"{hh:02d}:{mm:02d}"

        if p["hh_atual"] and p["hh_atual"] > 0:
            total_prev = p["hh_atual"] * 60
            percentual = (minutos / total_prev) * 100
            percentual_fmt = f"{percentual:.2f}%"
        else:
            percentual_fmt = "0%"

        html += f"""
        <tr>
          <td>{p['item_paint']}</td>
          <td>{p['classificacao']}</td>
          <td>{p['tipo_atividade'] or ''}</td>
          <td>{fmt(p['dt_ini'])}</td>
          <td>{fmt(p['dt_fim'])}</td>
          <td>{p['hh_atual']}</td>
          <td>{soma}</td>
          <td>{percentual_fmt}</td>
          <td>
            <a class='btn' href='/projeto/edit/{p["id"]}'>Editar</a>
            <a class='btn btn-danger'
               href='/projeto/delete/{p["id"]}'
               onclick="return confirm('Excluir este projeto?');">
               Excluir
            </a>
          </td>
        </tr>
        """

    html += "</table>"

    return render_template_string(
        BASE.replace("{% block content %}{% endblock %}", html),
        user=session['user'],
        perfil=session['perfil']
    )

from datetime import datetime

from datetime import datetime
from psycopg2 import IntegrityError

@app.route('/paint/import', methods=['POST'])
def paint_import():
    if 'user' not in session:
        return redirect('/')
    if session['perfil'] != 'admin':
        return 'Acesso negado'

    data = request.form.get('bulk_data')
    if not data:
        return "Nenhum dado fornecido"

    con = get_db()
    cur = con.cursor()

    linhas = data.strip().splitlines()
    inseridos = 0
    ignorados = 0

    def conv_data(d):
        d = d.strip()
        if not d or d == "***":
            return None
        try:
            return datetime.strptime(d, "%d/%m/%Y").date()
        except:
            return None

    for i, linha in enumerate(linhas):
        linha = linha.strip()
        if not linha:
            continue

        # ignora cabeçalho
        if i == 0 and linha.lower().startswith("classificação"):
            continue

        # Google Sheets → TAB
        cols = linha.split('\t')

        if len(cols) < 6:
            ignorados += 1
            continue

        classificacao = cols[0].strip()
        item_paint   = cols[1].strip()
        tipo         = cols[2].strip()
        objeto       = cols[3].strip()
        objetivo     = cols[4].strip()
        dt_ini       = conv_data(cols[5])

        dt_fim = conv_data(cols[6]) if len(cols) > 6 else None

        # HH atual
        hh_atual = 0
        if len(cols) > 7 and cols[7].strip():
            try:
                hh_atual = int(float(cols[7].replace(",", ".")))
            except:
                hh_atual = 0

        if not item_paint:
            ignorados += 1
            continue

        try:
            cur.execute("""
                INSERT INTO projeto_paint
                (classificacao, item_paint, tipo_atividade, objeto,
                 objetivo_geral, dt_ini, dt_fim, hh_atual)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                classificacao,
                item_paint,
                tipo,
                objeto,
                objetivo,
                dt_ini,
                dt_fim,
                hh_atual
            ))
            inseridos += 1

        except IntegrityError:
            con.rollback()
            ignorados += 1
            continue

    con.commit()
    con.close()

    return f"""
    <h3>Importação concluída</h3>
    <p>✅ Inseridos: <b>{inseridos}</b></p>
    <p>⚠️ Ignorados (duplicados ou inválidos): <b>{ignorados}</b></p>
    <a href="/paint">Voltar</a>
    """

@app.route('/projeto/delete/<int:id>')
def projeto_delete(id):
    if 'user' not in session:
        return redirect('/')
    if session['perfil'] != 'admin':
        return 'Acesso negado'

    con = get_db()
    cur = con.cursor()
    cur.execute("DELETE FROM projeto_paint WHERE id=%s", (id,))
    con.commit()
    con.close()
    return redirect('/paint')


@app.route('/projeto/delete_all')
def delete_all_projetos():
    if 'user' not in session:
        return redirect('/')
    if session['perfil'] != 'admin':
        return 'Acesso negado'

    con = get_db()
    cur = con.cursor()
    cur.execute("DELETE FROM projeto_paint")
    con.commit()
    con.close()
    return redirect('/paint')

# -------------------------
# OS - list / add
# -------------------------
@app.route('/os', methods=['GET', 'POST'])
def os_page():
    if 'user' not in session:
        return redirect('/')
    if session['perfil'] != 'admin':
        return 'Acesso negado'

    con = get_db()
    cur = con.cursor()

    # Carrega colaboradores
    cur.execute('SELECT nome FROM colaboradores ORDER BY nome')
    colabs = [r['nome'] for r in cur.fetchall()]

    if request.method == 'POST':
        codigo = request.form.get('codigo')
        item = request.form.get('item_paint')
        resumo = request.form.get('resumo')
        unidade = ", ".join(request.form.getlist("unidade"))
        supervisao = ", ".join(request.form.getlist('supervisao'))
        coordenacao = ", ".join(request.form.getlist('coordenacao'))
        equipe = ", ".join(request.form.getlist('equipe'))
        observacao = request.form.get('observacao')
        status = request.form.get('status')

        plan = 1 if request.form.get('plan') == 'on' else 0
        exec_ = 1 if request.form.get('exec') == 'on' else 0
        rp = 1 if request.form.get('rp') == 'on' else 0
        rf = 1 if request.form.get('rf') == 'on' else 0

        dt_inicio = request.form.get("dt_inicio") or None
        dt_previsao_fim = request.form.get("dt_previsao_fim") or None
        dt_conclusao = request.form.get("dt_conclusao") or None

        try:
            cur.execute(
                '''INSERT INTO os
                (codigo,item_paint,resumo,unidade,supervisao,coordenacao,equipe,
                 observacao,status,plan,exec,rp,rf,dt_inicio,dt_previsao_fim,dt_conclusao)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)''',
                (codigo, item, resumo, unidade, supervisao, coordenacao, equipe,
                 observacao, status, plan, exec_, rp, rf,
                 dt_inicio, dt_previsao_fim, dt_conclusao)
            )
            con.commit()
        except IntegrityError:
            con.rollback()

    # Lista OS
    cur.execute('SELECT * FROM os ORDER BY codigo')
    rows = cur.fetchall()

    # Lista itens PAINT
    cur.execute('SELECT item_paint FROM projeto_paint ORDER BY item_paint')
    items = [r['item_paint'] for r in cur.fetchall()]

    con.close()

    html = '<h3>Cadastrar O.S</h3>'
    html += "<form method='post'>"
    html += "<div>Código (ex: OS-001): <input name='codigo' required></div>"

    html += "<div>Item PAINT: <select name='item_paint'>"
    for it in items:
        html += f"<option value='{it}'>{it}</option>"
    html += "</select></div>"

    html += "<div>Resumo: <input name='resumo' style='width:300px'></div>"

    html += """
    <div>
        Unidade:
        <select name='unidade' multiple size='5'>
            <option value='DAC'>DAC</option>
            <option value='DIRAE'>DIRAE</option>
            <option value='DMAD'>DMAD</option>
            <option value='DOSE'>DOSE</option>
            <option value='DACGR'>DACGR</option>
        </select>
    </div>
    """

    html += "<div>Supervisão:<br><select name='supervisao' multiple size='6' style='width:260px'>"
    for c in colabs:
        html += f"<option value='{c}'>{c}</option>"
    html += "</select></div>"

    html += "<div>Coordenação:<br><select name='coordenacao' multiple size='6' style='width:260px'>"
    for c in colabs:
        html += f"<option value='{c}'>{c}</option>"
    html += "</select></div>"

    html += "<div>Equipe:<br><select name='equipe' multiple size='7' style='width:260px'>"
    for c in colabs:
        html += f"<option value='{c}'>{c}</option>"
    html += "</select></div>"

    html += "<div>Observação: <input name='observacao'></div>"
    html += "<div>Status: <select name='status'><option>Andamento</option><option>RP-Syria</option><option>RP-MariaCristina</option><option>Concluido</option></select></div>"

    html += "<div>Flags: \
        <label><input type='checkbox' name='plan'> PLAN</label> \
        <label><input type='checkbox' name='exec'> EXEC</label> \
        <label><input type='checkbox' name='rp'> RP</label> \
        <label><input type='checkbox' name='rf'> RF</label></div>"

    html += "<div>Data início: <input type='date' name='dt_inicio'></div>"
    html += "<div>Previsão fim: <input type='date' name='dt_previsao_fim'></div>"
    html += "<div>Data conclusão: <input type='date' name='dt_conclusao'></div>"

    html += "<div><button class='btn'>Adicionar OS</button></div>"
    html += "</form>"

    # Pesquisa
    html += """
    <div style="margin: 15px 0;">
        <input type="text" id="searchInput" placeholder="Pesquisar..."
               style="padding:8px;width:100%;font-size:16px;">
    </div>

    <script>
    document.addEventListener("DOMContentLoaded", function() {
        const input = document.getElementById("searchInput");

        input.addEventListener("keyup", function() {
            let filter = input.value.toLowerCase();
            let rows = document.querySelectorAll("#tabelaOS tbody tr");

            rows.forEach(row => {
                let text = row.innerText.toLowerCase();
                row.style.display = text.includes(filter) ? "" : "none";
            });
        });
    });
    </script>
    """

    html += """
    <h4>O.S cadastradas</h4>
    
    <div style="margin-bottom:15px; display:flex; gap:10px;">
        <a class='btn' href='/os/import'
           style="background:#2563eb; color:white; padding:6px 12px; border-radius:6px; text-decoration:none;">
           ⬆ Importar
        </a>
    
        <a class='btn btn-danger' href='/os/delete_all'
           style="background:#dc2626; color:white; padding:6px 12px; border-radius:6px; text-decoration:none;"
           onclick="return confirm('⚠ Deseja realmente EXCLUIR TODAS as O.S? Essa ação não pode ser desfeita!');">
           🗑 Excluir Todas
        </a>
    </div>
    """

    html += """
    <table id="tabelaOS">
    <thead>
    <tr>
        <th>Código</th>
        <th>Item PAINT</th>
        <th>Resumo</th>
        <th>Status</th>
        <th>Início</th>
        <th>Prev. Fim</th>
        <th>Prazo</th>
        <th>Restante</th>
        <th>PLAN</th>
        <th>EXEC</th>
        <th>RP</th>
        <th>RF</th>
        <th>Ações</th>
    </tr>
    </thead>
    <tbody>
    """

    for r in rows:
        prazo, restante = calcular_prazo(r['dt_inicio'], r['dt_previsao_fim'],r['dt_conclusao'])
        html += f"""
        <tr>
            <td>{r['codigo']}</td>
            <td>{r['item_paint']}</td>
            <td>{r['resumo']}</td>
            <td>{r['status']}</td>
            <td>{fmt(r['dt_inicio'])}</td>
            <td>{fmt(r['dt_previsao_fim'])}</td>
            <td>{prazo}</td>
            <td>{restante}</td>
            <td>{"<span style='color:#2563eb;font-size:18px;'>●</span>" if r['plan'] else "<span style='color:#dc2626;font-size:18px;'>●</span>"}</td>
            <td>{"<span style='color:#2563eb;font-size:18px;'>●</span>" if r['exec'] else "<span style='color:#dc2626;font-size:18px;'>●</span>"}</td>
            <td>{"<span style='color:#2563eb;font-size:18px;'>●</span>" if r['rp'] else "<span style='color:#dc2626;font-size:18px;'>●</span>"}</td>
            <td>{"<span style='color:#2563eb;font-size:18px;'>●</span>" if r['rf'] else "<span style='color:#dc2626;font-size:18px;'>●</span>"}</td>
            <td>
                <a class='btn' href='/os/view/{r["id"]}'>Ver</a>
                <a class='btn' href='/os/edit/{r["id"]}'>Editar</a>
                <a class='btn btn-danger' href='/os/delete/{r["id"]}'
                   onclick="return confirm('Deseja realmente excluir esta O.S?');">
                   Excluir
                </a>
            </td>
        </tr>
        """

    html += "</tbody></table>"

    return render_template_string(
        BASE.replace('{% block content %}{% endblock %}', html),
        user=session['user'],
        perfil=session['perfil'],
        fmt=fmt
    )

@app.route('/os/gestao')
def os_gestao():
    if 'user' not in session:
        return redirect('/')
    if session['perfil'] != 'admin':
        return 'Acesso negado'

    con = get_db()
    cur = con.cursor()

    cur.execute("SELECT * FROM os ORDER BY codigo")
    rows = cur.fetchall()

    con.close()

    def fmt_data(d):
        if not d:
            return '-'
        if isinstance(d, str):
            try:
                d = datetime.fromisoformat(d)
            except:
                return d
        return d.strftime('%d/%m/%Y')

    def pct(v):
        v = int(v or 0)
        if v == 100:
            return f"<span class='pct-ok'>🏁 {v}%</span>"
        return f"{v}%"

    html = """
    <style>
    .titulo-page {
        text-align:center;
        font-size:28px;
        font-weight:bold;
        margin:20px 0;
        color:#2c5aa0;
    }

    .filtro {
        margin-bottom:20px;
    }

    .filtro input {
        width:100%;
        padding:10px;
        font-size:16px;
        border-radius:8px;
        border:1px solid #ccc;
    }

    .linha {
        display:grid;
        grid-template-columns: 1fr 2fr 1fr 2fr 1fr 0.7fr 0.7fr 0.7fr 0.7fr 0.5fr 0.5fr;
        gap:10px;
        align-items:center;
        border:2px solid #2c5aa0;
        border-radius:12px;
        padding:10px;
        margin-bottom:10px;
        background:#eef3fb;
        font-size:14px;
    }

    .header {
        font-weight:bold;
        background:#dbe7ff;
    }

    .cell {
        border-right:2px solid #2c5aa0;
        padding-right:8px;
    }

    .cell:last-child {
        border-right:none;
    }

    .icon-btn {
        display:flex;
        justify-content:center;
        align-items:center;
        width:34px;
        height:34px;
        border-radius:50%;
        text-decoration:none;
        font-size:16px;
        font-weight:bold;
        transition:0.2s;
    }

    .icon-view {
        background:#dbeafe;
        color:#1d4ed8;
    }
    
    .icon-view:hover {
        background:#bfdbfe;
    }
    
    .icon-edit {
        background:#fef3c7;
        color:#b45309;
    }
    
    .icon-edit:hover {
        background:#fde68a;
    }

    .pct-ok {
        font-weight: bold;
        color: #15803d;
        background: #dcfce7;
        padding: 2px 6px;
        border-radius: 6px;
    }
    </style>

    <div class="titulo-page">
        Controle das Ordens de Serviço
    </div>

    <div class="filtro">
        <input type="text" id="searchInput" placeholder="Filtrar...">
    </div>

    <div class="linha header">
        <div class="cell">OS</div>
        <div class="cell">Descrição</div>
        <div class="cell">Diretoria</div>
        <div class="cell">Equipe</div>
        <div class="cell">Prazo</div>
        <div class="cell">Planej.</div>
        <div class="cell">Exec.</div>
        <div class="cell">RP</div>
        <div class="cell">RF</div>
        <div>Ver</div>
        <div>Editar</div>
    </div>
    """

    for r in rows:
        html += f"""
        <div class="linha row">
            <div class="cell">{r['codigo']}</div>
            <div class="cell">{r['resumo'] or ''}</div>
            <div class="cell">{r['unidade'] or ''}</div>
            <div class="cell">{r['equipe'] or ''}</div>
            <div class="cell">{fmt_data(r['dt_previsao_fim'])}</div>
            <div class="cell">{pct(r.get('plan0100'))}</div>
            <div class="cell">{pct(r.get('exec0100'))}</div>
            <div class="cell">{pct(r.get('rp0100'))}</div>
            <div class="cell">{pct(r.get('rf0100'))}</div>
            <a href="/os/view/{r['id']}" class="icon-btn icon-view">ℹ</a>
            <a href="/os/edit/{r['id']}" class="icon-btn icon-edit">✏</a>
        </div>
        """

    html += """
    <script>
    document.getElementById("searchInput").addEventListener("keyup", function() {
        let filter = this.value.toLowerCase();
        let rows = document.querySelectorAll(".row");

        rows.forEach(row => {
            let text = row.innerText.toLowerCase();
            row.style.display = text.includes(filter) ? "" : "none";
        });
    });
    </script>
    """

    return render_template_string(
        BASE.replace('{% block content %}{% endblock %}', html),
        user=session['user'],
        perfil=session['perfil']
    )

@app.route('/os/rh')
def os_rh():
    if 'user' not in session:
        return redirect('/')
    if session['perfil'] != 'admin':
        return 'Acesso negado'
        
    def fmt_horas(mins):
        if not mins:
            return "00:00"
        h = mins // 60
        m = mins % 60
        return f"{h:02d}:{m:02d}"

    con = get_db()
    cur = con.cursor()

    # JOIN entre OS e status por usuário
    cur.execute("""
        SELECT 
            c.nome AS colaborador,
            o.codigo AS os_codigo,
            o.resumo,
            s.status,
            s.observacao,
            COALESCE(SUM(h.duracao_minutos),0) AS total_min
        FROM os o
        
        JOIN colaboradores c ON 
            o.equipe ILIKE '%' || c.nome || '%' OR
            o.coordenacao ILIKE '%' || c.nome || '%' OR
            o.supervisao ILIKE '%' || c.nome || '%'
        
        LEFT JOIN os_status_user s
            ON s.os_codigo = o.codigo
            AND s.colaborador = c.nome
        
        LEFT JOIN horas h
            ON h.os_codigo = o.codigo
            AND h.colaborador_id = c.id
        
        GROUP BY c.nome, o.codigo, o.resumo, s.status, s.observacao
        ORDER BY o.codigo, c.nome
    """)

    rows = cur.fetchall()
    con.close()

    html = """
    <style>
    .titulo-page {
        text-align:center;
        font-size:28px;
        font-weight:bold;
        margin:20px 0;
        color:#2c5aa0;
    }

    .filtro {
        margin-bottom:20px;
    }

    .filtro input {
        width:100%;
        padding:10px;
        font-size:16px;
        border-radius:8px;
        border:1px solid #ccc;
    }

    .linha {
        display:grid;
        grid-template-columns: 1.2fr 1fr 2fr 1.2fr 3fr 1fr;
        gap:10px;
        align-items:center;
        border:2px solid #2c5aa0;
        border-radius:12px;
        padding:10px;
        margin-bottom:10px;
        background:#eef3fb;
        font-size:14px;
    }

    .header {
        font-weight:bold;
        background:#dbe7ff;
    }

    .cell {
        border-right:2px solid #2c5aa0;
        padding-right:8px;
    }

    .cell:last-child {
        border-right:none;
    }

    .badge {
        padding:4px 8px;
        border-radius:8px;
        font-size:12px;
        color:white;
        font-weight:bold;
        display:inline-block;
    }

    .st-nao { background:#7f8c8d; }
    .st-and { background:#3498db; }
    .st-pausado { background:#e67e22; }
    .st-aguard { background:#9b59b6; }
    .st-ok { background:#27ae60; }
    </style>

    <div class="titulo-page">
        Controle de Pessoal das Ordens de Serviço
    </div>

    <div class="filtro">
        <input type="text" id="searchInput" placeholder="Filtrar...">
    </div>

    <div class="linha header">
        <div class="cell">Servidor</div>
        <div class="cell">OS</div>
        <div class="cell">Descrição</div>
        <div class="cell">Status</div>
        <div>Observação</div>
        <div>Total Horas</div>
    </div>
    """

    def badge_class(s):
        return {
            "Não Iniciado": "st-nao",
            "Em Andamento": "st-and",
            "Pausado": "st-pausado",
            "Aguardando Servidor": "st-aguard",
            "Concluído": "st-ok"
        }.get(s, "")

    for r in rows:
        html += f"""
        <div class="linha row">
            <div class="cell">{r['colaborador']}</div>
            <div class="cell">{r['os_codigo']}</div>
            <div class="cell">{r['resumo'] or ''}</div>
            <div class="cell">
                <span class="badge {badge_class(r['status'])}">
                    {r['status'] or '-'}
                </span>
            </div>
            <div>{r['observacao'] or ''}</div>
            <div><b>{fmt_horas(r['total_min'])}</b></div>
        </div>
        """

    html += """
    <script>
    document.getElementById("searchInput").addEventListener("keyup", function() {
        let filter = this.value.toLowerCase();
        let rows = document.querySelectorAll(".row");

        rows.forEach(row => {
            let text = row.innerText.toLowerCase();
            row.style.display = text.includes(filter) ? "" : "none";
        });
    });
    </script>
    """

    return render_template_string(
        BASE.replace('{% block content %}{% endblock %}', html),
        user=session['user'],
        perfil=session['perfil']
    )

@app.route('/os/delete/<int:id>')
def os_delete(id):
    if 'user' not in session:
        return redirect('/')
    if session['perfil'] != 'admin':
        return 'Acesso negado'

    con = get_db()
    cur = con.cursor()
    cur.execute("DELETE FROM os WHERE id=%s", (id,))
    con.commit()
    con.close()
    return redirect('/os')

@app.route('/os/delete_all')
def os_delete_all():
    if 'user' not in session:
        return redirect('/')
    if session['perfil'] != 'admin':
        return 'Acesso negado'

    con = get_db()
    cur = con.cursor()
    cur.execute("DELETE FROM os")  # apaga todos os registros
    con.commit()
    con.close()
    return redirect('/os')


from datetime import datetime
from flask import request

@app.route('/os/view/<int:id>')
def os_view(id):
    import html as html_lib
    if 'user' not in session:
        return redirect('/')
        
    user = session['user']
    perfil = session['perfil']

    con = get_db()
    try:
        cur = con.cursor()

        # verifica se é participante da OS
        cur.execute("""
            SELECT 1
            FROM os
            WHERE id = %s
              AND (
                equipe ILIKE %s OR
                coordenacao ILIKE %s OR
                supervisao ILIKE %s
              )
        """, (id, f"%{user}%", f"%{user}%", f"%{user}%"))
        
        participa = cur.fetchone()
        
        # regra de acesso
        if perfil != 'admin' and not participa:
            con.close()
            return 'Acesso negado'
    
        # OS
        cur.execute("SELECT * FROM os WHERE id=%s", (id,))
        os = cur.fetchone()
    
        if not os:
            con.close()
            return "O.S não encontrada"
    
        # ---------------- FILTROS ----------------
        modo = request.args.get('modo', 'recentes')  # recentes | todos | mes
        mes = request.args.get('mes')
        ano = datetime.now().year
    
        # ---------------- QUERY DINÂMICA ----------------
        sql = """
            SELECT h.data, h.duracao_minutos, h.atividade, h.observacoes, c.nome as colaborador
            FROM horas h
            JOIN colaboradores c ON c.id = h.colaborador_id
            WHERE h.os_codigo = %s
            AND (
                h.observacoes IS NULL
                OR h.observacoes NOT ILIKE 'Lançamento automático%%'
            )
        """
        params = [os['codigo']]
    
        if modo == 'mes' and mes:
            sql += " AND EXTRACT(MONTH FROM h.data) = %s AND EXTRACT(YEAR FROM h.data) = %s"
            params += [mes, ano]
    
        sql += " ORDER BY h.data DESC, h.id DESC"
    
        if modo == 'todos':
            sql += " LIMIT 1000"  # proteção
        elif modo == 'recentes':
            sql += " LIMIT 100"
    
        cur.execute(sql, tuple(params))
        horas = cur.fetchall()
    
        # HORAS POR COLABORADOR
        cur.execute("""
            SELECT 
                c.nome as colaborador,
                SUM(h.duracao_minutos) as minutos
            FROM horas h
            JOIN colaboradores c ON c.id = h.colaborador_id
            WHERE h.os_codigo = %s
            GROUP BY c.nome
            ORDER BY minutos DESC
        """, (os['codigo'],))
        
        horas_colab = cur.fetchall()
        total_min = sum([hc['minutos'] or 0 for hc in horas_colab])
    
        cur.execute("""
            SELECT status,
                   COUNT(*) as qtd,
                   STRING_AGG(colaborador, ', ') as nomes
            FROM os_status_user
            WHERE os_codigo = %s
            GROUP BY status
        """, (os['codigo'],))
        
        status_data = cur.fetchall()
        labels = [s['status'] or 'Sem status' for s in status_data]
        values = [s['qtd'] for s in status_data]
        nomes = [s['nomes'] for s in status_data]
    finally:
        con.close()

    # ---------------- HELPERS ----------------
    def fmt_data(d):
        return d.strftime('%d/%m/%Y') if d else ''

    def fmt_horas(mins):
        if not mins:
            return "00:00"
        return f"{mins//60:02d}:{mins%60:02d}"

    def tipo_label(t):
        if not t:
            return ''
        return t.replace("1. ", "").replace("2. ", "").replace("3. ", "").replace("4. ", "")

    def active(cond):
        return "style='background:#2c5aa0;color:white;'" if cond else ""

    # ---------------- HTML ----------------
    html = f"""
    <style>
    .card {{
        border:2px solid #2c5aa0;
        border-radius:15px;
        padding:20px;
        margin-bottom:20px;
        background:#eef3fb;
    }}
    .titulo {{
        font-size:22px;
        font-weight:bold;
        display:flex;
        align-items:center;
        gap:10px;
        margin-bottom:10px;
        color:#2c5aa0;
    }}
    table {{
        width:100%;
        border-collapse:collapse;
        margin-top:10px;
    }}
    th {{
        background:#2c5aa0;
        color:white;
        padding:8px;
        text-align:left;
    }}
    td {{
        padding:8px;
        border-bottom:1px solid #ccc;
    }}
    .filtros {{
        display:flex;
        gap:8px;
        flex-wrap:wrap;
        margin-bottom:10px;
    }}
    .btn {{
        padding:5px 10px;
        border:1px solid #2c5aa0;
        border-radius:6px;
        text-decoration:none;
        font-weight:bold;
        color:#2c5aa0;
    }}
    </style>

    <div class="titulo">
        📄 Ordem de Serviço nº {os['codigo']}
    </div>

    <div class="card">
        <div class="titulo">🕒 Histórico de Lançamentos</div>

        <div class="filtros">
            <a class="btn" {active(modo=='recentes')} href="/os/view/{os['id']}?modo=recentes">Recentes</a>
            <a class="btn" {active(modo=='todos')} href="/os/view/{os['id']}?modo=todos">Todos</a>
    """

    meses = [
        ("01","Jan"),("02","Fev"),("03","Mar"),("04","Abr"),
        ("05","Mai"),("06","Jun"),("07","Jul"),("08","Ago"),
        ("09","Set"),("10","Out"),("11","Nov"),("12","Dez")
    ]

    for m, nome in meses:
        html += f"""
        <a class="btn" {active(modo=='mes' and mes==m)}
           href="/os/view/{os['id']}?modo=mes&mes={m}">
           {nome}
        </a>
        """

    html += """
        </div>

        <table>
            <tr>
                <th>Servidor</th>
                <th>Data</th>
                <th>Duração</th>
                <th>Tipo</th>
                <th>Observação</th>
            </tr>
    """

    for h in horas:
        obs_raw = h["observacoes"] or ""
    
        obs_full = html_lib.escape(obs_raw)
    
        obs_short = obs_raw.strip()
        if len(obs_short) > 120:
            obs_short = obs_short[:120] + "..."
        obs_short = html_lib.escape(obs_short)

        html += f"""
        <tr>
            <td>{h['colaborador']}</td>
            <td>{fmt_data(h['data'])}</td>
            <td>{fmt_horas(h['duracao_minutos'])}</td>
            <td>{tipo_label(h['atividade'])}</td>
            <td title="{obs_full}">{obs_short}</td>
        </tr>
        """

    html += """
        </table>
    </div>
    """

    html += """
    <div class="card">
        <div class="titulo">👥 Horas por Colaborador</div>
    
        <table>
            <tr>
                <th>Servidor</th>
                <th>Horas</th>
            </tr>
    """

    for hc in horas_colab:
        html += f"""
        <tr>
            <td>{hc['colaborador']}</td>
            <td>{fmt_horas(hc['minutos'])}</td>
        </tr>
        """
    
    # TOTAL FORA DO LOOP
    html += f"""
    <tr style="font-weight:bold; background:#f0f0f0;">
        <td>TOTAL GERAL</td>
        <td>{fmt_horas(total_min)}</td>
    </tr>
    """
    html += """
        </table>
    </div>
    """
    if status_data:
        html += """
        <div class="card" style="max-width:500px; margin:auto;">
            <div class="titulo">📊 Status dos Colaboradores</div>
            <canvas id="graficoStatus" style="max-height:300px;"></canvas>
        </div>
        
        <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
        
        <script>
        const ctx = document.getElementById('graficoStatus');
        
        if (ctx) {{
            const data = {{
                labels: {labels},
                datasets: [{{
                    data: {values}
                }}]
            }};
        
            const nomes = {nomes};
        
            new Chart(ctx, {{
                type: 'pie',
                data: data,
                options: {{
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {{
                        tooltip: {{
                            callbacks: {{
                                label: function(context) {{
                                    let i = context.dataIndex;
                                    let total = context.dataset.data.reduce((a,b)=>a+b,0);
                                    let val = context.dataset.data[i];
                                    let perc = ((val/total)*100).toFixed(1);
    
                                    return context.label + ": " + val + " (" + perc + "%)\\n" + nomes[i];
                                }}
                            }}
                        }}
                    }}
                }}
            }});
        }}
        </script>
        """.format(
            labels=labels,
            values=values,
            nomes=nomes
        )

    return render_template_string(
        BASE.replace('{% block content %}{% endblock %}', html),
        user=session['user'],
        perfil=session['perfil']
    )
    
@app.route('/os/edit/<int:id>', methods=['GET', 'POST'])
def os_edit(id):
    if 'user' not in session:
        return redirect('/')
    if session['perfil'] != 'admin':
        return 'Acesso negado'

    def to_int(v):
        try:
            return int(v)
        except:
            return 0

    def fmt_date_input(d):
        if not d:
            return ''
        if isinstance(d, str):
            return d[:10]  # já vem como 'YYYY-MM-DD' ou 'YYYY-MM-DD HH:MM:SS'
        return d.strftime('%Y-%m-%d')
    
    def fmt_horas(mins):
        return f"{mins//60:02d}:{mins%60:02d}"

    def checked(v):
        return 'checked' if str(v) == '1' else ''

    con = get_db()
    cur = con.cursor()

    cur.execute("SELECT * FROM os WHERE id=%s", (id,))
    os = cur.fetchone()

    if not os:
        con.close()
        return "O.S não encontrada"

    codigo_antigo = os["codigo"]

    # carregar dados auxiliares (igual ao seu código)
    cur.execute('SELECT nome FROM colaboradores ORDER BY nome')
    colabs = [r['nome'] for r in cur.fetchall()]

    cur.execute('SELECT item_paint FROM projeto_paint ORDER BY item_paint')
    items = [r['item_paint'] for r in cur.fetchall()]

    cur.execute("""
        SELECT
            atividade,
            SUM(duracao_minutos) AS minutos
        FROM horas
        WHERE os_codigo = %s
        GROUP BY atividade
    """, (os["codigo"],))

    horas_map = {r["atividade"]: r["minutos"] or 0 for r in cur.fetchall()}
    
    horas_plan = horas_map.get("1. Planejamento", 0)
    horas_exec = horas_map.get("2. Execução", 0)
    horas_rp   = horas_map.get("3. Relatório", 0)
    horas_rf   = horas_map.get("4. Relatório Final", 0)
    
    total_horas = horas_plan + horas_exec + horas_rp + horas_rf

    h_plan = fmt_horas(horas_plan)
    h_exec = fmt_horas(horas_exec)
    h_rp   = fmt_horas(horas_rp)
    h_rf   = fmt_horas(horas_rf)
    h_total = fmt_horas(total_horas)
    
    media_percentual = int((
        (os.get("plan0100") or 0) +
        (os.get("exec0100") or 0) +
        (os.get("rp0100") or 0) +
        (os.get("rf0100") or 0)
    ) / 4)

    if request.method == 'POST':
        codigo_novo = request.form.get('codigo')
        item = request.form.get('item_paint')
        resumo = request.form.get('resumo') or ""
        unidade = ", ".join(request.form.getlist("unidade"))
        supervisao = ", ".join(request.form.getlist('supervisao'))
        coordenacao = ", ".join(request.form.getlist('coordenacao'))
        equipe = ", ".join(request.form.getlist('equipe'))
        observacao = request.form.get('observacao')
        status = request.form.get('status')
        plan = 1 if request.form.get('plan') else 0
        exec_ = 1 if request.form.get('exec') else 0
        rp = 1 if request.form.get('rp') else 0
        rf = 1 if request.form.get('rf') else 0
        dt_inicio = request.form.get("dt_inicio") or None
        dt_previsao_fim = request.form.get("dt_previsao_fim") or None
        dt_conc = request.form.get('dt_conclusao') or None
        keys = request.form.get("keys")
        uo = ", ".join(request.form.getlist("uo"))
        
        plan0100 = to_int(request.form.get("plan0100"))
        exec0100 = to_int(request.form.get("exec0100"))
        rp0100   = to_int(request.form.get("rp0100"))
        rf0100   = to_int(request.form.get("rf0100"))
        
        rp_dt_envio_ua  = request.form.get("rp_dt_envio_ua") or None
        rf_dt_envio_sup = request.form.get("rf_dt_envio_sup") or None
        rf_dt_envio_ua  = request.form.get("rf_dt_envio_ua") or None

        try:
            # ---- atualiza OS ----
            cur.execute("""
                UPDATE os SET 
                    codigo=%s, item_paint=%s, resumo=%s, unidade=%s, supervisao=%s, 
                    coordenacao=%s, equipe=%s, observacao=%s, status=%s, 
                    plan=%s, exec=%s, rp=%s, rf=%s,
                    dt_inicio=%s, dt_previsao_fim=%s, dt_conclusao=%s,
            
                    keys=%s, uo=%s,
                    plan0100=%s, exec0100=%s, rp0100=%s, rf0100=%s,
                    rp_dt_envio_ua=%s,
                    rf_dt_envio_sup=%s, rf_dt_envio_ua=%s
            
                WHERE id=%s
            """, (
                codigo_novo, item, resumo, unidade, supervisao,
                coordenacao, equipe, observacao, status,
                plan, exec_, rp, rf,
                dt_inicio, dt_previsao_fim, dt_conc,
            
                keys, uo,
                plan0100, exec0100, rp0100, rf0100,
                rp_dt_envio_ua,
                rf_dt_envio_sup, rf_dt_envio_ua,
            
                id
            ))

            # ---- CASCADE MANUAL ----
            if codigo_antigo != codigo_novo:
                cur.execute(
                    "UPDATE horas SET os_codigo=%s WHERE os_codigo=%s",
                    (codigo_novo, codigo_antigo)
                )
                cur.execute(
                    "UPDATE delegacoes SET os_codigo=%s WHERE os_codigo=%s",
                    (codigo_novo, codigo_antigo)
                )

            con.commit()

        except Exception as e:
            con.rollback()
            con.close()
            return f"Erro ao atualizar O.S: {e}"

        con.close()
        return redirect('/os/gestao')

    con.close()

    # ---------------- CARREGAR DADOS ATUAIS ----------------
    unidade_atual = (os['unidade'] or "").split(", ")
    supervisao_atual = (os['supervisao'] or "").split(", ")
    coordenacao_atual = (os['coordenacao'] or "").split(", ")
    equipe_atual = (os['equipe'] or "").split(", ")

    resumo_atual = os['resumo'] if 'resumo' in os.keys() else ""
    
    # variaveis data input
    dt_inicio_val = fmt_date_input(os.get('dt_inicio'))
    dt_prev_val   = fmt_date_input(os.get('dt_previsao_fim'))
    dt_conc_val   = fmt_date_input(os.get('dt_conclusao'))
    
    rp_ua  = fmt_date_input(os.get('rp_dt_envio_ua'))
    rf_sup = fmt_date_input(os.get('rf_dt_envio_sup'))
    rf_ua  = fmt_date_input(os.get('rf_dt_envio_ua'))

    # ---------------- FORM HTML ----------------
    html = f"""
    <style>
    .card {{
        border:1px solid #cbd5e1;
        border-radius:10px;
        padding:15px;
        margin-bottom:15px;
        background:#f8fafc;
    }}
    .titulo {{
        font-weight:bold;
        font-size:16px;
        margin-bottom:10px;
        display:flex;
        align-items:center;
        gap:8px;
    }}
    .grid {{
        display:grid;
        grid-template-columns: repeat(auto-fit, minmax(220px,1fr));
        gap:10px;
    }}
    .badge {{
        background:#e2e8f0;
        padding:5px 8px;
        border-radius:6px;
        font-weight:bold;
    }}
    .box_horas {{
        background:#1e293b;
        color:white;
        padding:6px 10px;
        border-radius:6px;
        font-weight:bold;
        display:inline-block;
    }}
    input, select {{
        padding:6px;
        border-radius:6px;
        border:1px solid #cbd5e1;
        width:100%;
    }}
    </style>
    
    <h2 style="display:flex;align-items:center;gap:10px;">
        📄 <span>Ordem de Serviço nº {os['codigo']}</span>
    </h2>
    
    <form method='post'>
    
    <div class="card">
    <div class="titulo">📌 Dados Gerais</div>
    
    <div class="grid">
        <div>Código:<br><input name='codigo' value='{os['codigo']}'></div>
    
        <div>Item PAINT:<br>
            <select name='item_paint'>
    """
    
    for it in items:
        sel = "selected" if it == os['item_paint'] else ""
        html += f"<option value='{it}' {sel}>{it}</option>"
    
    html += f"""
            </select>
        </div>
    
        <div>Resumo:<br>
            <input name='resumo' value='{resumo_atual}'>
        </div>
    
        <div>Status:<br>
            <select name='status'>
                <option {'selected' if os['status']=='Andamento' else ''}>Andamento</option>
                <option {'selected' if os['status']=='Concluido' else ''}>Concluido</option>
            </select>
        </div>
    </div>
    
    <br>
    
    <div class="grid">
        <div>Palavras-chave:<br>
            <input name='keys' value="{os.get('keys','') or ''}">
        </div>
    
        <div>Unidade:<br>
            <select name='unidade' multiple size='5'>
    """
    
    unidades_opcoes = ["DAC","DIRAE","DMAD","DOSE","DACGR"]
    for u in unidades_opcoes:
        sel = "selected" if u in unidade_atual else ""
        html += f"<option value='{u}' {sel}>{u}</option>"
    
    html += f"""
            </select>
        </div>
    
        <div>UO:<br>
            <select name='uo' multiple size='5'>
    """
    
    uos = ["CM","SEGOV","SMGAS","PGM","SMA","SMF","SME","SMCT","SMS","SEDES",
           "SMAGRO","SEINFRA","SETTRAN","DMAE","IPREMU","FUTEL","FERUB",
           "EMAM","CGM","SESURB","SMH","SEJUV","SECOM","SEDEI","SMGE",
           "SEPLAN","SSEG","ARESAN","EXTERNO","OUTROS"]
    
    uo_atual = (os.get("uo") or "").split(", ")
    
    for u in uos:
        sel = "selected" if u in uo_atual else ""
        html += f"<option value='{u}' {sel}>{u}</option>"
    
    html += f"""
            </select>
        </div>
    </div>
    </div>
    
    <div class="card">
    <div class="titulo">⚙️ Gestão do Projeto</div>
    
    <div class="grid">
        <div>Data Início:<br>
            <input type='date' name='dt_inicio' value='{dt_inicio_val or ""}'>
        </div>
    
        <div>Previsão Inicial:<br>
            <input type='date' name='rf_dt_envio_sup' value='{rf_sup or ""}'>
        </div>

        <div>Previsão Atualizada:<br>
            <input type='date' name='dt_previsao_fim' value='{dt_prev_val or ""}'>
        </div>
    
        <div>Conclusão:<br>
            <input type='date' name='dt_conclusao' value='{dt_conc_val or ""}'>
        </div>
    </div>
    
    <br>
    <div class="grid">

    <div>
        <b>Planejamento</b><br>
        <div style="display:flex;align-items:center;gap:8px;">
            <input name='plan0100' type='number' min='0' max='100'
                   value='{os.get('plan0100',0)}'
                   style='width:70px;'>
            <span>%</span>

            <label style="display:flex;align-items:center;gap:4px;">
                <input type='checkbox' name='plan' {checked(os.get('plan'))}>
                PLAN
            </label>
        </div>
        <div class="box_horas">{h_plan}</div>
    </div>

    <div>
        <b>Execução</b><br>
        <div style="display:flex;align-items:center;gap:8px;">
            <input name='exec0100' type='number' min='0' max='100'
                   value='{os.get('exec0100',0)}'
                   style='width:70px;'>
            <span>%</span>

            <label style="display:flex;align-items:center;gap:4px;">
                <input type='checkbox' name='exec' {checked(os.get('exec'))}>
                EXEC
            </label>
        </div>
        <div class="box_horas">{h_exec}</div>
    </div>

    <div>
        <b>Relatório (RP)</b><br>
        <div style="display:flex;align-items:center;gap:8px;">
            <input name='rp0100' type='number' min='0' max='100'
                   value='{os.get('rp0100',0)}'
                   style='width:70px;'>
            <span>%</span>

            <label style="display:flex;align-items:center;gap:4px;">
                <input type='checkbox' name='rp' {checked(os.get('rp'))}>
                RP
            </label>
        </div>
        <div class="box_horas">{h_rp}</div>
        Enviado para Unid. Audit:
        <input type='date' name='rp_dt_envio_ua' value='{rp_ua or ""}'>
    </div>

    <div>
        <b>Relatório Final (RF)</b><br>
        <div style="display:flex;align-items:center;gap:8px;">
            <input name='rf0100' type='number' min='0' max='100'
                   value='{os.get('rf0100',0)}'
                   style='width:70px;'>
            <span>%</span>

            <label style="display:flex;align-items:center;gap:4px;">
                <input type='checkbox' name='rf' {checked(os.get('rf'))}>
                RF
            </label>
        </div>
        <div class="box_horas">{h_rf}</div>
    </div>

</div>
    
    <br>
    
    <div class="grid">
        <div>
            <b>Total Geral</b><br>
            <span class="badge">{media_percentual}%</span>
        </div>
    
        <div>
            <b>Total Horas</b><br>
            <span class="badge">{h_total}</span>
        </div>
    </div>
    
    </div>
    
    <div class="card">
    <div class="titulo">👥 Equipe</div>
    
    <div class="grid">
    """
    
    # supervisão
    html += "<div>Supervisão:<br><select name='supervisao' multiple size='5'>"
    for c in colabs:
        sel = "selected" if c in supervisao_atual else ""
        html += f"<option {sel}>{c}</option>"
    html += "</select></div>"
    
    # coordenação
    html += "<div>Coordenação:<br><select name='coordenacao' multiple size='5'>"
    for c in colabs:
        sel = "selected" if c in coordenacao_atual else ""
        html += f"<option {sel}>{c}</option>"
    html += "</select></div>"
    
    # equipe
    html += "<div>Equipe:<br><select name='equipe' multiple size='6'>"
    for c in colabs:
        sel = "selected" if c in equipe_atual else ""
        html += f"<option {sel}>{c}</option>"
    html += "</select></div>"
    
    html += f"""
    </div>
    </div>
    
    <div class="card">
    <div class="titulo">📝 Observações</div>
    <input name='observacao' value='{os['observacao'] or ''}' style='width:100%'>
    </div>
    
    <div style="margin-top:15px;">
        <button class='btn btn-primary'>Salvar</button>
        <a class='btn' href='/os'>Voltar</a>
    </div>
    
    </form>
    """

    return render_template_string(BASE.replace('{% block content %}{% endblock %}', html),
                                  user=session['user'], perfil=session['perfil'])

# -------------------------
# Importar OS por colar texto
# -------------------------
@app.route('/os/import', methods=['GET', 'POST'])
def os_import():
    if 'user' not in session:
        return redirect('/')
    if session['perfil'] != 'admin':
        return "Acesso negado"

    msg = ""

    def conv_bool(v):
        return 1 if v.strip().upper() == "TRUE" else 0

    def conv_data(v):
        v = v.strip()
        if not v:
            return None
        try:
            return datetime.strptime(v, "%d/%m/%Y").date()
        except:
            return None

    if request.method == 'POST':
        texto = request.form.get("texto", "").strip()

        if texto:
            linhas = texto.splitlines()
            con = get_db()
            cur = con.cursor()
        
            inseridos = 0
            ignorados = 0
        
            for i, linha in enumerate(linhas):
                linha = linha.strip()
        
                if not linha:
                    continue
        
                # ignora cabeçalho
                if i == 0 and linha.lower().startswith("os"):
                    continue
        
                partes = linha.split("\t")
        
                if len(partes) < 10:
                    ignorados += 1
                    continue
        
                codigo       = partes[0].strip()
                item_paint   = partes[1].strip()
                resumo       = partes[2].strip()
                unidade      = partes[3].strip()
                dt_inicio    = conv_data(partes[4])
                dt_fim_prev  = conv_data(partes[5])
                supervisao   = partes[6].strip()
                coordenacao  = partes[7].strip()
                equipe       = partes[8].strip()
        
                plan  = conv_bool(partes[9])  if len(partes) > 9  else 0
                exec_ = conv_bool(partes[10]) if len(partes) > 10 else 0
                rp    = conv_bool(partes[11]) if len(partes) > 11 else 0
                rf    = conv_bool(partes[12]) if len(partes) > 12 else 0
        
                status        = partes[13].strip() if len(partes) > 13 else None
                dt_conclusao  = conv_data(partes[14]) if len(partes) > 14 else None
                observacao    = partes[15].strip() if len(partes) > 15 else None
        
                if not codigo or not item_paint:
                    ignorados += 1
                    continue
        
                try:
                    cur.execute("""
                        INSERT INTO os (
                            codigo,
                            item_paint,
                            resumo,
                            unidade,
                            dt_inicio,
                            dt_previsao_fim,
                            supervisao,
                            coordenacao,
                            equipe,
                            observacao,
                            plan,
                            exec,
                            rp,
                            rf,
                            status,
                            dt_conclusao
                        )
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """, (
                        codigo,
                        item_paint,
                        resumo,
                        unidade,
                        dt_inicio,
                        dt_fim_prev,
                        supervisao,
                        coordenacao,
                        equipe,
                        observacao,
                        plan,
                        exec_,
                        rp,
                        rf,
                        status,
                        dt_conclusao
                    ))
                    inseridos += 1
        
                except IntegrityError:
                    con.rollback()
                    ignorados += 1
        
            # 🔹 AGORA FORA DO FOR
            con.commit()
            con.close()
        
            msg = f"✅ {inseridos} O.S inseridas | ⚠️ {ignorados} ignoradas"

    return render_template_string(
        BASE.replace(
            "{% block content %}{% endblock %}",
            f"""
            <h3>Importar O.S (Copiar e colar do Google Sheets)</h3>

            <form method="post">
                <textarea name="texto" rows="18" style="width:100%;"></textarea><br><br>
                <button class="btn btn-primary">Importar</button>
            </form>

            <p><b>{msg}</b></p>

            <p>Formato esperado (TAB):</p>
            <pre style="background:#eef; padding:10px;">
OS | ITEM PAINT | ID | UNIDADE | COORDENAÇÃO | EQUIPE | OBS | PLAN | EXEC | RP | RF | STATUS | DT_CONCLUSÃO
            </pre>

            <a class="btn" href="/os">Voltar</a>
            """
        ),
        user=session['user'],
        perfil=session['perfil']
    )

# -------------------------
# Lançar horas (colaborador)
# -------------------------
@app.route('/lancar', methods=['GET', 'POST'])
def lancar():
    if 'user' not in session:
        return redirect('/')

    from datetime import datetime, date

    con = get_db()
    cur = con.cursor()

    # -------------------------
    # CARREGAR OS
    # -------------------------
    cur.execute("SELECT codigo, item_paint, resumo FROM os ORDER BY codigo")
    oss = cur.fetchall()
    os_pre = request.args.get("os")

    # -------------------------
    # CARREGAR REQUISIÇÕES DO COLABORADOR
    # -------------------------
    cur.execute("""
        SELECT id, chave, tipo, criterio
        FROM requisicoes
        WHERE servidor_id = %s
          AND status_analise = 'ANDAMENTO'
        ORDER BY chave
    """, (session["user_id"],))
    requisicoes = [dict(r) for r in cur.fetchall()]

    # -------------------------
    # COLABORADORES
    # -------------------------
    cur.execute("SELECT id, nome FROM colaboradores ORDER BY nome")
    colaboradores = cur.fetchall()

    # -------------------------
    # PROCESSAR POST
    # -------------------------
    if request.method == 'POST':
    
        item = request.form.get('item')
        os_codigo = request.form.get('os')
        atividade = request.form.get('atividade')
        observacoes = request.form.get('observacoes')
        coparticipantes = request.form.getlist("coparticipantes[]")
    
        requisicoes_ids = request.form.getlist("requisicoes[]")
    
        datas = request.form.getlist("data[]")
        duracoes = request.form.getlist("duracao[]")
    
        if not datas:
            con.close()
            return "Nenhum lançamento informado"
    
        # lista de colaboradores que receberão o lançamento
        destinatarios = [session["user_id"]] + [int(c) for c in coparticipantes]
        for data, duracao in zip(datas, duracoes):
            if not duracao:
                continue
            try:
                h, m = map(int, duracao.split(":"))
                if m >= 60:
                    raise ValueError
            except:
                continue
                
            minutos = h * 60 + m
        
            dt = datetime.strptime(data, "%Y-%m-%d")
            if dt.year != 2026:
                con.close()
                return "Só é permitido lançar horas em 2026"
        
            duracao_fmt = f"{minutos//60:02d}:{minutos%60:02d}"
        
            for colab_id in destinatarios:
            
                if colab_id == session["user_id"]:
                    obs_final = observacoes
                else:
                    obs_final = f"Lançamento automático da O.S {os_codigo} por {session['user']}"
            
                    if observacoes:
                        obs_final += f" | {observacoes}"
            
                cur.execute("""
                    INSERT INTO horas
                    (colaborador_id, data, item_paint, os_codigo,
                     atividade, hora_inicio, hora_fim,
                     duracao, duracao_minutos, observacoes)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    RETURNING id
                """, (
                    colab_id,
                    data,
                    item,
                    os_codigo,
                    atividade,
                    None,
                    None,
                    duracao_fmt,
                    minutos,
                    obs_final
                ))
                hora_id = cur.fetchone()["id"]
                # vincular requisições (para TODOS)
                for req_id in requisicoes_ids:
                    cur.execute("""
                        INSERT INTO horas_requisicoes (hora_id, requisicao_id)
                        VALUES (%s, %s)
                    """, (hora_id, req_id))
    
            # -------------------------
            # OS 1.15 – Atendimento
            # -------------------------
            if os_codigo == "1.15/2026":
    
                responsaveis_ids = request.form.getlist("responsaveis[]")
                os_resumo = next((o["resumo"] for o in oss if o["codigo"] == os_codigo), None)
    
                cur.execute("""
                    INSERT INTO atendimentos (
                        hora_id, colaborador_id, os_codigo, os_resumo,
                        responsaveis_consultoria, macro, diretoria, atividade,
                        data_consultoria, assunto, participantes_externos,
                        entidades, meio_contato, observacao,
                        duracao_minutos, data_lancamento
                    ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    hora_id,
                    session["user_id"],
                    os_codigo,
                    os_resumo,
                    ", ".join(
                        c["nome"] for c in colaboradores
                        if str(c["id"]) in responsaveis_ids
                    ),
                    request.form.get("macro"),
                    request.form.get("diretoria"),
                    request.form.get("atividade_atendimento"),
                    request.form.get("data_consultoria"),
                    request.form.get("assunto"),
                    request.form.get("participantes_externos"),
                    ", ".join(request.form.getlist("entidades[]")),
                    request.form.get("meio_contato"),
                    request.form.get("observacao_atendimento"),
                    minutos,
                    data
                ))
    
            # -------------------------
            # OS 1.14 / 1.16 – Consultoria
            # -------------------------
            elif os_codigo in ("1.14/2026", "1.16/2026"):
    
                tipo = "consultoria" if os_codigo == "1.14/2026" else "treinamento"
                responsaveis = request.form.getlist("responsaveis2[]")
                os_resumo = next((o["resumo"] for o in oss if o["codigo"] == os_codigo), None)
    
                cur.execute("""
                    INSERT INTO consultorias (
                        hora_id, colaborador_id, os_codigo, os_resumo,
                        responsaveis, tipo, data_consul, assunto,
                        secretarias, meio, palavras_chave,
                        num_oficio, observacao,
                        duracao_minutos, data_lancamento
                    ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    hora_id,
                    session["user_id"],
                    os_codigo,
                    os_resumo,
                    ", ".join(responsaveis),
                    tipo,
                    request.form.get("data_consul"),
                    request.form.get("assunto_consultoria"),
                    ", ".join(request.form.getlist("secretarias[]")),
                    request.form.get("meio"),
                    request.form.get("palavras_chave"),
                    request.form.get("num_oficio"),
                    request.form.get("observacao"),
                    minutos,
                    data
                ))
        con.commit()
        con.close()
        return redirect('/menu')

    # -------------------------
    # HTML
    # -------------------------
    data_padrao = date.today().isoformat()
    con.close()

    form_html = """
<h3>Lançar Horas</h3>

<form method="post">

    <div>O.S:
        <select name="os" id="os_select" required>
            <option value=""></option>
            {% for o in oss %}
                <option value="{{ o.codigo }}"
                        data-item="{{ o.item_paint }}"
                        {% if os_pre == o.codigo %}selected{% endif %}>
                    {{ o.codigo }}{% if o.resumo %} - {{ o.resumo }}{% endif %}
                </option>
            {% endfor %}
        </select>
    </div>

    <div>Item PAINT:
        <input type="text" id="item_paint" name="item" readonly>
    </div>

    <!-- REQUISIÇÕES -->
    <div id="box_requisicoes" style="display:none; border:1px solid #ccc; padding:10px; margin-top:10px;">
        <h4>Requisições Delegadas</h4>

        <input type="text" id="busca_req" placeholder="Pesquisar..."
               style="width:100%; margin-bottom:6px;">

        <div id="lista_reqs" style="
            max-height:200px;
            overflow:auto;
        ">
            <style>
                .req_item:hover {
                    background:#f0f0f0;
                }
                .req_item input:checked + span {
                    font-weight:bold;
                }
            </style>
            {% for r in requisicoes %}
                <label class="req_item" style="
                    display:flex;
                    align-items:center;
                    gap:8px;
                    padding:6px 8px;
                    cursor:pointer;
                    border-radius:4px;
                ">
                    <input type="checkbox" name="requisicoes[]" value="{{ r.id }}"
                           style="margin:0;">
                    <span>
                        <strong>{{ r.chave }}</strong>
                        &nbsp;| {{ r.tipo }} | {{ r.criterio }}
                    </span>
                </label>
            {% endfor %}
        </div>
    </div>

    <div>Atividade:
        <select name="atividade" required>
            <option>1. Planejamento</option>
            <option>2. Execução</option>
            <option>3. Relatório</option>
        </select>
    </div>

    <h4>Registros de Horas</h4>

    <div id="registros">
        <div class="registro">
            <input type="date" name="data[]" value="{{ data_padrao }}"
                   min="2026-01-01" max="2026-12-31" required>

            <input type="text" name="duracao[]" placeholder="HH:MM" required pattern="^\d{1,4}:\d{2}$">

            <button type="button" onclick="remover(this)">❌</button>
        </div>
    </div>

    <!-- ATENDIMENTO OS 1.15 -->
    <div id="box_atendimento" style="display:none; border:1px solid #ccc; padding:10px; margin-top:10px;">
        <h4>Dados do Atendimento (O.S 1.15)</h4>

        <label>Responsáveis</label><br>
        <select name="responsaveis[]" multiple size="5">
            {% for c in colaboradores %}
                <option value="{{ c.id }}">{{ c.nome }}</option>
            {% endfor %}
        </select>

        <div>Macro: <input name="macro"></div>
        <div>Diretoria: <input name="diretoria"></div>

        <div>Atividade:
            <select name="atividade_atendimento">
                <option></option>
                <option>Consulta</option>
                <option>Esclarecimento</option>
                <option>Orientação</option>
                <option>Preventiva</option>
            </select>
        </div>

        <div>Data: <input type="date" name="data_consultoria"></div>
        <div>Assunto: <input name="assunto"></div>

        <div>Participantes Externos:
            <textarea name="participantes_externos"></textarea>
        </div>

        <div>Entidades:
            <select name="entidades[]" multiple size="6">
                <option>CM</option><option>SEGOV</option><option>SMGAS</option>
                <option>PGM</option><option>SMA</option><option>SMF</option>
                <option>SME</option><option>SMCT</option><option>SMS</option>
                <option>SEDES</option><option>SMAGRO</option><option>SEINFRA</option>
                <option>SETTRAN</option><option>DMAE</option><option>FUTEL</option>
                <option>EMAM</option><option>FERUB</option><option>IPREMU</option>
                <option>SESURB</option><option>SMH</option><option>SEJUV</option>
                <option>SECOM</option><option>SEDEI</option><option>SMGE</option>
                <option>SSEG</option><option>ARESAN</option>
            </select>
        </div>

        <div>Meio:
            <select name="meio_contato">
                <option></option>
                <option>Presencial</option>
                <option>Email</option>
                <option>Telefone</option>
            </select>
        </div>

        <div>Observação:
            <textarea name="observacao_atendimento"></textarea>
        </div>
    </div>

    <!-- CONSULTORIA -->
    <div id="box_consultoria" style="display:none; border:1px solid #ccc; padding:10px; margin-top:10px;">
        <h4>Consultoria / Treinamento</h4>

        <label>Responsáveis</label><br>
        <select name="responsaveis2[]" multiple size="5">
            {% for c in colaboradores %}
                <option value="{{ c.id }}">{{ c.nome }}</option>
            {% endfor %}
        </select>

        <div>Assunto: <textarea name="assunto_consultoria"></textarea></div>
        <div>Meio: <input name="meio"></div>
        <div>Ofício: <input name="num_oficio"></div>
        <div>Data: <input type="date" name="data_consul"></div>
        <div>Palavras-chave: <textarea name="palavras_chave"></textarea></div>

        <div>Secretarias:
            <select name="secretarias[]" multiple size="6">
                <option>CM</option><option>SEGOV</option><option>SMGAS</option>
                <option>PGM</option><option>SMA</option><option>SMF</option>
                <option>SME</option><option>SMCT</option><option>SMS</option>
                <option>SEDES</option><option>SMAGRO</option><option>SEINFRA</option>
                <option>SETTRAN</option><option>DMAE</option><option>FUTEL</option>
                <option>EMAM</option><option>FERUB</option><option>IPREMU</option>
                <option>SESURB</option><option>SMH</option><option>SEJUV</option>
                <option>SECOM</option><option>SEDEI</option><option>SMGE</option>
                <option>SSEG</option><option>ARESAN</option>
            </select>
        </div>

        <div>Observação:
            <textarea name="observacao"></textarea>
        </div>
    </div>

    <button type="button" onclick="adicionar()">➕ Adicionar registro</button>

    <div style="margin-top:10px;">
        <label>Observação geral:</label>
        <textarea name="observacoes" rows="4" style="width:100%;"></textarea>
    </div>

    <div id="box_coparticipantes" style="margin-top:10px;">
    <label>Co-participantes</label><br>
    <select name="coparticipantes[]" multiple size="6" style="width:100%;">
        {% for c in colaboradores %}
            {% if c.id != session["user_id"] %}
                <option value="{{ c.id }}">{{ c.nome }}</option>
            {% endif %}
        {% endfor %}
    </select>
</div>

    <button class="btn" style="margin-top:15px;">
        Registrar Lançamento(s)
    </button>

</form>

<script>
document.addEventListener("DOMContentLoaded", function () {

    const osSelect = document.getElementById("os_select");
    const itemInput = document.getElementById("item_paint");

    const boxReq = document.getElementById("box_requisicoes");
    const boxAtendimento = document.getElementById("box_atendimento");
    const boxConsultoria = document.getElementById("box_consultoria");

    osSelect.addEventListener("change", function () {

        const selected = this.selectedOptions[0];
        const codigoOS = this.value;

        itemInput.value = selected ? selected.dataset.item : "";

        // requisições
        if (codigoOS === "1.4/2026" ||
            codigoOS === "1.1/2026" ||
            codigoOS === "1.6/2026") {
            boxReq.style.display = "block";
        } else {
            boxReq.style.display = "none";
        }

        // OS específicas
        boxAtendimento.style.display = (codigoOS === "1.15/2026") ? "block" : "none";
        boxConsultoria.style.display =
            (codigoOS === "1.14/2026" || codigoOS === "1.16/2026") ? "block" : "none";
            
        // ✅ NOVO: esconder coparticipantes
        const boxCop = document.getElementById("box_coparticipantes");
    
        if (["1.14/2026", "1.15/2026", "1.16/2026"].includes(codigoOS)) {
            boxCop.style.display = "none";
        } else {
            boxCop.style.display = "block";
        }
    });

    // busca rápida
    document.getElementById("busca_req").addEventListener("keyup", function () {
        const f = this.value.toLowerCase();
        document.querySelectorAll(".req_item").forEach(e => {
            e.style.display = e.innerText.toLowerCase().includes(f) ? "" : "none";
        });
    });

});

// múltiplos registros
function adicionar() {
    const base = document.querySelector(".registro");
    const clone = base.cloneNode(true);
    clone.querySelector("input[name='duracao[]']").value = "";
    document.getElementById("registros").appendChild(clone);
}

function remover(btn) {
    const registros = document.querySelectorAll(".registro");
    if (registros.length > 1) btn.parentElement.remove();
}
</script>

<script>
document.addEventListener("input", function(e){
    if(e.target.name === "duracao[]"){
        let v = e.target.value.replace(/\D/g, "")

        if(v.length >= 3){
            e.target.value = v.slice(0, v.length-2) + ":" + v.slice(-2)
        } else {
            e.target.value = v
        }
    }
})

document.addEventListener("DOMContentLoaded", function () {

    const osSelect = document.getElementById("os_select");

    if (osSelect && osSelect.value) {
        osSelect.dispatchEvent(new Event('change'));
    }

});
</script>

"""

    return render_template_string(
        BASE.replace("{% block content %}{% endblock %}", form_html),
        oss=oss,
        requisicoes=requisicoes,   # ✅ É o que o HTML usa
        os_pre=os_pre,   # <-- aqui
        colaboradores=colaboradores,
        data_padrao=data_padrao,
        user=session['user'],
        perfil=session['perfil']
    )

# -------------------------
# Relatórios
# -------------------------
@app.route('/relatorios')
def relatorios():
    if 'user' not in session:
        return redirect('/')

    con = get_db()
    cur = con.cursor()

    perfil = session['perfil']
    user_id = session['user_id']

    # ------------------------------------------------------------------
    # 1) Total de horas por colaborador → SOMENTE ADMIN VÊ
    # ------------------------------------------------------------------
    por_colab = []
    if perfil == "admin":
        cur.execute("""
            SELECT c.nome,
                   SUM(
                        (SPLIT_PART(h.duracao, ':', 1)::int * 60) +
                         SPLIT_PART(h.duracao, ':', 2)::int
                   ) AS minutos
            FROM horas h
            JOIN colaboradores c ON h.colaborador_id = c.id
            GROUP BY c.nome
        """)
        por_colab = cur.fetchall()

    # ------------------------------------------------------------------
    # 2) Total por item_paint → TODOS VEEM
    # ------------------------------------------------------------------
    cur.execute("""
        SELECT item_paint,
               SUM(
                   (SPLIT_PART(duracao, ':', 1)::int * 60) +
                    SPLIT_PART(duracao, ':', 2)::int
               ) AS minutos
        FROM horas
        GROUP BY item_paint
    """)
    por_paint = cur.fetchall()

    # ------------------------------------------------------------------
    # 3) Minhas marcações com paginação
    # ------------------------------------------------------------------

    # ---- Tratamento de limite ----
    limit_param = request.args.get("limit", "100")

    if limit_param == "all":
        limite = None
    else:
        try:
            limite = int(limit_param)
        except:
            limite = 100

    mes_filtro = request.args.get("mes", "")

    sql = """
        SELECT
            h.*,
            p.item_paint as item,
            o.resumo AS os_resumo
        FROM horas h
        LEFT JOIN projeto_paint p ON p.item_paint = h.item_paint
        LEFT JOIN os o ON o.codigo = h.os_codigo
        WHERE h.colaborador_id = %s
    """

    params = [user_id]

    # filtro por mês escolhido
    if mes_filtro:
        sql += " AND EXTRACT(MONTH FROM h.data) = %s "
        params.append(mes_filtro)

    sql += " ORDER BY h.data DESC "

    # aplica LIMIT só se limite não for None
    if limite:
        sql += " LIMIT %s "
        params.append(limite)

    cur.execute(sql, tuple(params))
    minhas = cur.fetchall()

    con.close()

    # ------------------------- HTML -------------------------
    html = "<h3>Relatórios</h3>"

    # ==================================================================
    # TOTAL POR COLABORADOR
    # ==================================================================
    if perfil == "admin":
        html += "<h4>Total de horas por colaborador</h4>"
        html += "<table><tr><th>Colaborador</th><th>Total (HH:MM)</th></tr>"
        for r in por_colab:
            minutos = r['minutos'] or 0
            hh = minutos // 60
            mm = minutos % 60
            html += f"<tr><td>{r['nome']}</td><td>{hh:02d}:{mm:02d}</td></tr>"
        html += "</table><br>"

    # ==================================================================
    # TOTAL POR ITEM PAINT
    # ==================================================================
    html += "<h4>Total de horas por Item PAINT</h4>"
    html += "<table><tr><th>Item</th><th>Total (HH:MM)</th></tr>"
    for r in por_paint:
        minutos = r['minutos'] or 0
        hh = minutos // 60
        mm = minutos % 60
        html += f"<tr><td>{r['item_paint']}</td><td>{hh:02d}:{mm:02d}</td></tr>"
    html += "</table><br>"

    # ==================================================================
    # MINHAS MARCAÇÕES
    # ==================================================================
    html += "<h4>Minhas marcações</h4>"

    # ---------------- PAGINAÇÃO ----------------
    html += f"""
        <div style='margin:10px 0;'>
            Mostrar:
            <a href='/relatorios?limit=50'>50</a> |
            <a href='/relatorios?limit=100'>100</a> |
            <a href='/relatorios?limit=200'>200</a> |
            <a href='/relatorios?limit=500'>500</a> |
            <a href='/relatorios?limit=1000'>1000</a> |
            <a href='/relatorios?limit=all'>Todos</a>
        </div>
    """

    # ---------------- BOTÕES POR MÊS ----------------
    html += """
        <div style='margin-bottom:10px;'>
            <strong>Filtrar por mês:</strong><br>
    """

    meses = {
        "01": "Janeiro", "02": "Fevereiro", "03": "Março", "04": "Abril",
        "05": "Maio", "06": "Junho", "07": "Julho", "08": "Agosto",
        "09": "Setembro", "10": "Outubro", "11": "Novembro", "12": "Dezembro"
    }

    for num, nome in meses.items():
        html += f"<a class='btn' style='margin:3px;' href='/relatorios?mes={num}&limit={limit_param}'>{nome}</a>"

    html += """
        <a class='btn' style='margin:3px; background:#444;' href='/relatorios'>Limpar</a>
        </div>
    """

    # ---------------- FILTRO GERAL ----------------
    html += """
        <input type='text' id='filtroGeral' placeholder='Pesquisar em qualquer campo...'
               style='width:100%; padding:6px; margin-bottom:10px;'>
    """

    # ---------------- TABELA ----------------
    html += """
    <table id='tabelaMarcacoes'>
        <tr>
            <th>Data</th>
            <th>Item</th>
            <th>OS</th>
            <th>Atividade</th>
            <th>Observação</th>
            <th>Duração</th>
            <th>Ações</th>
        </tr>
    """
    
    for r in minhas:
        os_visual = r['os_codigo'] or ''
        if r['os_resumo']:
            os_visual += f" - {r['os_resumo']}"
        obs = (r["observacoes"] or "").strip()
        if len(obs) > 90:
            obs = obs[:90] + "..."

        html += f"""
            <tr>
                <td>{fmt(r['data'])}</td>
                <td>{r['item_paint']}</td>
                <td>{os_visual}</td>
                <td>{r['atividade']}</td>
                <td title="{r['observacoes'] or ''}">{obs}</td>
                <td>{r['duracao']}</td>
                <td style="white-space: nowrap;">
                <a class='btn' href='/editar/{r["id"]}'>Editar</a>
                <a class='btn' style='background:#c0392b; margin-left:5px;'
                   href='/excluir_hora/{r["id"]}'
                   onclick="return confirm('Confirma a exclusão deste lançamento?')">
                   Excluir
                </a>
                </td>

            </tr>
        """

    html += "</table>"

    if perfil == "admin":
        html += f"""
            <div style="margin-top:15px;">
            
            <select id="mes_export" style="padding:6px;">
                <option value="">Selecione o mês...</option>
        """
    
        for num, nome in meses.items():
            selected = "selected" if num == mes_filtro else ""
            html += f"<option value='{num}' {selected}>{nome}</option>"
    
        html += """
            </select>
    
            <button class="btn" onclick="exportarMes()" style="margin-left:10px;">
                Exportar mês
            </button>
    
            <a class="btn" style="margin-left:10px" href="/export_preventivas">
                Exportar Preventivas
            </a>
        </div>
        """

    # ---------------- SCRIPTS ----------------
    html += """
    <script>
    // FILTRO GERAL
    document.getElementById("filtroGeral").addEventListener("keyup", function() {
        let filtro = this.value.toLowerCase();
        let linhas = document.querySelectorAll("#tabelaMarcacoes tr");

        linhas.forEach((tr, i) => {
            if (i === 0) return; // pula cabeçalho
            tr.style.display = tr.innerText.toLowerCase().includes(filtro) ? "" : "none";
        });
    });

function exportarMes() {
    const mes = document.getElementById("mes_export").value;

    if (!mes) {
        alert("Selecione um mês para exportar.");
        return;
    }

    window.location.href = "/export?mes=" + mes;
}
    </script>
    """

    return render_template_string(
        BASE.replace("{% block content %}{% endblock %}", html),
        user=session['user'],
        perfil=session['perfil']
    )

# -------------------------
# Editar Registro de Hora
# -------------------------
@app.route("/editar/<int:hid>", methods=["GET", "POST"])
def editar(hid):
    if "user" not in session:
        return redirect("/")

    from datetime import datetime, date

    con = get_db()
    cur = con.cursor()

    cur.execute("SELECT * FROM horas WHERE id=%s", (hid,))
    base = cur.fetchone()
    if not base:
        con.close()
        return "Registro não encontrado."

    # -------------------------
    # Segurança
    # -------------------------
    if session["perfil"] != "admin" and base["colaborador_id"] != session["user_id"]:
        con.close()
        return "Acesso negado."

    # -------------------------
    # Registros a editar
    # 👉 SOMENTE o que aparece como agrupado no relatório
    # -------------------------
    cur.execute("""
        SELECT requisicao_id
        FROM horas_requisicoes
        WHERE hora_id = %s
    """, (hid,))

    reqs_base = sorted([r["requisicao_id"] for r in cur.fetchall()])
    sem_reqs = len(reqs_base) == 0
    
    cur.execute("""
        SELECT h.*
        FROM horas h
        LEFT JOIN horas_requisicoes hr ON hr.hora_id = h.id
        WHERE h.colaborador_id = %s
          AND h.data = %s
          AND h.os_codigo = %s
          AND h.atividade = %s
          AND COALESCE(h.observacoes,'') = COALESCE(%s,'')
        GROUP BY h.id
        HAVING
            CASE
                WHEN %s THEN COUNT(hr.requisicao_id) = 0
                ELSE array_agg(hr.requisicao_id ORDER BY hr.requisicao_id) = %s
            END
        ORDER BY h.data, h.id
    """, (
        base["colaborador_id"],
        base["data"],
        base["os_codigo"],
        base["atividade"],
        base["observacoes"],
        sem_reqs,
        reqs_base
    ))

    registros = cur.fetchall()

    if not registros:
        con.close()
        return "Registro não encontrado."

    ids_horas = [r["id"] for r in registros]
    
    cur.execute("""
        SELECT DISTINCT requisicao_id
        FROM horas_requisicoes
        WHERE hora_id = ANY(%s::int[])
    """, (ids_horas,))

    reqs_vinculadas = {r["requisicao_id"] for r in cur.fetchall()}
    
    primeiro = registros[0]

    # -------------------------
    # OS
    # -------------------------
    cur.execute("SELECT codigo, item_paint, resumo FROM os ORDER BY codigo")
    oss = cur.fetchall()

    # -------------------------
    # Requisições
    # -------------------------
    cur.execute("""
        SELECT DISTINCT r.id, r.chave, r.tipo, r.criterio, r.status_analise
        FROM requisicoes r
        JOIN horas_requisicoes hr ON hr.requisicao_id = r.id
        WHERE hr.hora_id = ANY(%s)
        ORDER BY r.chave
    """, (ids_horas,))

    requisicoes = [dict(r) for r in cur.fetchall()]
    # 🔀 Ordena: marcadas primeiro, depois as demais
    requisicoes.sort(
        key=lambda r: (r["id"] not in reqs_vinculadas, r["chave"])
    )

    # -------------------------
    # POST
    # -------------------------
    if request.method == "POST":

        os_codigo = request.form.get("os")
        item = request.form.get("item")
        atividade = request.form.get("atividade")
        observacoes = request.form.get("observacoes")
    
        requisicoes_ids = request.form.getlist("requisicoes[]")  # ✅ AQUI
    
        ids_form = request.form.getlist("hora_id[]")
        datas = request.form.getlist("data[]")
        duracoes = request.form.getlist("duracao[]")
    
        if not datas:
            con.close()
            return "Nenhum registro enviado."
    
        # 🔐 trava ano
        for d in datas:
            if datetime.strptime(d, "%Y-%m-%d").year != 2026:
                con.close()
                return "Só é permitido editar registros de 2026."
    
        ids_existentes = {r["id"] for r in registros}
        ids_enviados = set()
    
        if not (len(datas) == len(duracoes) == len(ids_form)):
            con.close()
            return "Erro nos dados enviados"
        
        for hid_atual, data, dur in zip(ids_form, datas, duracoes):
        
            if not dur:
                continue
        
            try:
                h, m = map(int, dur.split(":"))
                if m >= 60:
                    continue
            except:
                continue
        
            minutos = h * 60 + m
            duracao = f"{minutos//60:02d}:{minutos%60:02d}"
        
            if hid_atual:
                hid_atual = int(hid_atual)
        
                cur.execute("""
                    UPDATE horas SET
                        data=%s,
                        duracao=%s,
                        duracao_minutos=%s,
                        os_codigo=%s,
                        item_paint=%s,
                        atividade=%s,
                        observacoes=%s
                    WHERE id=%s
                """, (
                    data,
                    duracao,
                    minutos,
                    os_codigo,
                    item,
                    atividade,
                    observacoes,
                    hid_atual
                ))
        
                hora_id = hid_atual
        
            else:
                cur.execute("""
                    INSERT INTO horas
                    (colaborador_id, data, item_paint, os_codigo, atividade,
                     duracao, duracao_minutos, observacoes)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                    RETURNING id
                """, (
                    base["colaborador_id"],
                    data,
                    item,
                    os_codigo,
                    atividade,
                    duracao,
                    minutos,
                    observacoes
                ))
        
                hora_id = cur.fetchone()["id"]
        
            # ✅ FUNDAMENTAL (resolve seu bug)
            ids_enviados.add(hora_id)
        
            # requisicoes
            cur.execute("DELETE FROM horas_requisicoes WHERE hora_id = %s", (hora_id,))
        
            for req_id in requisicoes_ids:
                cur.execute("""
                    INSERT INTO horas_requisicoes (hora_id, requisicao_id)
                    VALUES (%s, %s)
                """, (hora_id, req_id))
    
        # 🗑️ remove os apagados
        for hid_del in ids_existentes - ids_enviados:
            cur.execute("DELETE FROM horas WHERE id=%s", (hid_del,))
    
        con.commit()
        con.close()
        return redirect("/relatorios")

    con.close()

    # -------------------------
    # HTML
    # -------------------------
    html = """
<style>
.req-box {
    max-height: 220px;
    overflow-y: auto;
    border: 1px solid #ccc;
    border-radius: 4px;
}

.req-item {
    display: flex;
    align-items: center;
    gap: 8px;
    padding: 6px 8px;
    cursor: pointer;
    border-bottom: 1px solid #eee;
}

.req-item:last-child {
    border-bottom: none;
}

.req-item:hover {
    background: #f5f7fa;
}

.req-item input {
    cursor: pointer;
}

.req-item.selecionada {
    background: #e8f0ff;
    font-weight: 600;
}

.req-text small {
    color: #666;
    font-weight: normal;
}
</style>

<h3>Editar Registro #{{ hid }}</h3>

<form method="post" style="max-width:650px">

<div id="registros">
{% for r in registros %}
<div class="registro">
    <input type="hidden" name="hora_id[]" value="{{ r.id }}">
    <input type="date" name="data[]" value="{{ r.data }}">
    <input type="text" name="duracao[]" value="{{ r.duracao }}" placeholder="HH:MM" required>
    <button type="button" onclick="remover(this)">🗑</button>
</div>
{% endfor %}
</div>

<button type="button" onclick="adicionar()">➕ Adicionar registro</button>

<br><br>

<label>OS:</label>
<select name="os" id="os_select">
<option></option>
{% for o in oss %}
<option value="{{ o.codigo }}" data-item="{{ o.item_paint }}"
{% if o.codigo == primeiro.os_codigo %}selected{% endif %}>
{{ o.codigo }}{% if o.resumo %} - {{ o.resumo }}{% endif %}
</option>
{% endfor %}
</select>

<br>
Item:
<input name="item" id="item_paint" value="{{ primeiro.item_paint }}" readonly>

<div id="box_requisicoes">
  <label>Requisições:</label>

  <input type="text" id="filtroReq"
         placeholder="🔍 Pesquisar chave, tipo ou critério..."
         style="width:100%; margin-bottom:6px; padding:6px;">

  <div class="req-box">
    {% for r in requisicoes %}
      <label class="req-item {% if r.id in reqs_vinculadas %}selecionada{% endif %}"
             data-text="{{ (r.chave ~ ' ' ~ r.tipo ~ ' ' ~ r.criterio)|lower }}">
        
        <input type="checkbox"
               name="requisicoes[]"
               value="{{ r.id }}"
               {% if r.id in reqs_vinculadas %}checked{% endif %}>
        
        <span class="req-text">
          {{ r.chave }} — {{ r.tipo }} — {{ r.criterio }}
          <small>({{ r.status_analise }})</small>
        </span>
      </label>
    {% endfor %}
  </div>
</div>

<br>
<br>
Atividade:
<select name="atividade">
<option {% if primeiro.atividade.startswith("1") %}selected{% endif %}>1. Planejamento</option>
<option {% if primeiro.atividade.startswith("2") %}selected{% endif %}>2. Execução</option>
<option {% if primeiro.atividade.startswith("3") %}selected{% endif %}>3. Relatório</option>
</select>

<br>
Observações:
<textarea name="observacoes">{{ primeiro.observacoes or '' }}</textarea>

<br><br>
<button class="btn">Salvar Alterações</button>
<a class="btn" href="/relatorios">Cancelar</a>

<script>
const osSelect = document.getElementById("os_select");
const itemInput = document.getElementById("item_paint");

function adicionar() {
    const base = document.querySelector(".registro");
    const clone = base.cloneNode(true);

    clone.querySelector("input[name='hora_id[]']").value = "";
    clone.querySelectorAll("input[type='date'], input[name='duracao[]']").forEach(i => i.value = "");

    document.getElementById("registros").appendChild(clone);
}

function remover(btn) {
    const registros = document.querySelectorAll(".registro");
    if (registros.length > 1) {
        btn.parentElement.remove();
    } else {
        alert("É necessário manter pelo menos um registro.");
    }
}

document.getElementById("filtroReq").addEventListener("keyup", function () {
    const termo = this.value.toLowerCase();
    document.querySelectorAll(".req-item").forEach(el => {
        el.style.display = el.dataset.text.includes(termo) ? "flex" : "none";
    });
});

// destaca ao clicar
document.querySelectorAll(".req-item input").forEach(chk => {
    chk.addEventListener("change", function () {
        this.closest(".req-item")
            .classList.toggle("selecionada", this.checked);
    });
});

document.addEventListener("input", function(e){
    if(e.target.name === "duracao[]"){
        let v = e.target.value.replace(/\D/g, "")

        if(v.length >= 3){
            e.target.value = v.slice(0, v.length-2) + ":" + v.slice(-2)
        } else {
            e.target.value = v
        }
    }
})

</script>
"""

    return render_template_string(
        BASE.replace("{% block content %}{% endblock %}", html),
        registros=registros,
        primeiro=primeiro,
        hid=hid,
        oss=oss,
        user=session["user"],
        perfil=session["perfil"],
        reqs_vinculadas=reqs_vinculadas,
        requisicoes=requisicoes,
    )

# -------------------------
# Excluir lançamento de horas
# -------------------------
@app.route('/excluir_hora/<int:id>')
def excluir_hora(id):
    if 'user' not in session:
        return redirect('/')

    con = get_db()
    cur = con.cursor()

    # -------------------------------------------------
    # Buscar registro (para validar dono e ajustar HH)
    # -------------------------------------------------
    cur.execute("""
        SELECT colaborador_id, item_paint, duracao_minutos
        FROM horas
        WHERE id = %s
    """, (id,))
    h = cur.fetchone()

    if not h:
        con.close()
        return "Registro não encontrado"

    # -------------------------------------------------
    # Segurança: colaborador só exclui o próprio registro
    # admin pode excluir qualquer um
    # -------------------------------------------------
    if session['perfil'] != 'admin' and h['colaborador_id'] != session['user_id']:
        con.close()
        return "Acesso negado"

    # -------------------------------------------------
    # Excluir registro
    # -------------------------------------------------
    cur.execute("DELETE FROM horas WHERE id = %s", (id,))
    con.commit()
    con.close()

    return redirect('/relatorios')

# -------- ADMIN - GERENCIAR PROJETOS --------
@app.route("/admin_projetos")
def admin_projetos():
    if session.get("perfil") not in ["admin", "comum"]:
        return redirect("/")

    def icon(v):
        if v == 1:
            return "<span style='color:green; font-weight:bold;'>✔</span>"
        else:
            return "<span style='color:red; font-weight:bold;'>✖</span>"

    conn = get_db()
    cur = conn.cursor()

    # =============================
    # 1) TOTAL DE REGISTROS
    # =============================
    cur.execute("SELECT COUNT(*) AS total FROM projeto_paint")
    total_paint = cur.fetchone()["total"]

    cur.execute("SELECT COUNT(*) AS total FROM os")
    total_os = cur.fetchone()["total"]

    # =============================
    # 2) TOTAL HH (soma hh_atual dos projetos)
    # =============================
    cur.execute("SELECT SUM(COALESCE(hh_atual,0)) AS total_hh FROM projeto_paint")
    total_hh = cur.fetchone()["total_hh"] or 0

    # =============================
    # 3) HH executadas
    # =============================
    cur.execute("""
        SELECT SUM(duracao_minutos) AS minutos
        FROM horas
    """)
    total_exec_min = cur.fetchone()["minutos"] or 0

    exec_hh = total_exec_min // 60
    exec_mm = total_exec_min % 60
    total_exec_hhmm = f"{int(exec_hh):02d}:{int(exec_mm):02d}"

    total_hh_min = int(total_hh) * 60
    percentual_global = (total_exec_min / total_hh_min * 100.0) if total_hh_min > 0 else 0
    percentual_global_fmt = f"{percentual_global:.2f}%"

    # =============================
    # 4) LISTAR PROJETOS PAINT
    # =============================
    cur.execute("SELECT * FROM projeto_paint ORDER BY item_paint")
    paint_rows = cur.fetchall()

    # ---- horas executadas por item_paint (1 query só) ----
    cur.execute("""
        SELECT 
            item_paint,
            SUM(duracao_minutos) AS minutos
        FROM horas
        GROUP BY item_paint
    """)

    horas_por_paint = {
        row["item_paint"]: row["minutos"] or 0
        for row in cur.fetchall()
    }

    paint_data = []

    for r in paint_rows:
        minutos = horas_por_paint.get(r["item_paint"], 0)

        hh = minutos // 60
        mm = minutos % 60
        soma = f"{hh:02d}:{mm:02d}"

        percentual_fmt = (
        f"{((minutos / 60) / r['hh_atual'] * 100):.2f}%"
        if r["hh_atual"] else "0%"
        )

        paint_data.append({
        "classificacao": r["classificacao"],
        "item_paint": r["item_paint"],
        "tipo_atividade": r["tipo_atividade"],
        "objeto": r["objeto"],
        "objetivo_geral": r["objetivo_geral"],
        "dt_ini": fmt(r["dt_ini"]),
        "dt_fim": fmt(r["dt_fim"]),
        "hh_atual": r["hh_atual"],
        "hh_exec": soma,
        "percentual": percentual_fmt
        })


    # =============================
    # 5) LISTAR OS
    # =============================
    cur.execute("SELECT * FROM os ORDER BY codigo")
    os_rows = cur.fetchall()
    
    os_data = []
    for r in os_rows:
        prazo, restante = calcular_prazo(r["dt_inicio"],r["dt_previsao_fim"],r["dt_conclusao"])
        os_data.append({
            "codigo": r["codigo"],
            "item_paint": r["item_paint"],
            "resumo": r["resumo"],
            "unidade": r["unidade"],
            "coordenacao": r["coordenacao"],
            "equipe": r["equipe"],
            "observacao": r["observacao"],
            "status": r["status"],
            "plan": r["plan"],
            "exec": r["exec"],
            "rp": r["rp"],
            "rf": r["rf"],
    
            # 👇 CAMPOS QUE FALTAVAM
            "dt_inicio": fmt(r["dt_inicio"]),
            "dt_fim": fmt(r["dt_previsao_fim"]),
            "prazo": prazo,
            "restante": restante,
    
            "dt_conclusao": fmt(r["dt_conclusao"])
        })

    # =============================
    # 6) HTML
    # =============================
    def hh_from_hours_integer(h):
        try:
            h = int(h or 0)
        except:
            h = 0
        return f"{h:02d}:00"

    total_hh_display = hh_from_hours_integer(total_hh)

    html = """
    <style>
        body, html { margin:0; padding:0; width:100vw; }
        .container, #content { width:100% !important; max-width:100% !important; margin:0 !important; padding:10px 20px !important; }
        table { width:100% !important; border-collapse: collapse; margin-bottom: 20px; }
        table th, table td { padding:10px 15px; border:1px solid #ccc; text-align:left; vertical-align:top; }
        table th { background:#f0f0f0; }
        input[type="text"] { width:400px; padding:6px 10px; margin-bottom:10px; }
    </style>
    <h2>Gerenciar Projetos</h2>
    <div style='display:flex; gap:20px; margin-bottom:18px; flex-wrap:wrap;'>
    """

    # === cards + gauge ===
    # garantir percent 0..100
    percent = max(0.0, min(100.0, percentual_global))
    # parâmetros do SVG
    import math
    circumference = 2 * math.pi * 30
    dash = (percent / 100.0) * circumference

    html += f"""
    <div style="display:flex; gap:20px; margin-bottom:18px; align-items:stretch; flex-wrap:wrap;">

        <div style='padding:18px; background:#d6e4ff; border-radius:12px;
                    text-align:center; flex:1; min-width:160px;
                    border:1px solid #9bbcff; box-shadow:0 2px 6px rgba(0,0,0,0.08);'>
            <h4 style='margin:6px 0; color:#1e3a8a;'>Total PAINT</h4>
            <p style='font-size:28px; font-weight:700; margin:6px 0; color:#1e40af;'>{total_paint}</p>
        </div>

        <div style='padding:18px; background:#d1fae5; border-radius:12px;
                    text-align:center; flex:1; min-width:160px;
                    border:1px solid #6ee7b7; box-shadow:0 2px 6px rgba(0,0,0,0.08);'>
            <h4 style='margin:6px 0; color:#065f46;'>Total OS</h4>
            <p style='font-size:28px; font-weight:700; margin:6px 0; color:#047857;'>{total_os}</p>
        </div>

        <div style='padding:18px; background:#e0f2fe; border-radius:12px;
                    text-align:center; flex:1; min-width:200px;
                    border:1px solid #7dd3fc; box-shadow:0 2px 6px rgba(0,0,0,0.08);'>
            <h4 style='margin:6px 0; color:#075985;'>Total HH (planejado)</h4>
            <p style='font-size:22px; font-weight:700; margin:6px 0; color:#0369a1;'>{total_hh_display}</p>
            <div class='small' style='color:#334155'>soma de hh_atual dos projetos</div>
        </div>

        <div style='padding:18px; background:#ffedd5; border-radius:12px;
                    text-align:center; flex:1; min-width:200px;
                    border:1px solid #fdba74; box-shadow:0 2px 6px rgba(0,0,0,0.08);'>
            <h4 style='margin:6px 0; color:#9a3412;'>HH Executadas</h4>
            <p style='font-size:22px; font-weight:700; margin:6px 0; color:#c2410c;'>{total_exec_hhmm}</p>
            <div class='small' style='color:#4b5563'>total de horas registradas</div>
        </div>

        <div style='padding:12px; background:#e8f1ff; border-radius:12px;
                    width:220px; text-align:center; min-width:220px;
                    border:1px solid #93c5fd; box-shadow:0 2px 6px rgba(0,0,0,0.08);'>
            <h4 style='margin:6px 0; color:#1e3a8a;'>% Executado</h4>

            <svg width='120' height='120' viewBox='0 0 120 120' style='display:block;margin:auto'>
              <defs>
                <linearGradient id='gaugeGrad' x1='0%' y1='0%' x2='100%' y2='0%'>
                  <stop offset='0%' stop-color='#2563eb'/>
                  <stop offset='100%' stop-color='#06b6d4'/>
                </linearGradient>
              </defs>
              <g transform='translate(60,60)'>
                <circle r='44' fill='transparent' stroke='#c7d2fe' stroke-width='16'/>
                <circle r='38' fill='transparent' stroke='url(#gaugeGrad)' stroke-width='12'
                        stroke-dasharray='{dash:.2f} {circumference - dash:.2f}'
                        stroke-linecap='round' transform='rotate(-90)' />
                <text x='0' y='6' text-anchor='middle' font-size='18' font-weight='700' fill='#1e40af'>
                    {percent:.2f}%
                </text>
              </g>
            </svg>

            <div class='small' style='color:#334155; margin-top:6px;'>
                {percentual_global_fmt} do total planejado
            </div>
        </div>

    </div>
    """

    # PAINT Table
    html += """
    <h3>Projetos PAINT</h3>
    <input type='text' id='searchPaint' onkeyup="filterTable('searchPaint','paintTable')" placeholder='Pesquisar...'>
    <table id='paintTable'>
        <tr>
            <th>Classificação</th><th>Item</th><th>Tipo</th><th>Objeto</th><th>Objetivo Geral</th>
            <th>Início</th><th>Fim</th><th>HH Atual</th><th>HH Executada</th><th>% Executado</th>
        </tr>
    """
    for r in paint_data:
        html += f"""
        <tr>
            <td>{r['classificacao']}</td><td>{r['item_paint']}</td><td>{r['tipo_atividade']}</td>
            <td>{r['objeto']}</td><td>{r['objetivo_geral']}</td><td>{r['dt_ini']}</td><td>{r['dt_fim']}</td>
            <td>{r['hh_atual']}</td><td>{r['hh_exec']}</td><td>{r['percentual']}</td>
        </tr>
        """
    html += "</table>"

    # OS Table
    html += """
    <h3>Ordens de Serviço (OS)</h3>
    <input type='text' id='searchOS' onkeyup="filterTable('searchOS','osTable')" placeholder='Pesquisar...'>
    <table id='osTable'>
        <tr>
            <th>Código</th><th>Item PAINT</th><th>Resumo</th><th>Unidade</th><th>Coordenação</th>
            <th>Equipe</th><th>Observação</th><th>Status</th><th>PLAN</th><th>EXEC</th>
            <th>RP</th><th>RF</th><th>Início</th><th>Fim</th><th>Prazo</th><th>Restante</th><th>Conclusão</th>
        </tr>
    """
    for r in os_data:
        html += f"""
        <tr>
            <td>{r['codigo']}</td><td>{r['item_paint']}</td><td>{r['resumo']}</td><td>{r['unidade']}</td>
            <td>{r['coordenacao']}</td><td>{r['equipe']}</td><td>{r['observacao']}</td><td>{r['status']}</td>
            <td>{icon(r['plan'])}</td><td>{icon(r['exec'])}</td><td>{icon(r['rp'])}</td><td>{icon(r['rf'])}</td>
            <td>{r['dt_inicio']}</td><td>{r['dt_fim']}</td><td>{r['prazo']}</td><td>{r['restante']}</td><td>{r['dt_conclusao']}</td>
        </tr>
        """
    html += "</table>"

    # JS filter
    html += """
    <script>
    function filterTable(inputId, tableId) {
        var input = document.getElementById(inputId);
        var filter = input.value.toLowerCase();
        var table = document.getElementById(tableId);
        var tr = table.getElementsByTagName("tr");
        for (var i = 1; i < tr.length; i++) {
            var tds = tr[i].getElementsByTagName("td");
            var found = false;
            for (var j = 0; j < tds.length; j++) {
                if (tds[j] && tds[j].innerText.toLowerCase().indexOf(filter) > -1) {
                    found = true; break;
                }
            }
            tr[i].style.display = found ? "" : "none";
        }
    }
    </script>
    """

    return render_template_string(
    BASE.replace("{% block content %}{% endblock %}", html),
    user=session["user"],
    perfil=session["perfil"]
)

@app.route('/visao')
def visao_consolidada():
    if 'user' not in session:
        return redirect('/')

    import json
    con = get_db()
    cur = con.cursor()

    MESES = [
        ("01", "janeiro"), ("02", "fevereiro"), ("03", "marco"),
        ("04", "abril"), ("05", "maio"), ("06", "junho"),
        ("07", "julho"), ("08", "agosto"), ("09", "setembro"),
        ("10", "outubro"), ("11", "novembro"), ("12", "dezembro"),
    ]

    # ============================================================
    # TABELA 1 – HORAS POR COLABORADOR
    # ============================================================
    selects_mes = []
    for num, nome in MESES:
        selects_mes.append(
            f"""
            SUM(CASE WHEN EXTRACT(MONTH FROM h.data) = {int(num)}
                THEN h.duracao_minutos ELSE 0 END) AS {nome}
            """
        )

    sql_colab = f"""
        SELECT c.nome,
               {",".join(selects_mes)},
               SUM(h.duracao_minutos) AS total
        FROM horas h
        JOIN colaboradores c ON c.id = h.colaborador_id
        GROUP BY c.nome
        ORDER BY c.nome
    """

    cur.execute(sql_colab)
    tabela_colab = cur.fetchall()

    total_colab = {nome: 0 for _, nome in MESES}
    total_colab["total"] = 0

    for row in tabela_colab:
        for _, nome in MESES:
            total_colab[nome] += row[nome] or 0
        total_colab["total"] += row["total"] or 0

    # ============================================================
    # TABELA 2 – HORAS POR OS
    # ============================================================
    selects_mes_os = []
    for num, nome in MESES:
        selects_mes_os.append(
            f"""
            SUM(CASE WHEN EXTRACT(MONTH FROM h.data) = {int(num)}
                THEN h.duracao_minutos ELSE 0 END) AS {nome}
            """
        )

    sql_os = f"""
        SELECT o.codigo, o.item_paint, o.resumo,
               p.tipo_atividade, p.objeto, p.objetivo_geral,
               {",".join(selects_mes_os)},
               SUM(h.duracao_minutos) AS total
        FROM horas h
        JOIN os o ON o.codigo = h.os_codigo
        LEFT JOIN projeto_paint p ON p.item_paint = o.item_paint
        GROUP BY o.codigo, o.item_paint, o.resumo,
                 p.tipo_atividade, p.objeto, p.objetivo_geral
        ORDER BY o.item_paint, o.codigo
    """

    cur.execute(sql_os)
    tabela_os = cur.fetchall()

    total_os = {nome: 0 for _, nome in MESES}
    total_os["total"] = 0

    for row in tabela_os:
        for _, nome in MESES:
            total_os[nome] += row[nome] or 0
        total_os["total"] += row["total"] or 0

    # ============================================================
    # GRÁFICO – UMA QUERY SÓ (🔥 ganho grande)
    # ============================================================
    cur.execute("""
        SELECT EXTRACT(MONTH FROM data) AS mes,
               SUM(duracao_minutos) AS minutos
        FROM horas
        GROUP BY mes
        ORDER BY mes
    """)

    mapa = {int(r["mes"]): round((r["minutos"] or 0) / 60, 2) for r in cur.fetchall()}
    totais_mensais = [mapa.get(i, 0) for i in range(1, 13)]
    labels_meses = [nome.capitalize() for _, nome in MESES]

    con.close()

    # ---------------------------------------------------------------------- #
    #                    HTML – RENDERIZAÇÃO                                #
    # ---------------------------------------------------------------------- #

    html = """
    <h2>Visão Consolidada</h2>

    <div class='card'>
        <h3>Horas por Colaborador</h3>
        <input id='filtroColab' onkeyup='filtrar("tabelaColab","filtroColab")' placeholder='Pesquisar colaborador...'>
        <table id='tabelaColab'>
            <thead>
                <tr>
                    <th>Nome</th>
    """

    for _, nome in MESES:
        html += f"<th>{nome.capitalize()}</th>"

    html += "<th>Total Geral</th></tr></thead><tbody>"

    # ---- Corpo ----
    for row in tabela_colab:
        html += "<tr>"
        html += f"<td>{row['nome']}</td>"
        for _, nome in MESES:
            minutos = row[nome] or 0
            html += f"<td>{minutos//60:02d}:{minutos%60:02d}</td>"
        tg = row["total"] or 0
        html += f"<td><b>{tg//60:02d}:{tg%60:02d}</b></td>"
        html += "</tr>"

    # ---- Linha TOTALIZADORA ----
    html += "<tr style='font-weight:bold;background:#eef;'>"
    html += "<td>TOTAL</td>"
    for _, nome in MESES:
        m = total_colab[nome]
        html += f"<td>{m//60:02d}:{m%60:02d}</td>"
    t = total_colab["total"]
    html += f"<td>{t//60:02d}:{t%60:02d}</td>"
    html += "</tr>"

    html += "</tbody></table></div>"
    
    # ------------------------- TABELA 2 (OS) -------------------------
    html += """
    <div class='card'>
        <h3>Horas por O.S</h3>
        <input id='filtroOS' onkeyup='filtrar("tabelaOS","filtroOS")' placeholder='Pesquisar OS / PAINT / Atividade...'>
        <table id='tabelaOS'>
            <thead>
                <tr>
                    <th>Código OS</th>
                    <th>Item PAINT</th>
                    <th>Resumo</th>
    """

    for _, nome in MESES:
        html += f"<th>{nome.capitalize()}</th>"

    html += "<th>Total Geral</th></tr></thead><tbody>"

    # ---- Corpo ----
    for row in tabela_os:
        html += "<tr>"
        html += f"<td>{row['codigo']}</td>"
        html += f"<td>{row['item_paint']}</td>"
        html += f"<td>{row['resumo'] or ''}</td>"

        for _, nome in MESES:
            minutos = row[nome] or 0
            html += f"<td>{minutos//60:02d}:{minutos%60:02d}</td>"

        tg = row["total"] or 0
        html += f"<td><b>{tg//60:02d}:{tg%60:02d}</b></td>"
        html += "</tr>"

    # ---- TOTALIZADOR ----
    html += "<tr style='font-weight:bold;background:#eef;'>"
    html += "<td colspan='3'>TOTAL</td>"
    for _, nome in MESES:
        m = total_os[nome]
        html += f"<td>{m//60:02d}:{m%60:02d}</td>"
    t = total_os["total"]
    html += f"<td>{t//60:02d}:{t%60:02d}</td>"
    html += "</tr>"

    html += "</tbody></table></div>"

    # Filtro JS
    html += """
    <script>
    function filtrar(idTabela, idFiltro) {
        let filtro = document.getElementById(idFiltro).value.toLowerCase();
        let linhas = document.getElementById(idTabela).getElementsByTagName("tr");

        for (let i = 1; i < linhas.length; i++) {
            let texto = linhas[i].innerText.toLowerCase();
            linhas[i].style.display = texto.includes(filtro) ? "" : "none";
        }
    }
    </script>
    """

    # -------------------- FILTROS (LISTAS) --------------------
    cur = get_db().cursor()

    cur.execute("SELECT DISTINCT nome FROM colaboradores ORDER BY nome")
    lista_colab = [r["nome"] for r in cur.fetchall()]
    
    cur.execute("SELECT DISTINCT item_paint FROM os WHERE item_paint IS NOT NULL ORDER BY item_paint")
    lista_paint = [r["item_paint"] for r in cur.fetchall()]
    
    cur.execute("SELECT DISTINCT codigo FROM os ORDER BY codigo")
    lista_os = [r["codigo"] for r in cur.fetchall()]

    f_colab = request.args.get("colaborador")
    f_paint = request.args.get("item_paint")
    f_os = request.args.get("os")

    filtros = []
    params = {}

    if f_colab:
        filtros.append("c.nome = %(colab)s")
        params["colab"] = f_colab
    
    if f_paint:
        filtros.append("o.item_paint = %(paint)s")
        params["paint"] = f_paint
    
    if f_os:
        filtros.append("o.codigo = %(os)s")
        params["os"] = f_os
    
    where_sql = ""
    if filtros:
        where_sql = "WHERE " + " AND ".join(filtros)


    # ---------------------- GRÁFICO (TOTAL MENSAL) ----------------------
    # Prepara lista de totais mensais (em horas float)
    totais_mensais = []

    for num, _ in MESES:
        where_mes = "AND" if where_sql else "WHERE"

        cur.execute(f"""
            SELECT
                SUM(h.duracao_minutos) AS minutos
            FROM horas h
            LEFT JOIN colaboradores c ON c.id = h.colaborador_id
            LEFT JOIN os o ON o.codigo = h.os_codigo
            {where_sql}
            {where_mes} EXTRACT(MONTH FROM h.data) = {int(num)}
        """, params)

        minutos = cur.fetchone()["minutos"] or 0
        totais_mensais.append(round(minutos / 60, 2))

        labels_meses = [nome.capitalize() for _, nome in MESES]
    html += f"""
<div class='card'>
    <h3>Total de Horas por Mês</h3>

    <form method="get" style="margin-bottom:15px;">
        <label>Colaborador:</label>
        <select name="colaborador">
            <option value="">Todos</option>
            {''.join(f"<option {'selected' if f_colab==c else ''}>{c}</option>" for c in lista_colab)}
        </select>

        <label>Item PAINT:</label>
        <select name="item_paint">
            <option value="">Todos</option>
            {''.join(f"<option {'selected' if f_paint==p else ''}>{p}</option>" for p in lista_paint)}
        </select>

        <label>OS:</label>
        <select name="os">
            <option value="">Todas</option>
            {''.join(f"<option {'selected' if f_os==o else ''}>{o}</option>" for o in lista_os)}
        </select>

        <button class="btn">Filtrar</button>
    </form>

    <canvas id="graficoHoras"></canvas>
</div>

<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-datalabels"></script>

<script>
new Chart(document.getElementById('graficoHoras'), {{
    type: 'bar',
    data: {{
        labels: {json.dumps(labels_meses)},
        datasets: [{{
            label: 'Total de horas',
            data: {json.dumps(totais_mensais)},
            borderWidth: 2
        }}]
    }},
    plugins: [ChartDataLabels],
    options: {{
        responsive: true,
        plugins: {{
            datalabels: {{
                anchor: 'end',
                align: 'top'
            }},
            tooltip: {{
                enabled: false
            }}
        }},
        scales: {{
            y: {{
                beginAtZero: true,
                title: {{
                    display: true,
                    text: 'Horas'
                }}
            }}
        }}
    }}
}});
</script>
"""

    return render_template_string(BASE.replace("{% block content %}{% endblock %}", html),
                                  user=session["user"], perfil=session["perfil"])

@app.route('/export')
def export_csv():
    if 'user' not in session or session["perfil"] != "admin":
        return redirect('/')

    import csv, io

    mes = request.args.get("mes")

    if not mes:
        return "Selecione um mês para exportar."

    con = get_db()
    cur = con.cursor()

    cur.execute("""
        SELECT
            h.id AS hora_id,
            c.nome AS colaborador,
            h.data,
            h.item_paint,
            h.os_codigo,
            h.atividade,
            h.duracao,
            h.duracao_minutos,
            h.observacoes,

            string_agg(r.chave, ', ') AS requisicoes,
            COUNT(r.id) AS qtd_requisicoes,
            COALESCE(SUM(r.valor_requisicao), 0) AS valor_total_requisicoes,

            MIN(r.data_inicio) AS data_inicio_requisicao,
            MAX(r.data_fim) AS data_fim_requisicao,
            string_agg(DISTINCT r.status_analise, ', ') AS status_requisicao,
            string_agg(DISTINCT r.tipo, ', ') AS tipo_requisicao,
            string_agg(DISTINCT r.criterio, ', ') AS criterio_requisicao

        FROM horas h
        JOIN colaboradores c ON c.id = h.colaborador_id
        LEFT JOIN horas_requisicoes hr ON hr.hora_id = h.id
        LEFT JOIN requisicoes r ON r.id = hr.requisicao_id

        WHERE EXTRACT(MONTH FROM h.data) = %s

        GROUP BY
            h.id, c.nome, h.data,
            h.item_paint, h.os_codigo, h.atividade,
            h.duracao, h.duracao_minutos, h.observacoes

        ORDER BY h.data
    """, (mes,))

    rows = cur.fetchall()
    con.close()

    si = io.StringIO()
    cw = csv.writer(si, delimiter=";")

    cw.writerow([
        "Hora ID", "Colaborador", "Data",
        "Item PAINT", "OS", "Atividade",
        "Duração", "Minutos", "Obs",
        "Requisições", "Qtd", "Valor Total",
        "Data Início", "Data Fim",
        "Status", "Tipo", "Critério"
    ])

    for r in rows:
        data_fmt = r["data"].strftime("%d/%m/%Y") if r["data"] else ""

        cw.writerow([
            r["hora_id"],
            r["colaborador"],
            data_fmt,
            r["item_paint"],
            r["os_codigo"],
            r["atividade"],
            r["duracao"],
            r["duracao_minutos"],
            r["observacoes"],
            r["requisicoes"] or "",
            r["qtd_requisicoes"] or 0,
            fmt_br(r["valor_total_requisicoes"] or 0),
            r["data_inicio_requisicao"],
            r["data_fim_requisicao"],
            r["status_requisicao"],
            r["tipo_requisicao"],
            r["criterio_requisicao"],
        ])

    output = io.BytesIO()
    output.write("\ufeff".encode("utf-8"))
    output.write(si.getvalue().encode("utf-8"))
    output.seek(0)

    return send_file(
        output,
        mimetype="text/csv",
        as_attachment=True,
        download_name=f"horas_mes_{mes}.csv"
    )

@app.route('/export_minhas')
def export_minhas():
    if 'user' not in session:
        return redirect('/')

    import csv, io
    from datetime import datetime

    user_id = session["user_id"]

    con = get_db()
    cur = con.cursor()

    cur.execute("""
        SELECT
            h.id AS hora_id,
            c.nome AS colaborador,
            h.data,
            h.item_paint,
            h.os_codigo,
            h.atividade,
            h.duracao,
            h.duracao_minutos,
            h.observacoes,

            string_agg(r.chave, ', ') AS requisicoes,
            COUNT(r.id)               AS qtd_requisicoes,
            COALESCE(SUM(r.valor_requisicao), 0) AS valor_total_requisicoes,

            MIN(r.data_inicio)        AS data_inicio_requisicao,
            MAX(r.data_fim)           AS data_fim_requisicao,
            string_agg(DISTINCT r.status_analise, ', ') AS status_requisicao,
            string_agg(DISTINCT r.tipo, ', ')            AS tipo_requisicao,
            string_agg(DISTINCT r.criterio, ', ')        AS criterio_requisicao

        FROM horas h
        JOIN colaboradores c ON c.id = h.colaborador_id
        LEFT JOIN horas_requisicoes hr ON hr.hora_id = h.id
        LEFT JOIN requisicoes r ON r.id = hr.requisicao_id

        WHERE h.colaborador_id = %s

        GROUP BY
            h.id, c.nome, h.data,
            h.item_paint, h.os_codigo, h.atividade,
            h.duracao, h.duracao_minutos, h.observacoes

        ORDER BY h.data;
    """, (user_id,))

    rows = cur.fetchall()
    con.close()

    si = io.StringIO()
    cw = csv.writer(si, delimiter=";")

    cw.writerow([
        "Hora ID", "Colaborador", "Data",
        "Item PAINT", "OS", "Atividade",
        "Duração", "Minutos", "Obs",
        "Requisições", "Qtd Requisições", "Valor Total",
        "Data Início Req", "Data Fim Req",
        "Status Req", "Tipo Req", "Critério Req"
    ])

    for r in rows:
        try:
            data_fmt = r["data"].strftime("%d/%m/%Y")
        except:
            data_fmt = r["data"]

        cw.writerow([
            r["hora_id"],
            r["colaborador"],
            data_fmt,
            r["item_paint"],
            r["os_codigo"],
            r["atividade"],
            r["duracao"],
            r["duracao_minutos"],
            r["observacoes"],
            r["requisicoes"] or "",
            r["qtd_requisicoes"] or 0,
            fmt_br(r["valor_total_requisicoes"] or 0),
            r["data_inicio_requisicao"],
            r["data_fim_requisicao"],
            r["status_requisicao"],
            r["tipo_requisicao"],
            r["criterio_requisicao"],
        ])

    output = io.BytesIO()
    output.write("\ufeff".encode("utf-8"))
    output.write(si.getvalue().encode("utf-8"))
    output.seek(0)

    return send_file(
        output,
        mimetype="text/csv",
        as_attachment=True,
        download_name="minhas_horas.csv"
    )

def minutos_para_hhmm(minutos):
    horas = minutos // 60
    mins = minutos % 60
    return f"{int(horas):02d}:{int(mins):02d}"

@app.route('/export_preventivas')
def export_preventivas():
    if 'user' not in session or session['perfil'] != 'admin':
        return redirect('/')

    import csv, io
    from datetime import datetime, date
    from collections import defaultdict
    from flask import Response

    con = get_db()
    cur = con.cursor()

    cur.execute("""
    SELECT
        r.id AS requisicao_id,
        r.chave,
        r.valor_requisicao,
        r.status_analise,
        r.tipo,
        r.criterio,
        r.os_codigo,
        r.data_inicio,
        r.data_fim,

        c.nome AS colaborador,

        h.data,
        h.duracao_minutos

        FROM requisicoes r
    
        JOIN colaboradores c ON c.id = r.servidor_id      -- 👈 garante que está atribuída
        LEFT JOIN horas_requisicoes hr ON hr.requisicao_id = r.id
        LEFT JOIN horas h ON h.id = hr.hora_id
    
        WHERE r.servidor_id IS NOT NULL                    -- 👈 filtro explícito
    
        ORDER BY r.id, h.data
    """)


    rows = cur.fetchall()
    con.close()

    # ---------------- AGRUPAR POR REQUISIÇÃO ----------------
    grupos = defaultdict(lambda: {
        "chave": "",
        "valor": 0,
        "status": "",
        "tipo": "",
        "criterio": "",
        "os": "",
        "colaborador": "",
        "data_inicio": "",
        "data_fim": "",
        "datas": [],
        "duracao_total_min": 0
    })

    for r in rows:
        g = grupos[r["requisicao_id"]]

        g["chave"] = r["chave"]
        g["valor"] = r["valor_requisicao"] or 0
        g["status"] = r["status_analise"]
        g["tipo"] = r["tipo"]
        g["criterio"] = r["criterio"]
        g["os"] = r["os_codigo"]
        g["colaborador"] = r["colaborador"]
        g["data_inicio"] = r["data_inicio"]
        g["data_fim"] = r["data_fim"]

        # horas (opcional)
        if r["data"]:
            if isinstance(r["data"], (datetime, date)):
                data_fmt = r["data"].strftime("%d/%m/%Y")
            else:
                data_fmt = str(r["data"])

            g["datas"].append(data_fmt)
            g["duracao_total_min"] += r["duracao_minutos"] or 0

    # ---------------- GERAR CSV ----------------
    output = io.StringIO()
    output.write("\ufeff")  # BOM Excel
    writer = csv.writer(output, delimiter=";")

    writer.writerow([
        "Chave Requisição",
        "Valor Requisição",
        "Colaborador",
        "Data Início",
        "Data Fim",
        "Datas Trabalhadas",
        "Horas Totais",
        "Status",
        "Tipo",
        "Critério",
        "OS"
    ])

    for g in grupos.values():
        duracao_total = minutos_para_hhmm(g["duracao_total_min"])

        writer.writerow([
            g["chave"],
            fmt_br(g["valor"]),                 # <<< formato BR
            g["colaborador"],
            g["data_inicio"],
            g["data_fim"],
            "\n".join(g["datas"]),
            duracao_total,
            g["status"],
            g["tipo"],
            g["criterio"],
            g["os"]
        ])

    return Response(
        output.getvalue(),
        mimetype="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": "attachment; filename=requisicoes.csv"
        }
    )

@app.route("/minhas_delegacoes")
def minhas_delegacoes():
    if "user" not in session:
        return redirect("/")

    if session["perfil"] == "admin":
        return redirect("/requisicoes")

    con = get_db()
    cur = con.cursor()

    limit_param = request.args.get("limit", "100")
    status_param = request.args.get("status")

    if limit_param == "all":
        limite = None
    else:
        try:
            limite = int(limit_param)
        except:
            limite = 100

    sql = """
        SELECT r.*
        FROM requisicoes r
        WHERE r.servidor_id = %s
    """
    params = [session["user_id"]]

    if status_param:
        sql += " AND r.status_analise = %s "
        params.append(status_param)

    sql += " ORDER BY r.data_inicio DESC NULLS LAST "

    if limite:
        sql += " LIMIT %s "
        params.append(limite)

    cur.execute(sql, tuple(params))
    requisicoes = cur.fetchall()
    con.close()

    return render_template_string(
        BASE.replace("{% block content %}{% endblock %}", """
<h2>Minhas Requisições Delegadas</h2>

<style>
tr.andamento {
    background-color: #ffe5e5; /* vermelho claro */
    
}

tr.analisando {
    background-color: #fff7cc; /* amarelo claro */
}

tr.analisado {
    background-color: #e5ffe5; /* verde claro */
}
tr.andamento td:first-child::before {
    content:"🔴 ";
}
tr.analisando td:first-child::before {
    content:"🟡 ";
}
tr.analisado td:first-child::before {
    content:"🟢 ";
}


/* Botões */
.btn.andamento { background:#d9534f; color:#fff }
.btn.analisando { background:#f0ad4e; color:#000 }
.btn.analisado { background:#5cb85c; color:#fff }
.btn.all { background:#0275d8; color:#fff }
.btn.ativo {
    outline: 3px solid #000;
    transform: scale(1.05);
    font-weight: bold;
}

/* ================= RESPONSIVIDADE ================= */

.table-container {
    overflow-x: auto;
    margin-top: 10px;
}

/* Largura mínima para evitar esmagamento */
.tabela-delegacoes {
    min-width: 1350px;
    border-collapse: collapse;
}

/* Cabeçalho fixo */
.tabela-delegacoes th {
    position: sticky;
    top: 0;
    background: #dce8fb;
    z-index: 2;
}

/* Primeira coluna congelada e negrito */
.tabela-delegacoes th:first-child,
.tabela-delegacoes td:first-child {
    position: sticky;
    left: 0;
    background: #fff;
    font-weight: bold;
    z-index: 3;
}

/* Ajustes específicos de largura */
.tabela-delegacoes th:nth-child(3),
.tabela-delegacoes td:nth-child(3) {
    min-width: 130px; /* Status maior */
}

.tabela-delegacoes th:nth-child(6),
.tabela-delegacoes td:nth-child(6) {
    min-width: 90px; /* Nota um pouco maior */
}

.tabela-delegacoes th:nth-child(10),
.tabela-delegacoes td:nth-child(10) {
    max-width: 180px; /* Resposta menor */
}

.tabela-delegacoes th:nth-child(11),
.tabela-delegacoes td:nth-child(11) {
    max-width: 180px; /* Obs menor */
}

/* Evita quebra feia */
.tabela-delegacoes td {
    white-space: nowrap;
}
</style>

<div style="margin-bottom:15px;display:flex;gap:10px;flex-wrap:wrap">
    <input type="text" id="filtro"
           placeholder="Pesquisar..."
           style="flex:1;padding:8px">

    <button onclick="filtrarStatus('')" class="btn all" id="btn-all">TODOS</button>
    <button onclick="filtrarStatus('ANDAMENTO')" class="btn andamento" id = "btn-andamento" >ANDAMENTO</button>
    <button onclick="filtrarStatus('ANALISANDO')" class="btn analisando" id = "btn-analisando">ANALISANDO</button>
    <button onclick="filtrarStatus('ANALISADO')" class="btn analisado" id = "btn-analisado">ANALISADO</button>
</div>

{% if requisicoes %}
<div class="table-container">
<table class="tabela-delegacoes">
    <tr>
        <th>Chave</th>
        <th>Início</th>
        <th>Status</th>
        <th>Tipo</th>
        <th>Critério</th>
        <th>Nota</th>
        <th>Nº Nota</th>
        <th>Ofício</th>
        <th>Monit?</th>
        <th>Obs</th>
        <th>Ação</th>
    </tr>
    {% for r in requisicoes %}
    <tr class="{{ r.status_analise|lower }}">
        <td>{{ r.chave }}</td>
        <td>{{ fmt(r.data_inicio) }}</td>
    
        <!-- STATUS -->
        <td>
            <select class="status-select" data-id="{{ r.id }}">
                <option value="ANDAMENTO" {{ "selected" if r.status_analise=="ANDAMENTO" else "" }}>ANDAMENTO</option>
                <option value="ANALISANDO" {{ "selected" if r.status_analise=="ANALISANDO" else "" }}>ANALISANDO</option>
                <option value="ANALISADO" {{ "selected" if r.status_analise=="ANALISADO" else "" }}>ANALISADO</option>
            </select>
        </td>
    
        <td>{{ r.tipo or '' }}</td>
        <td>{{ r.criterio or '' }}</td>
    
        <!-- NOTA -->
        <td>
            <select class="campo-inline" data-id="{{ r.id }}" data-campo="nota">
                <option value=""></option>
                <option value="SIM" {{ "selected" if r.nota=="SIM" else "" }}>SIM</option>
                <option value="NÃO" {{ "selected" if r.nota=="NÃO" else "" }}>NÃO</option>
            </select>
        </td>
    
        <!-- Nº NOTA -->
        <td>
            <input type="text"
                   value="{{ r.num_nota or '' }}"
                   class="campo-inline"
                   data-id="{{ r.id }}"
                   data-campo="num_nota"
                   style="width:90px;">
        </td>
    
        <!-- OFÍCIO -->
        <td>
            <input type="text"
                   value="{{ r.oficio or '' }}"
                   class="campo-inline"
                   data-id="{{ r.id }}"
                   data-campo="oficio"
                   style="width:100px;">
        </td>
    
        <!-- MONITORAMENTO -->
        <td>
            <select class="campo-inline" data-id="{{ r.id }}" data-campo="monitoramento">
                <option value=""></option>
                <option value="SIM" {{ "selected" if r.monitoramento=="SIM" else "" }}>SIM</option>
                <option value="NÃO" {{ "selected" if r.monitoramento=="NÃO" else "" }}>NÃO</option>
            </select>
        </td>

        <!-- OBSERVAÇÕES -->
        <td>
            <input
                type="text"
                class="campo-inline"
                data-id="{{ r.id }}"
                data-campo="observacoes"
                value="{{ r.observacoes or '' }}"
                style="width:260px;">
        </td>
        
        <td>
            <a class="btn" href="/requisicao/{{ r.id }}">Ver</a>
        </td>
    </tr>
    {% endfor %}
</table>
</div>

<script>
document.querySelectorAll(".status-select").forEach(sel => {
    sel.addEventListener("change", function () {

        fetch("/requisicoes", {
            method: "POST",
            headers: { "Content-Type": "application/x-www-form-urlencoded" },
            body: new URLSearchParams({
                acao: "atualizar",
                id: this.dataset.id,
                status_analise: this.value
            })
        })
        .then(r => r.text())
        .then(resp => {
            if (resp !== "OK") {
                alert("Erro ao atualizar status");
                return;
            }
            let tr = this.closest("tr");
            tr.className = this.value.toLowerCase();
        })
        .catch(err => {
            alert("Erro de rede");
            console.error(err);
        });

    });
});

let statusAtual = "ANDAMENTO"; // padrão

function ativarBotao(status){
    document.querySelectorAll(".btn").forEach(b => b.classList.remove("ativo"));

    if(status === ""){
        document.getElementById("btn-all").classList.add("ativo");
    } else {
        document.getElementById("btn-" + status.toLowerCase()).classList.add("ativo");
    }
}

document.getElementById("filtro").addEventListener("keyup", aplicarFiltros);

function filtrarStatus(status){
    statusAtual = status;
    ativarBotao(status);
    aplicarFiltros();
}

function aplicarFiltros(){
    let texto = document.getElementById("filtro").value.toLowerCase();

    document.querySelectorAll("table tr").forEach((tr, i) => {
        if (i === 0) return;

        let matchTexto = tr.innerText.toLowerCase().includes(texto);
        let matchStatus = !statusAtual || tr.classList.contains(statusAtual.toLowerCase());

        tr.style.display = (matchTexto && matchStatus) ? "" : "none";
    });
}

/* APLICA O FILTRO AO CARREGAR A PÁGINA */
document.addEventListener("DOMContentLoaded", () => {
    ativarBotao("ANDAMENTO");
    aplicarFiltros();
});

document.querySelectorAll(".campo-inline, .status-select").forEach(el => {

    let evento = el.tagName === "SELECT" ? "change" : "blur";

    el.addEventListener(evento, function(){

        let tr = this.closest("tr");
        let id = this.dataset.id || tr.querySelector(".status-select").dataset.id;

        let fd = new FormData();
        fd.append("acao","atualizar");
        fd.append("id", id);

        fd.append("status_analise", tr.querySelector(".status-select").value);
        fd.append("nota", tr.querySelector("[data-campo='nota']").value);
        fd.append("num_nota", tr.querySelector("[data-campo='num_nota']").value);
        fd.append("oficio", tr.querySelector("[data-campo='oficio']").value);
        fd.append("monitoramento", tr.querySelector("[data-campo='monitoramento']").value);
        fd.append("observacoes", tr.querySelector("[data-campo='observacoes']").value);

        fetch("/requisicoes", { method:"POST", body: fd })
        .then(r => r.text())
        .then(resp => {
            if(resp !== "OK"){
                alert("Erro ao salvar");
                console.error(resp);
            }
        })
        .catch(err => {
            alert("Erro de rede");
            console.error(err);
        });

    });

});

</script>


{% else %}
<p>Nenhuma requisição delegada para você.</p>
{% endif %}
        """),
        requisicoes=requisicoes,
        fmt=fmt,
        user=session["user"],
        perfil=session["perfil"]
    )

# ---------------------------------------------------------
# Visualizar uma requisição + histórico de horas associadas
# ---------------------------------------------------------
@app.route("/requisicao/<int:id>")
def ver_requisicao(id):
    if "user" not in session:
        return redirect("/")

    con = get_db()
    cur = con.cursor()

    # -------------------------------
    # Buscar dados da requisição
    # -------------------------------
    cur.execute("""
        SELECT r.*, c.nome AS colaborador
        FROM requisicoes r
        LEFT JOIN colaboradores c ON c.id = r.servidor_id
        WHERE r.id = %s
    """, (id,))
    req = cur.fetchone()

    if not req:
        con.close()
        return "Requisição não encontrada"

    # -------------------------------
    # REGRA DE ACESSO
    # -------------------------------
    if session["perfil"] != "admin":
        if req["servidor_id"] != session["user_id"]:
            con.close()
            return "Acesso negado", 403

    # -------------------------------
    # Buscar horas vinculadas
    # -------------------------------
    cur.execute("""
        SELECT
            h.data,
            h.hora_inicio,
            h.hora_fim,
            h.duracao,
            h.atividade,
            col.nome AS colaborador
        FROM horas h
        JOIN horas_requisicoes hr ON hr.hora_id = h.id
        LEFT JOIN colaboradores col ON col.id = h.colaborador_id
        WHERE hr.requisicao_id = %s
        ORDER BY h.data, h.hora_inicio
    """, (id,))

    horas = cur.fetchall()
    con.close()

    # -------------------------------
    # HTML
    # -------------------------------
    html = f"""
    <h2>Requisição {req['chave']}</h2>

    <fieldset>
        <legend><b>Dados da Requisição</b></legend>

        <b>Sigla:</b> {req['sigla']}<br>
        <b>Tipo:</b> {req['tipo'] or '-'}<br>
        <b>Critério:</b> {req['criterio'] or '-'}<br>
        <b>Status:</b> {req['status_analise'] or '-'}<br>
        <b>Responsável:</b> {req['colaborador'] or '-'}<br>
        <b>O.S:</b> {req['os_codigo'] or '-'}<br>
        <b>Data Início:</b> {fmt(req['data_inicio'])}<br>
        <b>Data Fim:</b> {fmt(req['data_fim'])}<br>
        <br>
        <b>Nota:</b> {req['nota'] or '-'}<br>
        <b>Nº Nota:</b> {req['num_nota'] or '-'}<br>
        <b>Ofício:</b> {req['oficio'] or '-'}<br>
        <b>Monitoramento:</b> {req['monitoramento'] or '-'}<br>
        <b>Resposta Monitoramento:</b><br>
        <div style="background:#f5f5f5;padding:8px;border-radius:4px;margin-bottom:6px;">
            {req['monitoramento_resposta'] or '-'}
        </div>
        <b>Observações:</b><br>
        <div style="background:#f5f5f5;padding:8px;border-radius:4px;">
            {req['observacoes'] or '-'}
        </div>
    </fieldset>

    <br>

    <h3>Horas Lançadas nesta Requisição</h3>
    """

    if not horas:
        html += "<p>Nenhum lançamento encontrado.</p>"
    else:
        html += """
        <table border="1" cellpadding="6" cellspacing="0">
            <tr>
                <th>Data</th>
                <th>Hora Início</th>
                <th>Hora Fim</th>
                <th>Duração</th>
                <th>Atividade</th>
                <th>Colaborador</th>
            </tr>
        """
        for h in horas:
            html += f"""
            <tr>
                <td>{fmt(h['data'])}</td>
                <td>{h['hora_inicio']}</td>
                <td>{h['hora_fim']}</td>
                <td>{h['duracao']}</td>
                <td>{h['atividade']}</td>
                <td>{h['colaborador']}</td>
            </tr>
            """
        html += "</table>"

    html += """
    <br>
    <a class="btn" href="javascript:history.back()">⬅ Voltar</a>
    """

    return render_template_string(
        BASE.replace("{% block content %}{% endblock %}", html),
        user=session["user"],
        perfil=session["perfil"]
    )

def minutos_para_hhmm(minutos):
    if minutos is None:
        return "00:00"
    horas = minutos // 60
    mins = minutos % 60
    return f"{horas:02d}:{mins:02d}"

# -------------------------
# Atendimentos (tela)
# -------------------------
@app.route('/atendimentos')
def atendimentos():
    if 'user' not in session:
        return redirect('/')

    mes = request.args.get('mes')  # YYYY-MM
    perfil = session.get("perfil")
    colaborador_id = session.get("user_id")
    nome_user = session.get("user")

    con = get_db()
    cur = con.cursor()

    where = []
    params = []

    # -------------------------
    # REGRA DE VISUALIZAÇÃO
    # -------------------------
    if perfil != "admin":
        where.append("""
            (
                h.colaborador_id = %s
                OR a.responsaveis_consultoria ILIKE %s
            )
        """)
        params.extend([colaborador_id, f"%{nome_user}%"])

    if mes:
        where.append("to_char(a.data_consultoria, 'YYYY-MM') = %s")
        params.append(mes)

    where_sql = "WHERE " + " AND ".join(where) if where else ""

    # -------------------------
    # CONSULTA
    # -------------------------
    sql = f"""
        SELECT
            a.id,
            a.data_consultoria,
            a.assunto,
            a.macro,
            a.diretoria,
            a.atividade,
            a.meio_contato,
            a.duracao_minutos,
            c.nome AS colaborador,
            h.os_codigo,
            h.item_paint
        FROM atendimentos a
        JOIN horas h ON h.id = a.hora_id
        JOIN colaboradores c ON c.id = h.colaborador_id
        {where_sql}
        ORDER BY a.data_consultoria DESC
        {"LIMIT 100" if not mes else ""}
    """

    cur.execute(sql, tuple(params))
    rows = cur.fetchall()
    con.close()

    # -------------------------
    # CONVERTER MINUTOS → HH:MM
    # -------------------------
    atendimentos = []
    for a in rows:
        a = dict(a)
        a["duracao_hhmm"] = minutos_para_hhmm(a["duracao_minutos"])
        a["data_consultoria"] = fmt(a["data_consultoria"])
        atendimentos.append(a)

    # -------------------------
    # HTML
    # -------------------------
    html = """
<h3>Atendimentos</h3>

<form method="get" style="margin-bottom:15px;">
    <label>Mês:</label>
    <input type="month" name="mes" value="{{ mes or '' }}">
    <button class="btn">Filtrar</button>
    <a href="/atendimentos">Limpar</a>
</form>

<a href="/atendimentos/exportar" class="btn" style="margin-bottom:10px; display:inline-block;">
    Exportar atendimentos
</a>

<table border="1" cellpadding="6" cellspacing="0" width="100%">
<tr>
    <th>Data</th>
    {% if perfil == 'admin' %}<th>Colaborador</th>{% endif %}
    <th>OS</th>
    <th>Item</th>
    <th>Assunto</th>
    <th>Macro</th>
    <th>Diretoria</th>
    <th>Meio</th>
    <th>Duração</th>
    <th>Ações</th>
</tr>

{% for a in atendimentos %}
<tr>
    <td>{{ a.data_consultoria }}</td>

    {% if perfil == 'admin' %}
        <td>{{ a.colaborador }}</td>
    {% endif %}

    <td>{{ a.os_codigo }}</td>
    <td>{{ a.item_paint }}</td>
    <td>{{ a.assunto }}</td>
    <td>{{ a.macro }}</td>
    <td>{{ a.diretoria }}</td>
    <td>{{ a.meio_contato }}</td>
    <td style="text-align:right;">{{ a.duracao_hhmm }}</td>
    <td>
        <a href="/atendimentos/ver/{{ a.id }}">Ver</a> |
        <a href="/atendimentos/editar/{{ a.id }}">Editar</a>
    </td>
</tr>
{% endfor %}
</table>
"""

    return render_template_string(
        BASE.replace("{% block content %}{% endblock %}", html),
        atendimentos=atendimentos,
        mes=mes,
        perfil=perfil,
        user=session['user']
    )

@app.route('/atendimentos/ver/<int:id>')
def ver_atendimento(id):
    if 'user' not in session:
        return redirect('/')

    con = get_db()
    cur = con.cursor()

    cur.execute("""
        SELECT *
        FROM atendimentos
        WHERE id = %s
    """, (id,))

    a = cur.fetchone()
    a = dict(a)
    a["duracao_hhmm"] = minutos_para_hhmm(a["duracao_minutos"])
    a["data_consultoria"] = fmt(a["data_consultoria"])
    con.close()

    html = """
<h3>Detalhes do Atendimento</h3>

<p><b>Data:</b> {{ a.data_consultoria }}</p>
<p><b>Assunto:</b> {{ a.assunto }}</p>
<p><b>Macro:</b> {{ a.macro }}</p>
<p><b>Diretoria:</b> {{ a.diretoria }}</p>
<p><b>Atividade:</b> {{ a.atividade }}</p>
<p><b>Meio:</b> {{ a.meio_contato }}</p>
<p><b>Responsáveis:</b> {{ a.responsaveis_consultoria }}</p>
<p><b>Duração:</b> {{ a.duracao_hhmm }}</p>
<p><b>Observação:</b> {{ a.observacao }}</p>

<a href="/atendimentos">Voltar</a>
"""

    return render_template_string(
        BASE.replace("{% block content %}{% endblock %}", html),
        a=a,
        user=session['user'],
        perfil=session['perfil']
    )

@app.route('/atendimentos/editar/<int:id>', methods=['GET', 'POST'])
def editar_atendimento(id):
    if 'user' not in session:
        return redirect('/')

    con = get_db()
    cur = con.cursor()

    if request.method == 'POST':
        cur.execute("""
            UPDATE atendimentos SET
                assunto = %s,
                macro = %s,
                diretoria = %s,
                atividade = %s,
                meio_contato = %s,
                observacao = %s,
                responsaveis_consultoria = %s
            WHERE id = %s
        """, (
            request.form['assunto'],
            request.form['macro'],
            request.form['diretoria'],
            request.form['atividade_atendimento'],
            request.form['meio_contato'],
            request.form['observacao'],
            request.form['responsaveis'],
            id
        ))
        con.commit()
        con.close()
        return redirect('/atendimentos')

    cur.execute("SELECT * FROM atendimentos WHERE id = %s", (id,))
    a = cur.fetchone()
    con.close()

    html = """
<h3>Editar Atendimento</h3>

<form method="post">
    <label>Assunto</label><br>
    <input name="assunto" value="{{ a.assunto }}"><br><br>

    <label>Macro</label><br>
    <input name="macro" value="{{ a.macro }}"><br><br>

    <label>Diretoria</label><br>
    <input name="diretoria" value="{{ a.diretoria }}"><br><br>

    <label>Atividade</label><br>
    <input name="atividade_atendimento" value="{{ a.atividade }}"><br><br>

    <label>Meio</label><br>
    <input name="meio_contato" value="{{ a.meio_contato }}"><br><br>

    <label>Responsáveis</label><br>
    <input name="responsaveis" value="{{ a.responsaveis_consultoria }}"><br><br>

    <label>Observação</label><br>
    <textarea name="observacao">{{ a.observacao }}</textarea><br><br>

    <button class="btn">Salvar</button>
    <a href="/atendimentos">Cancelar</a>
</form>
"""

    return render_template_string(
        BASE.replace("{% block content %}{% endblock %}", html),
        a=a,
        user=session['user'],
        perfil=session["perfil"]
    )

# -------------------------
# Exportar Atendimentos (CSV)
# -------------------------
@app.route('/atendimentos/exportar')
def exportar_atendimentos():
    if 'user' not in session:
        return redirect('/')

    perfil = session.get("perfil")
    colaborador_id = session.get("user_id")

    con = get_db()
    cur = con.cursor()

    where = []
    params = []

    if perfil != "admin":
        where.append("h.colaborador_id = %s")
        params.append(colaborador_id)

    where_sql = "WHERE " + " AND ".join(where) if where else ""

    cur.execute(f"""
        SELECT
            a.data_consultoria,
            c.nome AS colaborador,
            h.os_codigo,
            h.item_paint,
            a.macro,
            a.diretoria,
            a.responsaveis_consultoria,
            a.atividade,
            a.data_consultoria,
            a.assunto,
            a.participantes_externos,
            a.entidades,
            a.meio_contato,
            a.observacao,
            a.duracao_minutos
        FROM atendimentos a
        JOIN horas h ON h.id = a.hora_id
        JOIN colaboradores c ON c.id = h.colaborador_id
        {where_sql}
        ORDER BY a.data_consultoria
    """, tuple(params))

    rows = cur.fetchall()
    con.close()

    import csv, io
    from flask import send_file

    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')

    # ---- cabeçalho amigável
    writer.writerow([
        "Data",
        "Colaborador",
        "OS",
        "Item Paint",
        "Macro",
        "Diretoria",
        "Responsaveis",
        "Atividade",
        "Data",
        "Assunto",
        "Participantes",
        "Entidades",
        "Meio Contato",
        "Obs",
        "Duração (min)"
    ])

    # ---- dados CORRETOS (valores, não as chaves)
    for r in rows:
        writer.writerow([
            r["data_consultoria"],
            r["colaborador"],
            r["os_codigo"],
            r["item_paint"],
            r["macro"],
            r["diretoria"],
            r["responsaveis_consultoria"],
            r["atividade"],
            r["data_consultoria"],
            r["assunto"],
            r["participantes_externos"],
            r["entidades"],
            r["meio_contato"],
            r["observacao"],
            r["duracao_minutos"]
        ])

    output.seek(0)

    return send_file(
        io.BytesIO(output.getvalue().encode("utf-8-sig")),
        mimetype="text/csv",
        as_attachment=True,
        download_name="atendimentos.csv"
    )

@app.route('/consultorias')
def consultorias():
    if 'user' not in session:
        return redirect('/')

    mes = request.args.get('mes')  # YYYY-MM
    perfil = session.get("perfil")
    colaborador_id = session.get("user_id")
    nome_user = session.get("user")

    con = get_db()
    cur = con.cursor()

    where = []
    params = []

    # -------------------------
    # REGRA DE VISUALIZAÇÃO
    # -------------------------
    if perfil != "admin":
        where.append("""
            (
                h.colaborador_id = %s
                OR c.responsaveis ILIKE %s
            )
        """)
        params.extend([colaborador_id, f"%{nome_user}%"])

    if mes:
        where.append("to_char(c.data_consul, 'YYYY-MM') = %s")
        params.append(mes)

    where_sql = "WHERE " + " AND ".join(where) if where else ""

    sql = f"""
        SELECT
            c.id,
            c.data_consul,
            c.assunto,
            c.secretarias,
            c.meio,
            c.tipo,
            c.duracao_minutos,
            col.nome AS colaborador,
            h.os_codigo,
            h.item_paint
        FROM consultorias c
        JOIN horas h ON h.id = c.hora_id
        JOIN colaboradores col ON col.id = h.colaborador_id
        {where_sql}
        ORDER BY c.data_consul DESC
        {"LIMIT 100" if not mes else ""}
    """

    cur.execute(sql, tuple(params))
    rows = cur.fetchall()
    con.close()

    consultorias = []
    for c in rows:
        c = dict(c)
        c["duracao_hhmm"] = minutos_para_hhmm(c["duracao_minutos"])
        c["data_consul"] = fmt(c["data_consul"])
        consultorias.append(c)

    html = """
<h3>Consultorias</h3>

<form method="get" style="margin-bottom:15px;">
    <label>Mês:</label>
    <input type="month" name="mes" value="{{ mes or '' }}">
    <button class="btn">Filtrar</button>
    <a href="/consultorias">Limpar</a>
</form>

<a href="/consultorias/exportar" class="btn" style="margin-bottom:10px; display:inline-block;">
    Exportar consultorias
</a>

<table border="1" cellpadding="6" cellspacing="0" width="100%">
<tr>
    <th>Data</th>
    {% if perfil == 'admin' %}<th>Colaborador</th>{% endif %}
    <th>OS</th>
    <th>Item</th>
    <th>Assunto</th>
    <th>Secretarias</th>
    <th>Meio</th>
    <th>Tipo</th>
    <th>Duração</th>
    <th>Ações</th>
</tr>

{% for c in consultorias %}
<tr>
    <td>{{ c.data_consul }}</td>

    {% if perfil == 'admin' %}
        <td>{{ c.colaborador }}</td>
    {% endif %}

    <td>{{ c.os_codigo }}</td>
    <td>{{ c.item_paint }}</td>
    <td>{{ c.assunto }}</td>
    <td>{{ c.secretarias }}</td>
    <td>{{ c.meio }}</td>
    <td>{{ c.tipo }}</td>
    <td style="text-align:right;">{{ c.duracao_hhmm }}</td>
    <td>
        <a href="/consultorias/ver/{{ c.id }}">Ver</a> |
        <a href="/consultorias/editar/{{ c.id }}">Editar</a>
    </td>
</tr>
{% endfor %}
</table>
"""

    return render_template_string(
        BASE.replace("{% block content %}{% endblock %}", html),
        consultorias=consultorias,
        mes=mes,
        perfil=perfil,
        user=session['user']
    )

@app.route('/consultorias/ver/<int:id>')
def ver_consultoria(id):
    if 'user' not in session:
        return redirect('/')

    con = get_db()
    cur = con.cursor()

    cur.execute("SELECT * FROM consultorias WHERE id=%s", (id,))
    c = dict(cur.fetchone())
    con.close()

    c["duracao_hhmm"] = minutos_para_hhmm(c["duracao_minutos"])
    c["data_consul"] = fmt(c["data_consul"])

    html = """
<h3>Detalhes da Consultoria / Treinamento</h3>
<p><b>Tipo:</b> {{ c.tipo }}</p>
<p><b>Data:</b> {{ c.data_consul }}</p>
<p><b>Assunto:</b> {{ c.assunto }}</p>
<p><b>Secretarias:</b> {{ c.secretarias }}</p>
<p><b>Meio:</b> {{ c.meio }}</p>
<p><b>Responsáveis:</b> {{ c.responsaveis }}</p>
<p><b>Palavras-chave:</b> {{ c.palavras_chave }}</p>
<p><b>Nº Ofício:</b> {{ c.num_oficio }}</p>
<p><b>Duração:</b> {{ c.duracao_hhmm }}</p>
<p><b>Observação:</b> {{ c.observacao }}</p>

<a href="/consultorias">Voltar</a>
"""

    return render_template_string(
        BASE.replace("{% block content %}{% endblock %}", html),
        c=c,
        user=session['user'],
        perfil=session['perfil']
    )

@app.route('/consultorias/editar/<int:id>', methods=['GET', 'POST'])
def editar_consultoria(id):
    if 'user' not in session:
        return redirect('/')

    con = get_db()
    cur = con.cursor()

    if request.method == 'POST':
        cur.execute("""
            UPDATE consultorias SET
                assunto = %s,
                secretarias = %s,
                meio = %s,
                tipo = %s,
                responsaveis = %s,
                palavras_chave = %s,
                num_oficio = %s,
                observacao = %s
            WHERE id = %s
        """, (
            request.form['assunto'],
            request.form['secretarias'],
            request.form['meio'],
            request.form['tipo'],
            request.form['responsaveis'],
            request.form['palavras_chave'],
            request.form['num_oficio'],
            request.form['observacao'],
            id
        ))
        con.commit()
        con.close()
        return redirect('/consultorias')

    cur.execute("SELECT * FROM consultorias WHERE id=%s", (id,))
    c = cur.fetchone()
    con.close()

    html = """
<h3>Editar Consultoria</h3>

<form method="post">
    <label>Assunto</label><br>
    <input name="assunto" value="{{ c.assunto }}"><br><br>

    <label>Secretarias</label><br>
    <input name="secretarias" value="{{ c.secretarias }}"><br><br>

    <label>Meio</label><br>
    <input name="meio" value="{{ c.meio }}"><br><br>

    <label>Tipo</label><br>
    <input name="tipo" value="{{ c.tipo }}"><br><br>

    <label>Responsáveis</label><br>
    <input name="responsaveis" value="{{ c.responsaveis }}"><br><br>

    <label>Palavras-chave</label><br>
    <input name="palavras_chave" value="{{ c.palavras_chave }}"><br><br>

    <label>Nº Ofício</label><br>
    <input name="num_oficio" value="{{ c.num_oficio }}"><br><br>

    <label>Observação</label><br>
    <textarea name="observacao">{{ c.observacao }}</textarea><br><br>

    <button class="btn">Salvar</button>
    <a href="/consultorias">Cancelar</a>
</form>
"""

    return render_template_string(
        BASE.replace("{% block content %}{% endblock %}", html),
        c=c,
        user=session['user'],
        perfil=session["perfil"]
    )

@app.route('/consultorias/exportar')
def exportar_consultorias():
    if 'user' not in session:
        return redirect('/')

    perfil = session.get("perfil")
    colaborador_id = session.get("user_id")

    con = get_db()
    cur = con.cursor()

    where = []
    params = []

    if perfil != "admin":
        where.append("h.colaborador_id = %s")
        params.append(colaborador_id)

    where_sql = "WHERE " + " AND ".join(where) if where else ""

    cur.execute(f"""
        SELECT
            c.data_consul,
            col.nome AS colaborador,
            h.os_codigo,
            h.item_paint,
            c.secretarias,
            c.responsaveis,
            c.assunto,
            c.meio,
            c.tipo,
            c.palavras_chave,
            c.num_oficio,
            c.observacao,
            c.duracao_minutos
        FROM consultorias c
        JOIN horas h ON h.id = c.hora_id
        JOIN colaboradores col ON col.id = h.colaborador_id
        {where_sql}
        ORDER BY c.data_consul
    """, tuple(params))

    rows = cur.fetchall()
    con.close()

    import csv, io
    from flask import send_file

    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')

    writer.writerow([
        "Data",
        "Colaborador",
        "OS",
        "Item Paint",
        "Secretarias",
        "Responsáveis",
        "Assunto",
        "Meio",
        "Tipo",
        "Palavras-chave",
        "Nº Ofício",
        "Observação",
        "Duração (min)"
    ])

    for r in rows:
        writer.writerow([
            r["data_consul"],
            r["colaborador"],
            r["os_codigo"],
            r["item_paint"],
            r["secretarias"],
            r["responsaveis"],
            r["assunto"],
            r["meio"],
            r["tipo"],
            r["palavras_chave"],
            r["num_oficio"],
            r["observacao"],
            r["duracao_minutos"]
        ])

    output.seek(0)

    return send_file(
        io.BytesIO(output.getvalue().encode("utf-8-sig")),
        mimetype="text/csv",
        as_attachment=True,
        download_name="consultorias.csv"
    )

from datetime import datetime
from flask import request, redirect, session, render_template_string, flash, jsonify
import threading
import io

importando_requisicoes = False

progresso_import = {
    "total": 0,
    "processados": 0,
    "inseridos": 0,
    "duplicados": 0,
    "erros": 0,
    "finalizado": False,
    "mensagem": ""
}

def importar_requisicoes_background(arquivo_bytes, data_corte):
    global importando_requisicoes, progresso_import

    SIGLAS = {
        "02": "SEGOV", "03": "SMGAS", "04": "PGM", "05": "SMA",
        "06": "SMF", "07": "SME", "08": "SMCT", "09": "SMS",
        "10": "SMDES", "12": "SMAGRO", "13": "SEINFRA",
        "15": "SETTRAN", "17": "DMAE", "18": "IPREMU",
        "19": "FUTEL", "20": "FERUB", "21": "EMAM",
        "23": "CGM", "24": "SESURB", "25": "SMH",
        "27": "SEJUV", "28": "SECOM", "29": "SEDEI",
        "33": "SMGE", "34": "SEPLAN", "35": "SSEG",
        "38": "ARESAN"
    }

    def parse_data_excel(valor):
        if not valor:
            return None
        if isinstance(valor, datetime):
            return valor
        for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
            try:
                return datetime.strptime(str(valor).strip(), fmt)
            except:
                pass
        return None

    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(arquivo_bytes), read_only=True, data_only=True)
        ws = wb.active

        conn = get_db()
        cur = conn.cursor()

        cur.execute("TRUNCATE requisicoes_staging")

        buffer = io.StringIO()
        BATCH = 2000

        rows = ws.iter_rows(min_row=2, values_only=True)
        progresso_import["total"] = ws.max_row - 1
        
        for idx, row in enumerate(rows, start=1):
            progresso_import["processados"] = idx
            
            try:
                secretaria = row[0]
                num = row[1]
                if not secretaria or not num:
                    continue

                codigo = str(secretaria)[:2]
                sigla = SIGLAS.get(codigo)
                if not sigla:
                    continue

                chave = f"{num}/{sigla}"

                linha = [
                    chave, num, sigla,
                    row[0], row[2], row[3], row[4],
                    parse_data_excel(row[5]),
                    row[6],
                    parse_data_excel(row[7]),
                    row[8], row[9], row[11],
                    row[12], row[13],
                    parse_data_excel(row[14]),
                    parse_data_excel(row[15]),
                    row[16], row[17],
                    data_corte
                ]

                buffer.write("\t".join("" if v is None else str(v) for v in linha) + "\n")

                if idx % BATCH == 0:
                    buffer.seek(0)
                    cur.copy_from(buffer, "requisicoes_staging", sep="\t", null="")
                    buffer.close()
                    buffer = io.StringIO()

            except Exception:
                progresso_import["erros"] += 1

        buffer.seek(0)
        cur.copy_from(buffer, "requisicoes_staging", sep="\t", null="")
        buffer.close()

        cur.execute("""
            INSERT INTO requisicoes (
                chave, requisicao_num, sigla,
                secretaria, tipo_documento, valor_requisicao,
                nome_solicitante, data_criacao, status_atual,
                data_tramitacao, natureza_despesa, item_despesa,
                nome_fornecedor, edital, contrato,
                data_medicao, data_liquidacao, empenho,
                ficha_despesa, data_corte
            )
            SELECT
                chave, requisicao_num, sigla,
                secretaria, tipo_documento, valor_requisicao,
                nome_solicitante, data_criacao, status_atual,
                data_tramitacao, natureza_despesa, item_despesa,
                nome_fornecedor, edital, contrato,
                data_medicao, data_liquidacao, empenho,
                ficha_despesa, data_corte
            FROM requisicoes_staging
            ON CONFLICT (chave) DO UPDATE
            SET
                valor_requisicao = CASE
                    WHEN (
                        EXCLUDED.data_tramitacao IS NOT NULL
                        AND (
                            requisicoes.data_tramitacao IS NULL
                            OR EXCLUDED.data_tramitacao > requisicoes.data_tramitacao
                        )
                        AND requisicoes.valor_requisicao IS DISTINCT FROM EXCLUDED.valor_requisicao
                    )
                    THEN EXCLUDED.valor_requisicao
                    ELSE requisicoes.valor_requisicao
                END,
            
                data_tramitacao = GREATEST(
                    requisicoes.data_tramitacao,
                    EXCLUDED.data_tramitacao
                )
        """)

        progresso_import["inseridos"] = cur.rowcount
        progresso_import["duplicados"] = progresso_import["total"] - cur.rowcount

        conn.commit()
        conn.close()

        progresso_import["finalizado"] = True
        progresso_import["mensagem"] = "Importação concluída com sucesso."

    finally:
        importando_requisicoes = False

@app.route("/requisicoes/importar", methods=["GET", "POST"])
def importar_requisicoes():
    global importando_requisicoes, progresso_import

    if "user" not in session or session["perfil"] != "admin":
        return "Acesso negado", 403

    if request.method == "POST":

        if importando_requisicoes:
            flash("⚠ Já existe uma importação em andamento.")
            return redirect(request.url)

        arquivo = request.files.get("arquivo")
        data_corte = request.form.get("data_corte")

        if not arquivo or not data_corte:
            flash("Arquivo e data de corte são obrigatórios.")
            return redirect(request.url)

        progresso_import = {
            "total": 0,
            "processados": 0,
            "inseridos": 0,
            "duplicados": 0,
            "erros": 0,
            "finalizado": False,
            "mensagem": ""
        }

        importando_requisicoes = True

        t = threading.Thread(
            target=importar_requisicoes_background,
            args=(arquivo.read(), data_corte)
        )
        t.start()

        flash("⏳ Importação iniciada em background.")
        return redirect(request.url)

    html = """
    <h3>📥 Importar Requisições</h3>
    <br>
    <a href="/requisicoes/importar-completo"
       class="btn"
       style="display:inline-block; margin-bottom:15px; background:#0d6efd;">
       Importar Completo - Modelo Drive
    </a>
    
    {% with msgs = get_flashed_messages() %}
      {% if msgs %}
        <div style="background:#fff3cd;padding:10px;border-left:4px solid #f59e0b;margin-bottom:15px">
          {% for m in msgs %}{{ m }}<br>{% endfor %}
        </div>
      {% endif %}
    {% endwith %}

    <form method="post" enctype="multipart/form-data">
    
        <label style="font-weight:600;">Arquivo:</label><br>
        <input type="file" name="arquivo" required><br><br>
    
        <label style="font-weight:600;">Data de Corte:</label><br>
        <input type="date" name="data_corte" required><br><br>
    
        <button class="btn">Importar</button>
    </form>

    <hr>

    <div id="progresso-box" style="display:none;background:#f9fafb;padding:15px;border-radius:8px;">
        <p><b>Processados:</b> <span id="proc">0</span> / <span id="total">0</span></p>
        <p><b>Inseridos:</b> <span id="ins">0</span></p>
        <p><b>Duplicados:</b> <span id="dup">0</span></p>
        <p><b>Erros:</b> <span id="err">0</span></p>

        <progress id="barra" value="0" max="100" style="width:100%;height:20px;"></progress>
    </div>

    <script>
    function atualizarStatus() {
        fetch("/requisicoes/importar/status")
            .then(r => r.json())
            .then(d => {
                if (d.total > 0) {
                    document.getElementById("progresso-box").style.display = "block";
                    document.getElementById("proc").innerText = d.processados;
                    document.getElementById("total").innerText = d.total;
                    document.getElementById("ins").innerText = d.inseridos;
                    document.getElementById("dup").innerText = d.duplicados;
                    document.getElementById("err").innerText = d.erros;

                    let perc = Math.round((d.processados / d.total) * 100);
                    document.getElementById("barra").value = perc;
                }

                if (!d.finalizado) {
                    setTimeout(atualizarStatus, 1000);
                }
            });
    }

    atualizarStatus();
    </script>
    """

    return render_template_string(
        BASE.replace("{% block content %}{% endblock %}", html),
        user=session["user"],
        perfil=session["perfil"]
    )

@app.route("/requisicoes/importar/status")
def status_importacao():
    return jsonify(progresso_import)

import re
def limpar(v):
    if v is None:
        return ""

    v = str(v)

    # remove TAB e quebras
    v = v.replace("\t", " ")
    v = v.replace("\n", " ")
    v = v.replace("\r", " ")

    # remove caracteres invisíveis
    v = re.sub(r'[\x00-\x1f\x7f]', '', v)

    return v.strip()

def safe(r, i):
    if i >= len(r):
        return None
    return r[i]

def importar_requisicoes_completa_background(arquivo_bytes):
    global importando_requisicoes, progresso_import

    from openpyxl import load_workbook
    import io

    try:
        wb = load_workbook(io.BytesIO(arquivo_bytes), read_only=True, data_only=True)
        ws = wb.active

        conn = get_db()
        cur = conn.cursor()

        # Limpa staging
        cur.execute("TRUNCATE requisicoes_staging_completa")

        buffer = io.StringIO()
        BATCH = 2000

        progresso_import["total"] = 0

        for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
            # se não existir número da requisição, parar leitura
            if not r or r[3] is None:
                break
            progresso_import["processados"] = i
            progresso_import["total"] += 1
        
            try:
                # =========================
                # TRATAMENTO DO VALOR
                # =========================
                valor = r[5]
        
                if isinstance(valor, str):
                    valor = valor.strip()
        
                    if valor.startswith("R$"):
                        valor = (
                            valor.replace("R$", "")
                                 .replace(".", "")
                                 .replace(",", ".")
                                 .strip()
                        )
        
                    if valor == "":
                        valor = None
        
                # =========================
                # MONTAGEM DA LINHA
                # =========================
                linha = [
                safe(r,0),   # 1 chave
                safe(r,1),   # 2 data_corte
                safe(r,2), # 3 secretaria
                safe(r,3), # 4 requisicao_num
                safe(r,4), # 5 tipo_documento
                valor, # 6 valor_requisicao (tratado)
                safe(r,6),    # 7 nome_solicitante
                safe(r,7),    # 8 data_criacao
                safe(r,8),    # 9 status_atual
                safe(r,9),    # 10 data_tramitacao
                safe(r,10),   # 11 natureza_despesa
                safe(r,11),   # 12 item_despesa
                safe(r,13),   # 14 nome_fornecedor
                safe(r,14),   # 15 edital
                safe(r,15),   # 16 contrato
                safe(r,16),   # 17 data_medicao
                safe(r,17),   # 18 data_liquidacao
                safe(r,18),   # 19 empenho
                safe(r,19),   # 20 ficha_despesa
                safe(r,20),   # 21 tipo
                safe(r,24),   # 25 criterio
                safe(r,25),   # 26 servidor_nome
                safe(r,26),   # 27 nota
                safe(r,27),   # 28 num_nota
                safe(r,28),   # 29 oficio
                safe(r,29),   # 30 monitoramento
                safe(r,31),   # 32 monitoramento_resposta
                safe(r,32),   # 33 observacoes
            ]
        
                buffer.write(
                    "\t".join(limpar(v) for v in linha) + "\n"
                )
        
                if i % BATCH == 0:
                    if buffer.tell() > 0:
                        buffer.seek(0)
                        cur.copy_from(
                            buffer,
                            "requisicoes_staging_completa",
                            sep="\t",
                            null=""
                        )
                        buffer = io.StringIO()
        
            except Exception as e:
                progresso_import["erros"] += 1
                progresso_import["linhas_com_erro"].append({
                    "linha_excel": i + 1,
                    "requisicao": r[3],
                    "erro": str(e)
                })
                print("ERRO NA LINHA:", i)
                print("DADOS:", linha)
                print("ERRO:", e)

        if buffer.tell() > 0:
            buffer.seek(0)
            cur.copy_from(buffer, "requisicoes_staging_completa", sep="\t", null="")
        buffer.close()

        # STAGING → FINAL
        cur.execute("""
            INSERT INTO requisicoes (
                chave, data_corte, secretaria, requisicao_num,
                tipo_documento, valor_requisicao, nome_solicitante,
                data_criacao, status_atual, data_tramitacao,
                natureza_despesa, item_despesa, nome_fornecedor,
                edital, contrato, data_medicao, data_liquidacao,
                empenho, ficha_despesa, tipo,
                criterio, servidor_id, nota, num_nota,
                oficio, monitoramento, monitoramento_resposta,
                observacoes, status_analise,
                sigla
            )
            SELECT
                s.chave,
                s.data_corte,
                s.secretaria,
                s.requisicao_num,
                s.tipo_documento,
                s.valor_requisicao,
                s.nome_solicitante,
                s.data_criacao,
                s.status_atual,
                s.data_tramitacao,
                s.natureza_despesa,
                s.item_despesa,
                s.nome_fornecedor,
                s.edital,
                s.contrato,
                s.data_medicao,
                s.data_liquidacao,
                s.empenho,
                s.ficha_despesa,
        
                -- ===============================
                -- REGRA DO TIPO (inalterada)
                -- ===============================
                CASE
                    WHEN TRIM(s.tipo) IN ('CONTRATAÇÃO', 'LIQUIDAÇÃO', 'ADITAMENTO')
                        THEN TRIM(s.tipo)
        
                    WHEN s.tipo IS NULL
                         OR TRIM(s.tipo) = ''
                         OR TRIM(s.tipo) = '∄'
                         OR TRIM(s.tipo) NOT IN ('CONTRATAÇÃO', 'LIQUIDAÇÃO', 'ADITAMENTO')
                    THEN
                        CASE
                            WHEN s.tipo_documento IN (
                                'REQUISIÇÕES DE COMPRAS',
                                'REQUERIMENTO PARA REGISTRO DE PREÇOS',
                                'REQUISIÇÃO DE TERMO DE COLABORAÇÃO',
                                'REQUISIÇÃO DE TERMO DE FOMENTO',
                                'REQUISIÇÕES CONSOME SALDO',
                                'REQUISIÇÃO DE PAGAMENTOS DIVERSOS',
                                'REQUISIÇÃO EXTRA ORÇAMENTARIA',
                                'CONTRATO DE GESTÃO',
                                'REQUISIÇÃO EXTRA ORÇAMENTARIA GERAL',
                                'REQUISIÇÃO DE REQUERIMENTO PERMISSÃO DE USO',
                                'REQUISIÇÃO DE COMPRAS - EMENDAS IMPOSITIVAS',
                                'REQUISIÇÃO PAGAMENTO DIVERSOS - EMENDAS IMPOSITIVAS',
                                'REQUISIÇAO TERMO DE FOMENTO-EMENDAS IMPOSITIVAS',
                                'REQUERIMENTO DE COMPRAS',
                                'REQUISIÇÃO COTAÇÃO',
                                'REQUISIÇÃO CONSOME RESERVA COMPRAS'
                            ) THEN 'CONTRATAÇÃO'
        
                            WHEN s.tipo_documento IN (
                                'REQUISIÇÕES P/ LIQUIDAR',
                                'REQUISIÇÕES P/ LIQUIDAR PAGAMENTOS DIVERSOS',
                                'REQUISIÇÕES P/ LIQUIDAR TFD',
                                'REQUISIÇÕES DE LIQUIDAÇÃO-EMENDAS IMPOSITIVAS',
                                'REQUISIÇOES P/ LIQUIDAR DIARIAS E ADIANTAMENTOS VIAGENS',
                                'REQUISIÇÕES P/ REAJUSTAR / REALINHAR - PAGAMENTO DIFERENÇA'
                            ) THEN 'LIQUIDAÇÃO'
        
                            WHEN s.tipo_documento IN (
                                'REQUISIÇÕES P/ ADITAR',
                                'REQUISIÇÕES P/ REAJUSTAR / REALINHAR - ACRÉSCIMO',
                                'REQUISIÇÕES P/ REAJUSTAR / REALINHAR - SUPRESSÃO',
                                'REQUERIMENTO DE SUPRESSÃO',
                                'REQUISIÇÕES DE SUBSTITUIÇÃO À DE PRÓXIMO ORÇAMENTO-ADITIVOS',
                                'REQUISIÇÕES P/ ADITAR - ACRÉSCIMO'
                            ) THEN 'ADITAMENTO'
        
                            ELSE NULL
                        END
                    ELSE NULL
                END,
        
                s.criterio,
        
                -- servidor_id
                CASE UPPER(TRIM(s.servidor_nome))
                    WHEN 'ANA PAULA' THEN 2
                    WHEN 'ALEXANDRA' THEN 1
                    WHEN 'MICHELLE' THEN 14
                    WHEN 'PAULA' THEN 15
                    WHEN 'PRISCILLA' THEN 17
                    WHEN 'SYRIA' THEN 18
                    WHEN 'THAMY' THEN 19
                    WHEN 'MARIANA CAVANHA' THEN 13
                    
                END,
        
                s.nota,
                s.num_nota,
                s.oficio,
                s.monitoramento,
                s.monitoramento_resposta,
                s.observacoes,
        
                -- status_analise
                CASE
                    WHEN s.criterio IS NOT NULL AND TRIM(s.criterio) <> ''
                    THEN 'ANALISADO'
                END,
        
                -- ===============================
                -- SIGLA (MELHOR PRÁTICA)
                -- ===============================
                COALESCE(
                    -- 1️⃣ tenta extrair da chave: 40800/2026/SMS
                    UPPER(NULLIF(split_part(s.chave, '/', 3), '')),
        
                    -- 2️⃣ fallback: mapeamento pela secretaria
                    CASE SUBSTRING(LPAD(s.secretaria::text, 2, '0'), 1, 2)
                        WHEN '02' THEN 'SEGOV'
                        WHEN '03' THEN 'SMGAS'
                        WHEN '04' THEN 'PGM'
                        WHEN '05' THEN 'SMA'
                        WHEN '06' THEN 'SMF'
                        WHEN '07' THEN 'SME'
                        WHEN '08' THEN 'SMCT'
                        WHEN '09' THEN 'SMS'
                        WHEN '10' THEN 'SMDES'
                        WHEN '12' THEN 'SMAGRO'
                        WHEN '13' THEN 'SEINFRA'
                        WHEN '15' THEN 'SETTRAN'
                        WHEN '17' THEN 'DMAE'
                        WHEN '18' THEN 'IPREMU'
                        WHEN '19' THEN 'FUTEL'
                        WHEN '20' THEN 'FERUB'
                        WHEN '21' THEN 'EMAM'
                        WHEN '23' THEN 'CGM'
                        WHEN '24' THEN 'SESURB'
                        WHEN '25' THEN 'SMH'
                        WHEN '27' THEN 'SEJUV'
                        WHEN '28' THEN 'SECOM'
                        WHEN '29' THEN 'SEDEI'
                        WHEN '33' THEN 'SMGE'
                        WHEN '34' THEN 'SEPLAN'
                        WHEN '35' THEN 'SSEG'
                        WHEN '38' THEN 'ARESAN'
                    END
                ) AS sigla
        
            FROM (
                SELECT DISTINCT ON (chave) *
                FROM requisicoes_staging_completa
                ORDER BY
                    chave,
                    CASE 
                        WHEN servidor_nome IS NOT NULL 
                             AND TRIM(servidor_nome) <> '' 
                        THEN 0 
                        ELSE 1 
                    END
            ) s
            
            ON CONFLICT (chave) DO UPDATE
            SET
                data_corte = EXCLUDED.data_corte,
                secretaria = EXCLUDED.secretaria,
                requisicao_num = EXCLUDED.requisicao_num,
                tipo_documento = EXCLUDED.tipo_documento,
            
                -- 💰 VALOR (regra correta)
                valor_requisicao = CASE
                    WHEN
                        EXCLUDED.valor_requisicao IS NOT NULL
                        AND requisicoes.valor_requisicao IS DISTINCT FROM EXCLUDED.valor_requisicao
                        AND (
                            requisicoes.data_tramitacao IS NULL
                            OR EXCLUDED.data_tramitacao IS NULL
                            OR EXCLUDED.data_tramitacao > requisicoes.data_tramitacao
                        )
                    THEN EXCLUDED.valor_requisicao
                    ELSE requisicoes.valor_requisicao
                END,
            
                nome_solicitante = EXCLUDED.nome_solicitante,
                data_criacao = EXCLUDED.data_criacao,
                status_atual = EXCLUDED.status_atual,
            
                -- 📅 DATA (nunca retrocede)
                data_tramitacao = CASE
                    WHEN EXCLUDED.data_tramitacao IS NULL
                        THEN requisicoes.data_tramitacao
                    WHEN requisicoes.data_tramitacao IS NULL
                        THEN EXCLUDED.data_tramitacao
                    ELSE GREATEST(requisicoes.data_tramitacao, EXCLUDED.data_tramitacao)
                END,
            
                natureza_despesa = EXCLUDED.natureza_despesa,
                item_despesa = EXCLUDED.item_despesa,
                nome_fornecedor = EXCLUDED.nome_fornecedor,
                edital = EXCLUDED.edital,
                contrato = EXCLUDED.contrato,
                data_medicao = EXCLUDED.data_medicao,
                data_liquidacao = EXCLUDED.data_liquidacao,
                empenho = EXCLUDED.empenho,
                ficha_despesa = EXCLUDED.ficha_despesa,
            
                -- 🔁 TIPO e CRITÉRIO
                tipo = COALESCE(EXCLUDED.tipo, requisicoes.tipo),
                criterio = COALESCE(EXCLUDED.criterio, requisicoes.criterio),
            
                -- 🧑‍💼 SERVIDOR (não sobrescreve com NULL)
                servidor_id = COALESCE(EXCLUDED.servidor_id, requisicoes.servidor_id),
            
                nota = COALESCE(EXCLUDED.nota, requisicoes.nota),
                num_nota = COALESCE(EXCLUDED.num_nota, requisicoes.num_nota),
                oficio = COALESCE(EXCLUDED.oficio, requisicoes.oficio),
                monitoramento = COALESCE(EXCLUDED.monitoramento, requisicoes.monitoramento),
                monitoramento_resposta = COALESCE(EXCLUDED.monitoramento_resposta, requisicoes.monitoramento_resposta),
                observacoes = COALESCE(EXCLUDED.observacoes, requisicoes.observacoes),
            
                status_analise = COALESCE(EXCLUDED.status_analise, requisicoes.status_analise),
                sigla = COALESCE(EXCLUDED.sigla, requisicoes.sigla)
            
            -- ⚠️ IMPORTANTE: não bloqueia update inteiro por causa do servidor
            WHERE
                (
                    EXCLUDED.servidor_id IS NOT NULL
                    OR requisicoes.servidor_id IS NULL
                );
        """)
        progresso_import["inseridos"] = cur.rowcount
        progresso_import["duplicados"] = progresso_import["total"] - cur.rowcount

        conn.commit()
        conn.close()

        progresso_import["finalizado"] = True
        progresso_import["mensagem"] = "Importação completa concluída com sucesso."

    finally:
        importando_requisicoes = False

@app.route("/requisicoes/importar-completo", methods=["GET", "POST"])
def importar_requisicoes_completo():
    global importando_requisicoes, progresso_import

    if "user" not in session or session["perfil"] != "admin":
        return "Acesso negado", 403

    if request.method == "POST":

        if importando_requisicoes:
            flash("⚠ Já existe uma importação em andamento.")
            return redirect(request.url)

        arquivo = request.files.get("arquivo")
        if not arquivo:
            flash("Arquivo obrigatório.")
            return redirect(request.url)

        progresso_import = {
            "total": 0,
            "processados": 0,
            "inseridos": 0,
            "duplicados": 0,
            "erros": 0,
            "linhas_com_erro": [],   # ← FALTAVA ISSO
            "finalizado": False,
            "mensagem": ""
        }

        importando_requisicoes = True

        threading.Thread(
            target=importar_requisicoes_completa_background,
            args=(arquivo.read(),)
        ).start()

        flash("⏳ Importação completa iniciada.")
        return redirect(request.url)

    return render_template_string("""
    <h2>📥 Importação Completa de Requisições</h2>

    <form method="POST" enctype="multipart/form-data">
        <input type="file" name="arquivo" required><br><br>

        <button class="btn btn-danger"
                onclick="return confirm('Confirma importação COMPLETA?')">
            Importar Arquivo
        </button>
    </form>

    <hr>
    <div id="status"></div>

    <script>
    setInterval(() => {
    fetch("/requisicoes/importar/status")
        .then(r => r.json())
        .then(d => {

            let erros = "";

            if (d.linhas_com_erro && d.linhas_com_erro.length > 0) {

                erros = "<br><b>Erros:</b><br>";

                d.linhas_com_erro.forEach(e => {
                    erros += `Linha ${e.linha_excel} - Req ${e.requisicao}<br>`;
                });
            }

            document.getElementById("status").innerHTML = `
                Processados: ${d.processados}/${d.total}<br>
                Inseridos: ${d.inseridos}<br>
                Duplicados: ${d.duplicados}<br>
                Erros: ${d.erros}<br>
                ${erros}
                ${d.finalizado ? "✅ Finalizado" : "⏳ Em andamento"}
            `;
        });
}, 1500);
    </script>
    """)

@app.route("/requisicoes", methods=["GET", "POST"])
def requisicoes():
    if "user" not in session:
        return redirect("/")

    con = get_db()
    cur = con.cursor()

    # ======================================================
    # ATUALIZAÇÃO INLINE / EXCLUSÃO
    # ======================================================
    if request.method == "POST":
        try:
            acao = request.form.get("acao")
            req_id = request.form.get("id")
    
            if not acao or not req_id:
                return "Ação ou ID inválido", 400
    
            # ==========================================
            # EXCLUIR → SÓ ADMIN
            # ==========================================
            if acao == "excluir":
                if session["perfil"] != "admin":
                    return "Acesso negado", 403
    
                cur.execute("DELETE FROM requisicoes WHERE id = %s", (req_id,))
                con.commit()
                return "OK"
    
            # ==========================================
            # ATUALIZAR CAMPOS
            # ==========================================
            if acao == "atualizar":
    
                status = request.form.get("status_analise")
                tipo = request.form.get("tipo")
                criterio = request.form.get("criterio")
                servidor_id = request.form.get("servidor_id")
                nota = request.form.get("nota")
                num_nota = request.form.get("num_nota")
                oficio = request.form.get("oficio")
                monitoramento = request.form.get("monitoramento")
                monitoramento_resposta = request.form.get("monitoramento_resposta")
                observacoes = request.form.get("observacoes")
                
                if monitoramento == "undefined":
                    monitoramento = None
                    
                if nota == "undefined":
                    nota = None
                
                if session["perfil"] != "admin":
                    cur.execute(
                        "SELECT servidor_id FROM requisicoes WHERE id=%s",
                        (req_id,)
                    )
                    dono = cur.fetchone()
                
                    if not dono or dono["servidor_id"] != session["user_id"]:
                        return "Acesso negado", 403
                
                    # Pode alterar status + campos operacionais
                    cur.execute("""
                        UPDATE requisicoes
                        SET
                            status_analise = %s,
                            nota = NULLIF(%s,''),
                            num_nota = NULLIF(%s,''),
                            oficio = NULLIF(%s,''),
                            monitoramento = NULLIF(%s,''),
                            monitoramento_resposta = NULLIF(%s,''),
                            observacoes = NULLIF(%s,'')
                        WHERE id = %s
                    """, (
                        status,
                        nota,
                        num_nota,
                        oficio,
                        monitoramento,
                        monitoramento_resposta,
                        observacoes,
                        req_id
                    ))
    
                else:
                    cur.execute("""
                        UPDATE requisicoes
                        SET
                            status_analise = NULLIF(%s,''),
                            tipo = NULLIF(%s,''),
                            criterio = NULLIF(%s,''),
                            servidor_id = NULLIF(%s,'')::INTEGER,
                            nota = NULLIF(%s,''),
                            num_nota = NULLIF(%s,''),
                            oficio = NULLIF(%s,''),
                            monitoramento = NULLIF(%s,''),
                            monitoramento_resposta = NULLIF(%s,''),
                            observacoes = NULLIF(%s,'')
                        WHERE id = %s
                    """, (
                        status,
                        tipo,
                        criterio,
                        servidor_id,
                        nota,
                        num_nota,
                        oficio,
                        monitoramento,
                        monitoramento_resposta,
                        observacoes,
                        req_id
                    ))

                if status == "ANALISADO":
                    cur.execute("""
                        SELECT MAX(h.data) AS ultima_data
                        FROM horas h
                        JOIN horas_requisicoes hr ON hr.hora_id = h.id
                        WHERE hr.requisicao_id = %s
                    """, (req_id,))
    
                    row = cur.fetchone()
                    if row and row["ultima_data"]:
                        cur.execute("""
                            UPDATE requisicoes
                            SET data_fim=%s, data_conclusao=NOW()
                            WHERE id=%s
                        """, (row["ultima_data"], req_id))
    
                con.commit()
                return "OK"
    
            # ==========================================
            # ATUALIZAR DATA → ADMIN
            # ==========================================
            if acao == "atualizar_campo":
                if session["perfil"] != "admin":
                    return "Acesso negado", 403
    
                campo = request.form.get("campo")
                valor = request.form.get("valor")
    
                if campo not in ("data_inicio", "data_fim"):
                    return "Campo inválido", 400
    
                cur.execute(
                    f"UPDATE requisicoes SET {campo}=%s WHERE id=%s",
                    (valor or None, req_id)
                )
    
                con.commit()
                return "OK"
    
            return "Ação desconhecida", 400
    
        except Exception as e:
            con.rollback()
            print("ERRO POST /requisicoes:", e)
            return "Erro interno", 500
    
        finally:
            con.close()

    # ======================
    # LISTAGEM (OTIMIZADA)
    # ======================
    page = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 200))  # 200 é o ideal
    offset = (page - 1) * per_page
    
    status = request.args.get("status")
    q = request.args.get("q")
    
    base_sql = """
    FROM requisicoes r
    LEFT JOIN colaboradores c ON c.id = r.servidor_id
    WHERE 1=1
    """
    params = []
    
    if status:
        base_sql += " AND r.status_analise = %s"
        params.append(status)
    
    if q:
        base_sql += """
        AND (
            r.chave ILIKE %s OR
            r.sigla ILIKE %s OR
            r.valor_requisicao::text ILIKE %s OR
            c.nome ILIKE %s
        )
        """
        like = f"%{q}%"
        params.extend([like, like, like, like])
    
    # =================================================
    # BUSCA UMA A MAIS PARA SABER SE TEM PRÓXIMA PÁGINA
    # =================================================
    cur.execute(
        f"""
        SELECT r.*, c.nome AS servidor
        {base_sql}
        ORDER BY r.created_at DESC
        LIMIT %s OFFSET %s
        """,
        params + [per_page + 1, offset]
    )
    
    rows = cur.fetchall()
    
    # Detecta se existe próxima página
    tem_proxima = len(rows) > per_page
    rows = rows[:per_page]
    
    # Detecta se existe página anterior
    tem_anterior = page > 1
    
    # Colaboradores (mantém como está)
    cur.execute("SELECT id, nome FROM colaboradores ORDER BY nome")
    colaboradores = cur.fetchall()
    
    from datetime import date
    hoje = date.today().isoformat()
    
    con.close()

    html = """
    
    <style>
    tr.andamento { background:#ffe5e5; }
    tr.analisando { background:#fff7cc; }
    tr.analisado { background:#e5ffe5; }
    
    .btn.andamento { background:#d9534f; color:#fff }
    .btn.analisando { background:#f0ad4e }
    .btn.analisado { background:#5cb85c; color:#fff }
    .btn.all { background:#0275d8; color:#fff }
    .btn.ativo { outline:2px solid #000; }
    
   #tbl {
    border-collapse: collapse;
    position: relative;
    }
    
    /* cabeçalho fixo */
    #tbl th {
        position: sticky;
        top: 0;
        background: #dce8fb;
        z-index: 10;
    }
    
    /* coluna chave */
    #tbl th:nth-child(1),
    #tbl td:nth-child(1) {
        position: sticky;
        left: 0;
        background: white;
        z-index: 9;
        min-width:170px;
    }
    
    /* coluna corte */
    #tbl th:nth-child(2),
    #tbl td:nth-child(2) {
        position: sticky;
        left: 170px;
        background: white;
        z-index: 9;
        min-width:110px;
    }
    
    /* cabeçalho acima das colunas */
    #tbl th:nth-child(1),
    #tbl th:nth-child(2) {
        z-index: 11;
    }
    </style>
    
    <h3>Requisições</h3>
    
     <div style="margin-bottom:10px;display:flex;gap:8px;flex-wrap:wrap">
        <a class="btn all {% if not status %}ativo{% endif %}" href="/requisicoes">TODOS</a>
        <a class="btn andamento {% if status=='ANDAMENTO' %}ativo{% endif %}"
           href="/requisicoes?status=ANDAMENTO{% if q %}&q={{ q }}{% endif %}">ANDAMENTO</a>
        <a class="btn analisando {% if status=='ANALISANDO' %}ativo{% endif %}"
           href="/requisicoes?status=ANALISANDO{% if q %}&q={{ q }}{% endif %}">ANALISANDO</a>
        <a class="btn analisado {% if status=='ANALISADO' %}ativo{% endif %}"
           href="/requisicoes?status=ANALISADO{% if q %}&q={{ q }}{% endif %}">ANALISADO</a>
    </div>

    <form method="get" style="margin-bottom:10px;">
        {% if status %}
            <input type="hidden" name="status" value="{{ status }}">
        {% endif %}
        <input type="text"
               name="q"
               value="{{ q or '' }}"
               placeholder="Pesquisar em todas as requisições..."
               style="width:100%;padding:8px;">
    </form>

      <div style="margin:10px 0; display:flex; gap:10px; align-items:center; flex-wrap:wrap;">
    
        {% if page > 1 %}
            <a href="?page=1{% if status %}&status={{status}}{% endif %}{% if q %}&q={{q}}{% endif %}">
                ⏮ Primeira
            </a>
    
            <a href="?page={{ page-1 }}{% if status %}&status={{status}}{% endif %}{% if q %}&q={{q}}{% endif %}">
                ◀ Anterior
            </a>
        {% endif %}
    
        <span>Página {{ page }}</span>
    
        {% if tem_proxima %}
            <a href="?page={{ page+1 }}{% if status %}&status={{status}}{% endif %}{% if q %}&q={{q}}{% endif %}">
                Próxima ▶
            </a>
        {% endif %}
    
    </div>
    <div style="overflow:auto; margin-top:10px; max-height:70vh;">
    <table id="tbl" style="min-width:1400px;">
        <tr>
            <th>Chave</th>
            <th>Corte</th>
            <th>Valor</th>
            <th>Status</th>
            <th>Tipo</th>
            <th>Critério</th>
            <th>Responsável</th>
            <th>Início</th>
            <th>Fim</th>
            <th>Nota</th>
            <th>Nº Nota</th>
            <th>Ofício</th>
            <th>Monitoramento</th>
            <th>Ações</th>
        </tr>

        {% for r in rows %}
        <tr class="{{ r.status_analise|lower }}">
            <td>{{ r.chave }}</td>
            <td>{{ fmt(r.data_corte) if r.data_corte else "" }}</td>
            <td>{{ fmt_br(r.valor_requisicao) }}</td>

            <td style="white-space:nowrap;">
                <select onchange="salvar({{ r.id }})"
                        id="status_{{ r.id }}"
                        style="min-width:90px; padding:2px;">
                    <option value=""></option>
                    {% for s in ['ANDAMENTO','ANALISANDO','ANALISADO'] %}
                        <option value="{{ s }}" {% if r.status_analise==s %}selected{% endif %}>
                            {{ s }}
                        </option>
                    {% endfor %}
                </select>
            </td>

            <td style="white-space:nowrap;">
                <select onchange="salvar({{ r.id }})"
                        id="tipo_{{ r.id }}"
                        style="min-width:70px; padding:2px;">
                    <option value=""></option>
                    {% for t in ['CONTRATAÇÃO','LIQUIDAÇÃO','ADITAMENTO'] %}
                        <option value="{{ t }}" {% if r.tipo==t %}selected{% endif %}>
                            {{ t }}
                        </option>
                    {% endfor %}
                </select>
            </td>

            <td style="white-space:nowrap;">
                <select onchange="salvar({{ r.id }})"
                        id="criterio_{{ r.id }}"
                        style="min-width:70px; padding:2px;">

                    <option value=""></option>
                    {% for c in ['MATERIALIDADE','RELEVÂNCIA','RISCO','ENGENHARIA'] %}
                        <option value="{{ c }}" {% if r.criterio==c %}selected{% endif %}>
                            {{ c }}
                        </option>
                    {% endfor %}
                </select>
            </td>

            <td>
                <select onchange="salvar({{ r.id }})" id="servidor_{{ r.id }}">
                    <option value=""></option>
                    {% for col in colaboradores %}
                        <option value="{{ col.id }}" {% if r.servidor_id==col.id %}selected{% endif %}>
                            {{ col.nome }}
                        </option>
                    {% endfor %}
                </select>
            </td>

            <td>
                <input type="date"
                       value="{{ r.data_inicio if r.data_inicio else '' }}"
                       onchange="atualizarCampo({{ r.id }}, 'data_inicio', this.value)">

            </td>

            <td>
                <input type="date"
                       value="{{ r.data_fim or '' }}"
                       onchange="atualizarCampo({{ r.id }}, 'data_fim', this.value)">
            </td>

            <td>
                <select onchange="salvar({{ r.id }})"
                id="nota_{{ r.id }}"
                style="min-width:75px;">
                    <option value=""></option>
                    <option value="SIM" {% if r.nota=='SIM' %}selected{% endif %}>SIM</option>
                    <option value="NÃO" {% if r.nota=='NÃO' %}selected{% endif %}>NÃO</option>
                </select>
            </td>

            <td>
                <input type="text"
                       value="{{ r.num_nota or '' }}"
                       id="num_nota_{{ r.id }}"
                       onchange="salvar({{ r.id }})"
                       style="width:90px;">
            </td>

            <td>
                <input type="text"
                       value="{{ r.oficio or '' }}"
                       id="oficio_{{ r.id }}"
                       onchange="salvar({{ r.id }})"
                       style="width:100px;">
            </td>

            <td>
                <select onchange="salvar({{ r.id }})"
                        id="monitoramento_{{ r.id }}">
                    <option value=""></option>
                    <option value="SIM" {% if r.monitoramento=='SIM' %}selected{% endif %}>SIM</option>
                    <option value="NÃO" {% if r.monitoramento=='NÃO' %}selected{% endif %}>NÃO</option>
                </select>
            </td>
            
            <td style="white-space:nowrap;">
                <a href="/requisicao/{{ r.id }}" title="Ver">🔍</a>
                &nbsp;
                <a href="/requisicoes/editar/{{ r.id }}" title="Editar">✏️</a>
                &nbsp;
                <button onclick="excluir({{ r.id }})" title="Excluir">🗑️</button>
            </td>

        </tr>
        {% endfor %}
    </table>
    </div>
    <script>

    function salvar(id){
        let fd = new FormData();
        fd.append("acao","atualizar");
        fd.append("id",id);
        fd.append("status_analise", document.getElementById("status_"+id).value);
        fd.append("tipo", document.getElementById("tipo_"+id).value);
        fd.append("criterio", document.getElementById("criterio_"+id).value);
        fd.append("servidor_id", document.getElementById("servidor_"+id).value);

        fd.append("nota", document.getElementById("nota_"+id).value);
        fd.append("num_nota", document.getElementById("num_nota_"+id).value);
        fd.append("oficio", document.getElementById("oficio_"+id).value);
        fd.append("monitoramento", document.getElementById("monitoramento_"+id).value);

        fetch("/requisicoes", { method:"POST", body: fd })
            .then(r => r.text())
            .then(resp => {
                if (resp !== "OK") {
                    alert("Erro ao salvar");
                    console.error(resp);
                }
            })
            .catch(err => {
                alert("Erro de rede");
                console.error(err);
            });
    }

    function excluir(id){
        if(!confirm("Excluir esta requisição?")) return;
        let fd = new FormData();
        fd.append("acao","excluir");
        fd.append("id",id);

        fetch("/requisicoes",{method:"POST", body:fd})
            .then(()=>location.reload());
    }
    
function atualizarCampo(id, campo, valor){
    let fd = new FormData();
    fd.append("acao","atualizar_campo");
    fd.append("id", id);
    fd.append("campo", campo);
    fd.append("valor", valor);

    fetch("/requisicoes", { method:"POST", body:fd });
}
    </script>
    """

    return render_template_string(
        BASE.replace("{% block content %}{% endblock %}", html),
        rows=rows,
        colaboradores=colaboradores,
        hoje=hoje,
        user=session["user"],
        perfil=session["perfil"],
        page=page,
        tem_proxima=tem_proxima,  # ✅ substitui total_pages
        status=status,
        q=q,                     # ✅ mantém busca
        fmt=fmt,
        fmt_br=fmt_br,
    )

@app.route("/requisicoes/editar/<int:id>", methods=["GET","POST"])
def editar_requisicao(id):
    if "user" not in session:
        return redirect("/")

    if session["perfil"] != "admin":
        return "Acesso negado", 403

    con = get_db()
    cur = con.cursor()

    if request.method == "POST":
        cur.execute("""
            UPDATE requisicoes
            SET
                tipo = NULLIF(%s,''),
                status_analise = NULLIF(%s,''),
                criterio = NULLIF(%s,''),
                servidor_id = %s,
                nota = NULLIF(%s,''),
                num_nota = NULLIF(%s,''),
                oficio = NULLIF(%s,''),
                monitoramento = NULLIF(%s,''),
                monitoramento_resposta = NULLIF(%s,''),
                observacoes = NULLIF(%s,'')
            WHERE id = %s
        """, (
            request.form.get("tipo"),
            request.form.get("status_analise"),
            request.form.get("criterio"),
            request.form.get("servidor_id") or None,
            request.form.get("nota"),
            request.form.get("num_nota"),
            request.form.get("oficio"),
            request.form.get("monitoramento"),
            request.form.get("monitoramento_resposta"),
            request.form.get("observacoes"),
            id
        ))
        con.commit()
        con.close()
        return redirect("/requisicoes")

    cur.execute("SELECT * FROM requisicoes WHERE id = %s", (id,))
    r = cur.fetchone()

    cur.execute("SELECT id, nome FROM colaboradores ORDER BY nome")
    colaboradores = cur.fetchall()

    con.close()

    html = """
    <h3>Editar Requisição</h3>
    
    <form method="post">
    
    <fieldset style="border:1px solid #ccc;padding:10px">
    <legend><b>Dados da Requisição (somente leitura)</b></legend>
    
    <table style="width:100%">
    <tr><td><b>Chave</b></td><td>{{ r.chave }}</td></tr>
    <tr><td><b>Nº Requisição</b></td><td>{{ r.requisicao_num }}</td></tr>
    <tr><td><b>Sigla</b></td><td>{{ r.sigla }}</td></tr>
    <tr><td><b>Secretaria</b></td><td>{{ r.secretaria }}</td></tr>
    <tr><td><b>Tipo Documento</b></td><td>{{ r.tipo_documento }}</td></tr>
    <tr><td><b>Valor</b></td><td>{{ fmt_br(r.valor_requisicao) }}</td></tr>
    <tr><td><b>Solicitante</b></td><td>{{ r.nome_solicitante }}</td></tr>
    <tr><td><b>Status Atual</b></td><td>{{ r.status_atual }}</td></tr>
    <tr><td><b>Data Criação</b></td><td>{{ r.data_criacao }}</td></tr>
    <tr><td><b>Data Tramitação</b></td><td>{{ r.data_tramitacao }}</td></tr>
    <tr><td><b>Natureza Despesa</b></td><td>{{ r.natureza_despesa }}</td></tr>
    <tr><td><b>Item Despesa</b></td><td>{{ r.item_despesa }}</td></tr>
    <tr><td><b>Fornecedor</b></td><td>{{ r.nome_fornecedor }}</td></tr>
    <tr><td><b>Edital</b></td><td>{{ r.edital }}</td></tr>
    <tr><td><b>Contrato</b></td><td>{{ r.contrato }}</td></tr>
    <tr><td><b>Empenho</b></td><td>{{ r.empenho }}</td></tr>
    <tr><td><b>Ficha Despesa</b></td><td>{{ r.ficha_despesa }}</td></tr>
    <tr><td><b>Data Corte</b></td><td>{{ r.data_corte }}</td></tr>
    </table>
    
    </fieldset>
    
    <br>
    
    <fieldset style="border:1px solid #ccc;padding:10px">
    <legend><b>Análise CGM</b></legend>
    
    <label>Status da Análise</label><br>
    <select name="status_analise">
        <option value=""></option>
        {% for s in ['ANDAMENTO','ANALISANDO','ANALISADO'] %}
        <option value="{{s}}" {% if r.status_analise==s %}selected{% endif %}>{{s}}</option>
        {% endfor %}
    </select><br><br>
    
    <label>Tipo</label><br>
    <select name="tipo">
        <option value=""></option>
        {% for t in ['CONTRATAÇÃO','LIQUIDAÇÃO','ADITAMENTO'] %}
        <option value="{{t}}" {% if r.tipo==t %}selected{% endif %}>{{t}}</option>
        {% endfor %}
    </select><br><br>
    
    <label>Critério</label><br>
    <select name="criterio">
        <option value=""></option>
        {% for c in ['MATERIALIDADE','RELEVÂNCIA','RISCO','ENGENHARIA'] %}
        <option value="{{c}}" {% if r.criterio==c %}selected{% endif %}>{{c}}</option>
        {% endfor %}
    </select><br><br>
    
    <label>Responsável</label><br>
    <select name="servidor_id">
        <option value=""></option>
        {% for col in colaboradores %}
        <option value="{{col.id}}" {% if r.servidor_id==col.id %}selected{% endif %}>
            {{col.nome}}
        </option>
        {% endfor %}
    </select><br><br>

    <label>Nota</label><br>
    <select name="nota">
        <option value=""></option>
        <option value="SIM" {% if r.nota=='SIM' %}selected{% endif %}>SIM</option>
        <option value="NÃO" {% if r.nota=='NÃO' %}selected{% endif %}>NÃO</option>
    </select><br><br>
    
    <label>Nº Nota</label><br>
    <input type="text" name="num_nota"
           value="{{ r.num_nota or '' }}"
           style="width:200px;"><br><br>
    
    <label>Ofício</label><br>
    <input type="text" name="oficio"
           value="{{ r.oficio or '' }}"
           style="width:200px;"><br><br>
    
    <label>Monitoramento</label><br>
    <select name="monitoramento">
        <option value=""></option>
        <option value="SIM" {% if r.monitoramento=='SIM' %}selected{% endif %}>SIM</option>
        <option value="NÃO" {% if r.monitoramento=='NÃO' %}selected{% endif %}>NÃO</option>
    </select><br><br>
    
    <label>Resposta do Monitoramento</label><br>
    <input type="text"
           name="monitoramento_resposta"
           value="{{ r.monitoramento_resposta or '' }}"
           style="width:100%;"><br><br>
    
    <label>Observações</label><br>
    <textarea name="observacoes" style="width:100%;min-height:80px;">
    {{ r.observacoes or '' }}
    </textarea>
    </fieldset>
    
    <br>
    
    <button class="btn">Salvar</button>
    <a class="btn" href="/requisicoes">Cancelar</a>
    
    </form>
    """

    return render_template_string(
        BASE.replace("{% block content %}{% endblock %}", html),
        r=r,
        colaboradores=colaboradores,
        user=session["user"],
        perfil=session["perfil"],
        fmt_br=fmt_br,
    )

@app.route("/painel_requisicoes")
def dashboard():
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # =====================================================
    # UNIVERSO (DISTINTO POR CHAVE)
    # =====================================================
    cur.execute("""
        WITH universo AS (
            SELECT DISTINCT ON (chave)
                chave, sigla, secretaria, tipo, criterio,
                valor_requisicao, status_analise
            FROM requisicoes
            ORDER BY chave, created_at DESC
        )
        SELECT
            COUNT(*) AS qtd_universo,
            SUM(valor_requisicao) AS valor_universo,
            COUNT(*) FILTER (WHERE status_analise='ANALISADO') AS qtd_analisadas,
            SUM(valor_requisicao)
                FILTER (WHERE status_analise='ANALISADO') AS valor_analisado
        FROM universo;
    """)
    cards = cur.fetchone()

    # =====================================================
    # TABELA COMPARATIVA POR SIGLA + TOTAL
    # =====================================================
    cur.execute("""
        WITH universo AS (
            SELECT DISTINCT ON (chave)
                chave, sigla, valor_requisicao, status_analise, num_nota
            FROM requisicoes
            ORDER BY chave, created_at DESC
        ),
        base AS (
            SELECT
                sigla,
        
                COUNT(*) AS qtd_total,
                COUNT(*) FILTER (WHERE status_analise='ANALISADO') AS qtd_analisadas,
        
                SUM(valor_requisicao) AS valor_total,
                SUM(valor_requisicao)
                    FILTER (WHERE status_analise='ANALISADO') AS valor_analisado,
        
                COUNT(DISTINCT num_nota) FILTER (
                    WHERE num_nota IS NOT NULL AND num_nota <> ''
                ) AS qtd_notas,
        
                COUNT(*) FILTER (
                    WHERE num_nota IS NOT NULL AND num_nota <> ''
                ) AS qtd_req_com_nota,
        
                SUM(valor_requisicao) FILTER (
                    WHERE num_nota IS NOT NULL AND num_nota <> ''
                ) AS valor_req_com_nota
        
            FROM universo
            GROUP BY sigla
        )
        
        SELECT *,
            ROUND((valor_analisado/NULLIF(valor_total,0))*100,2) AS perc_valor,
            ROUND((qtd_analisadas::numeric/NULLIF(qtd_total,0))*100,2) AS perc_qtd
        FROM base
        ORDER BY sigla;
    """)
    tabela = cur.fetchall()

    # TOTAL GERAL
    total = {
    "sigla": "TOTAL",
    "qtd_total": cards["qtd_universo"],
    "qtd_analisadas": cards["qtd_analisadas"],
    "valor_total": cards["valor_universo"],
    "valor_analisado": cards["valor_analisado"],

    "perc_valor": round(cards["valor_analisado"] / cards["valor_universo"] * 100, 2)
        if cards["valor_universo"] else 0,

    "perc_qtd": round(cards["qtd_analisadas"] / cards["qtd_universo"] * 100, 2)
        if cards["qtd_universo"] else 0,

    # NOVOS CAMPOS
    "qtd_notas": sum(r["qtd_notas"] or 0 for r in tabela),

    "qtd_req_com_nota": sum(r["qtd_req_com_nota"] or 0 for r in tabela),

    "valor_req_com_nota": sum(r["valor_req_com_nota"] or 0 for r in tabela),
}

    # =====================================================
    # 4 – PIZZA CRITÉRIO (ANALISADAS)
    # =====================================================
    cur.execute("""
    WITH universo AS (
        SELECT DISTINCT ON (chave)
            chave, criterio
        FROM requisicoes
        WHERE status_analise = 'ANALISADO'
          AND criterio IS NOT NULL
        ORDER BY chave, created_at DESC
    )
    SELECT criterio, COUNT(*)::int AS qtd
    FROM universo
    GROUP BY criterio
    ORDER BY criterio;

    """)
    pizza_criterio = cur.fetchall()

    # =====================================================
    # 5 – QTD ANALISADA POR TIPO
    # =====================================================
    cur.execute("""
    WITH universo AS (
        SELECT DISTINCT ON (chave)
            chave, tipo
        FROM requisicoes
        WHERE status_analise='ANALISADO'
          AND tipo IS NOT NULL
        ORDER BY chave, created_at DESC
    )
    SELECT tipo, COUNT(*)::int AS qtd
    FROM universo
    GROUP BY tipo
    ORDER BY tipo;

    """)
    graf_tipo = cur.fetchall()

    # =====================================================
    # 6 – CRITÉRIO × SIGLA (EMPILHADO)
    # =====================================================
    cur.execute("""
        WITH universo AS (
            SELECT DISTINCT ON (chave)
                chave, sigla, criterio
            FROM requisicoes
            WHERE status_analise='ANALISADO'
              AND criterio IS NOT NULL
            ORDER BY chave, created_at DESC
        )
        SELECT sigla, criterio, COUNT(*)::int AS qtd
        FROM universo
        GROUP BY sigla, criterio
        ORDER BY sigla, criterio;
        """)
    empilhado = cur.fetchall()

    # =====================================================
    # 7 – VALOR ANALISADO POR CRITÉRIO
    # =====================================================
    cur.execute("""
        WITH universo AS (
            SELECT DISTINCT ON (chave)
                chave, criterio, valor_requisicao
            FROM requisicoes
            WHERE status_analise='ANALISADO'
              AND criterio IS NOT NULL
            ORDER BY chave, created_at DESC
        )
        SELECT criterio, SUM(valor_requisicao) AS valor
        FROM universo
        GROUP BY criterio
        ORDER BY criterio;
    
        """)
    barras_valor = cur.fetchall()

    # =====================================================
    # 8 – TABELA / GRÁFICO POR COLABORADOR
    # =====================================================
    cur.execute("""
        WITH base AS (
            SELECT DISTINCT ON (r.chave)
                r.chave,
                r.servidor_id,
                r.valor_requisicao
            FROM requisicoes r
            WHERE r.status_analise = 'ANALISADO'
              AND r.servidor_id IS NOT NULL
            ORDER BY r.chave, r.created_at DESC
        ),
        agregado AS (
            SELECT
                c.nome AS colaborador,
                COUNT(*) AS qtd_analisada,
                SUM(b.valor_requisicao) AS valor_analisado
            FROM base b
            JOIN colaboradores c ON c.id = b.servidor_id
            GROUP BY c.nome
        ),
        totais AS (
            SELECT
                SUM(qtd_analisada) AS total_qtd,
                SUM(valor_analisado) AS total_valor
            FROM agregado
        )
        
        -- 🔹 linhas por colaborador
        SELECT
            a.colaborador,
            a.qtd_analisada,
            ROUND(a.qtd_analisada::numeric / t.total_qtd * 100, 2) AS perc_qtd,
            a.valor_analisado,
            ROUND(a.valor_analisado / t.total_valor * 100, 2) AS perc_valor,
            1 AS ordem
        FROM agregado a
        CROSS JOIN totais t
        
        UNION ALL
        
        -- 🔹 linha TOTAL GERAL
        SELECT
            'TOTAL GERAL' AS colaborador,
            t.total_qtd,
            100.00 AS perc_qtd,
            t.total_valor,
            100.00 AS perc_valor,
            2 AS ordem
        FROM totais t
        
        ORDER BY ordem, qtd_analisada DESC;

    """)
    tabela_colaboradores = cur.fetchall()

    # =====================================================
    # 9 – CARDS
    # =====================================================
    cur.execute("""
    WITH notas AS (
        SELECT
            r.num_nota,
            SUM(r.valor_requisicao) AS valor_nota
        FROM requisicoes r
        WHERE r.num_nota IS NOT NULL
          AND r.num_nota <> ''
        GROUP BY r.num_nota
    ),
    auditoria AS (
        SELECT
            n.num_nota,
            n.valor_nota,
            na.valor_posterior
        FROM notas n
        LEFT JOIN notas_auditoria na
            ON na.num_nota = n.num_nota
    )
    SELECT
        COUNT(*) AS qtd_notas,
        SUM(valor_nota) AS valor_notas,
    
        -- 🔹 benefício só quando tem valor_posterior E valor_nota > valor_posterior
        SUM(
            CASE
                WHEN valor_posterior IS NOT NULL
                 AND valor_nota > valor_posterior
                THEN valor_nota - valor_posterior
                ELSE 0
            END
        ) AS beneficio,
    
        -- 🔹 base continua sendo TOTAL das notas (correto)
        SUM(valor_nota) AS base_total
    
    FROM auditoria
    """)
    cards_notas = cur.fetchone()

    perc_beneficio = (
    (cards_notas["beneficio"] / cards_notas["base_total"]) * 100
    if cards_notas["base_total"] else 0
)

    # 10 – Gráfico Notas por Sigla
    cur.execute("""
    SELECT
        sigla,
        COUNT(DISTINCT num_nota) AS qtd_notas,
        COUNT(*) FILTER (
            WHERE num_nota IS NOT NULL AND num_nota <> ''
        ) AS qtd_req_nota
    FROM requisicoes
    WHERE num_nota IS NOT NULL
      AND num_nota <> ''
    GROUP BY sigla
    ORDER BY sigla
    """)
    graf_notas_sigla = cur.fetchall()
    
    # 11 – Cards Reqs com Nota
    cur.execute("""
    SELECT
        COUNT(*) FILTER (
            WHERE num_nota IS NOT NULL AND num_nota <> ''
        ) AS qtd_req_nota,
    
        COUNT(*) FILTER (
            WHERE status_analise='ANALISADO'
        ) AS total_analisadas,
    
        SUM(valor_requisicao) FILTER (
            WHERE num_nota IS NOT NULL AND num_nota <> ''
        ) AS valor_req_nota,
    
        SUM(valor_requisicao) FILTER (
            WHERE status_analise='ANALISADO'
        ) AS valor_total_analisado
    
    FROM requisicoes
    """)
    card_req_nota = cur.fetchone()

    perc_qtd_req_nota = (
    card_req_nota["qtd_req_nota"] / card_req_nota["total_analisadas"] * 100
    if card_req_nota["total_analisadas"] else 0
)

    perc_valor_req_nota = (
        card_req_nota["valor_req_nota"] / card_req_nota["valor_total_analisado"] * 100
        if card_req_nota["valor_total_analisado"] else 0
    )
    
    cur.close()
    conn.close()

    # =====================================================
    # HTML
    # =====================================================
    return render_template_string("""
<!DOCTYPE html>
<html>
<head>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4"></script>
<script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-datalabels@2"></script>
<meta charset="utf-8">
<title>Dashboard</title>

<style>
th, td {
    text-align: left;
}

.chart-box {
    background:white;
    padding:10px;
    border-radius:12px;
    box-shadow:0 6px 14px rgba(0,0,0,.08);
    height:280px;

    /* ❌ REMOVER */
    /* display:flex;
       align-items:center;
       justify-content:center; */

    position: relative;
}

.chart-box canvas {
    width:100% !important;
    height:100% !important;
}


body { font-family: 'Segoe UI', sans-serif; background:#f4f6fa; padding:20px; }
.cards { display:grid; grid-template-columns:repeat(4,1fr); gap:20px; }
.card {
    background:white;
    border-radius:12px;
    padding:18px;
    box-shadow:0 6px 14px rgba(0,0,0,.08);
}
.card h4 { margin:0; color:#666; }
.card strong { font-size:22px; color:#1a3c8b; }

table {
    width:100%;
    background:white;
    margin-top:25px;
    border-collapse:collapse;
    box-shadow:0 6px 14px rgba(0,0,0,.08);
}
th,td { padding:10px; border-bottom:1px solid #eee; }
th { background:#1a3c8b; color:white; }
tfoot td { font-weight:bold; background:#eef2ff; }

canvas { background:white; border-radius:12px;
         box-shadow:0 6px 14px rgba(0,0,0,.08); }
.grid { display:grid; grid-template-columns:1fr 1fr; gap:25px; margin-top:25px; }
</style>
</head>

<body>

<h2>📊 Dashboard de Requisições</h2>

<div class="cards">
    <div class="card"><h4>Qtd Universo</h4><strong>{{ cards.qtd_universo }}</strong></div>
    <div class="card"><h4>Qtd Analisadas</h4><strong>{{ cards.qtd_analisadas }}</strong></div>
    <div class="card"><h4>Valor Universo</h4><strong>R$ {{ fmt_br(cards.valor_universo) }}</strong></div>
    <div class="card"><h4>Valor Analisado</h4><strong>R$ {{ fmt_br(cards.valor_analisado) }}</strong></div>
</div>
<br>
<div class="cards">
    <div class="card">
        <h4>Qtd Notas</h4>
        <strong>{{ cards_notas.qtd_notas }}</strong>
    </div>

    <div class="card">
        <h4>Valor Notas</h4>
        <strong>R$ {{ fmt_br(cards_notas.valor_notas) }}</strong>
    </div>

    <div class="card">
        <h4>Requisições com Nota</h4>

        <div style="margin-top:8px;">
            <div>
                <span style="color:#666;">Qtd:</span><br>
                <strong>
                    {{ card_req_nota.qtd_req_nota }}
                    ({{ perc_qtd_req_nota|round(2) }}%)
                </strong>
            </div>

            <div style="margin-top:10px;">
                <span style="color:#666;">Valor:</span><br>
                <strong>
                    R$ {{ fmt_br(card_req_nota.valor_req_nota) }}
                    ({{ perc_valor_req_nota|round(2) }}%)
                </strong>
            </div>
        </div>
    </div>

    <div class="card">
        <h4>Benefício Financeiro</h4>
        <strong>
            R$ {{ fmt_br(cards_notas.beneficio) }}
            ({{ perc_beneficio|round(2) }}%)
        </strong>
    </div>
</div>

<h3>📋 Comparativo por Sigla</h3>
<table>
<tr>
<th>Sigla</th><th>Qtd</th><th>Analis.</th><th>%</th>
<th>Valor</th><th>Vlr Analis.</th><th>%</th><th>Qtd Not.</th><th>Reqs Not.</th><th>Valor Reqs. Not.</th>
</tr>
{% for r in tabela %}
<tr>
<td>{{ r.sigla }}</td>
<td>{{ r.qtd_total }}</td>
<td>{{ r.qtd_analisadas }}</td>
<td>{{ r.perc_qtd }}%</td>
<td>R$ {{ fmt_br(r.valor_total) }}</td>
<td>R$ {{ fmt_br(r.valor_analisado) }}</td>
<td>{{ r.perc_valor }}%</td>
<td>{{ r.qtd_notas }}</td>
<td>{{ r.qtd_req_com_nota }}</td>
<td>R$ {{ fmt_br(r.valor_req_com_nota) }}</td>
</tr>
{% endfor %}
<tfoot>
<tr>
<td>{{ total.sigla }}</td>
<td>{{ total.qtd_total }}</td>
<td>{{ total.qtd_analisadas }}</td>
<td>{{ total.perc_qtd }}%</td>
<td>R$ {{ fmt_br(total.valor_total) }}</td>
<td>R$ {{ fmt_br(total.valor_analisado) }}</td>
<td>{{ total.perc_valor }}%</td>
<td>{{ total.qtd_notas }}</td>
<td>{{ total.qtd_req_com_nota }}</td>
<td>R$ {{ fmt_br(total.valor_req_com_nota) }}</td>
</tr>
</tfoot>
</table>

<h3>👥 Análise por Colaborador</h3>

<table>
<thead>
<tr>
    <th>Colaborador</th>
    <th>Qtd Analisada</th>
    <th>% Qtd</th>
    <th>Valor Analisado</th>
    <th>% Valor</th>
</tr>
</thead>

<tbody>
{% for r in tabela_colaboradores if r.colaborador != 'TOTAL GERAL' %}
<tr>
    <td>{{ r.colaborador }}</td>
    <td>{{ r.qtd_analisada }}</td>
    <td>{{ r.perc_qtd }}%</td>
    <td>R$ {{ fmt_br(r.valor_analisado) }}</td>
    <td>{{ r.perc_valor }}%</td>
</tr>
{% endfor %}
</tbody>

<tfoot>
{% for r in tabela_colaboradores if r.colaborador == 'TOTAL GERAL' %}
<tr>
    <td>{{ r.colaborador }}</td>
    <td>{{ r.qtd_analisada }}</td>
    <td>{{ r.perc_qtd }}%</td>
    <td>R$ {{ fmt_br(r.valor_analisado) }}</td>
    <td>{{ r.perc_valor }}%</td>
</tr>
{% endfor %}
</tfoot>
</table>

<div class="grid">
    <div class="chart-box"><canvas id="pizza"></canvas></div>
    <div class="chart-box"><canvas id="tipo"></canvas></div>
    <div class="chart-box"><canvas id="empilhado"></canvas></div>
    <div class="chart-box"><canvas id="valor"></canvas></div>
</div>
<div class="chart-box" style="height:400px;">
    <canvas id="notas_sigla"></canvas>
</div>

<script>
Chart.register(ChartDataLabels);

/* ============================
   DADOS (ANTES DOS GRÁFICOS)
============================ */
const pizzaLabels = {{ pizza_criterio | map(attribute='criterio') | list | tojson }};
const pizzaData   = {{ pizza_criterio | map(attribute='qtd') | list | tojson }};

const tipoLabels  = {{ graf_tipo | map(attribute='tipo') | list | tojson }};
const tipoData    = {{ graf_tipo | map(attribute='qtd') | list | tojson }};

const dadosEmpilhado = {{ empilhado | tojson }};
const siglas = [...new Set(dadosEmpilhado.map(d => d.sigla))];
const criterios = [...new Set(dadosEmpilhado.map(d => d.criterio))];

const valorLabels = {{ barras_valor | map(attribute='criterio') | list | tojson }};
const valorData   = {{ barras_valor | map(attribute='valor') | list | tojson }};

const notasSiglaLabels = {{ graf_notas_sigla | map(attribute='sigla') | list | tojson }};
const notasData        = {{ graf_notas_sigla | map(attribute='qtd_notas') | list | tojson }};
const reqsData         = {{ graf_notas_sigla | map(attribute='qtd_req_nota') | list | tojson }};
</script>


<script>
/* ============================
   1 - PIZZA CRITÉRIO
============================ */

if (pizzaData.length) {
    new Chart(document.getElementById("pizza"), {
        type: 'pie',
        data: {
            labels: pizzaLabels,
            datasets: [{
                data: pizzaData,
                backgroundColor: ['#1a3c8b','#4caf50','#ff9800','#e53935']
            }]
        },
        options: {
            responsive: true,
            maintainAspectRatio: false,
            plugins: {
                datalabels: {
                    color: '#fff',
                    font: { weight: 'bold' },
                    formatter: (v, ctx) => {
                        const total = ctx.chart.data.datasets[0].data
                            .reduce((a,b)=>a+b,0);
                        return total
                            ? ((v/total)*100).toFixed(0)+'%'
                            : '';
                    }
                }
            }
        }
    });
}

/* ============================
   5 - QTD POR TIPO
============================ */

if (tipoData.length) {
    new Chart(document.getElementById("tipo"), {
        type: 'pie',
        data: {
            labels: tipoLabels,
            datasets: [{
                data: tipoData,
                backgroundColor: ['#2196f3','#9c27b0','#ff5722']
            }]
        },
        options: {
            responsive: true,
            maintainAspectRatio: false,
            plugins: {
                datalabels: {
                    color: '#fff',
                    font: { weight: 'bold' }
                }
            }
        }
    });
}


/* ============================
   6 - EMPILHADO CRITÉRIO x SIGLA
============================ */

if (siglas.length && criterios.length) {
    new Chart(document.getElementById("empilhado"), {
        type: 'bar',
        data: {
            labels: siglas,
            datasets: criterios.map((c, i) => ({
                label: c,
                data: siglas.map(s => {
                    const r = dadosEmpilhado.find(
                        d => d.sigla === s && d.criterio === c
                    );
                    return r ? Number(r.qtd) : 0;
                }),
                backgroundColor: ['#1a3c8b','#4caf50','#ff9800','#e53935'][i % 4],
                stack: 'stack1'
            }))
        },
        options: {
            responsive: true,
            maintainAspectRatio: false,
            scales: {
                x: { stacked: true },
                y: { stacked: true, beginAtZero: true }
            }
        }
    });
}

/* ============================
   7 - VALOR POR CRITÉRIO (HORIZONTAL)
============================ */
if (valorData.length) {
    new Chart(document.getElementById("valor"), {
        type: 'bar',
        data: {
            labels: valorLabels,
            datasets: [{
                label: 'Valor Analisado (R$)',
                data: valorData.map(Number),
                backgroundColor: '#1a3c8b'
            }]
        },
        options: {
            responsive: true,
            maintainAspectRatio: false,
            indexAxis: 'y',
            scales: {
                x: { beginAtZero: true }
            },
            plugins: {
                datalabels: {
                    anchor: 'end',
                    align: 'right',
                    formatter: v =>
                        'R$ ' + v.toLocaleString('pt-BR',
                            { minimumFractionDigits:2 })
                }
            }
        }
    });
}

if (notasData.length) {
    new Chart(document.getElementById("notas_sigla"), {
        type: 'bar',
        data: {
            labels: notasSiglaLabels,
            datasets: [
                {
                    label: 'Qtd Notas',
                    data: notasData
                },
                {
                    label: 'Req com Nota',
                    data: reqsData
                }
            ]
        },
        options: {
            indexAxis: 'y',  // 👈 horizontal
            responsive: true,
            maintainAspectRatio: false,
            scales: {
                x: {
                    beginAtZero: true
                }
            },
            plugins: {
                datalabels: {
                    anchor: 'end',
                    align: 'right'
                }
            }
        }
    });
}
</script>

</body>
</html>
""",
cards=cards,
tabela=tabela,
total=total,
pizza_criterio=pizza_criterio,
graf_tipo=graf_tipo,
empilhado=empilhado,
barras_valor=barras_valor,
tabela_colaboradores=tabela_colaboradores,  # 👈 FALTAVA ISSO
cards_notas=cards_notas,
perc_beneficio=perc_beneficio,
graf_notas_sigla=graf_notas_sigla,
card_req_nota=card_req_nota,
perc_qtd_req_nota=perc_qtd_req_nota,
perc_valor_req_nota=perc_valor_req_nota,
fmt_br=fmt_br
)

@app.route("/notas-auditoria", methods=["GET","POST"])
def notas_auditoria():

    if "user" not in session:
        return redirect("/")

    con = get_db()
    cur = con.cursor()

    # =========================
    # PAGINAÇÃO
    # =========================
    try:
        page = int(request.args.get("page", 1))
        if page < 1:
            page = 1
    except:
        page = 1
    
    try:
        per_page = int(request.args.get("per_page", 25))
        if per_page not in (10, 25, 50, 100):
            per_page = 25
    except:
        per_page = 25

    offset = (page - 1) * per_page

    # =========================
    # SALVAR EDIÇÃO DA NOTA
    # =========================
    if request.method == "POST":

        num_nota = request.form.get("num_nota")
        valor_posterior = request.form.get("valor_posterior")

        if valor_posterior:
            valor_posterior = valor_posterior.strip()
            valor_posterior = valor_posterior.replace(".", "").replace(",", ".")
        else:
            valor_posterior = None

        observacoes = request.form.get("observacoes")
        status = request.form.get("status")

        cur.execute("""
        INSERT INTO notas_auditoria
        (num_nota, valor_posterior, observacoes, status)
        VALUES (%s,%s,%s,%s)
        ON CONFLICT (num_nota)
        DO UPDATE SET
            valor_posterior = EXCLUDED.valor_posterior,
            observacoes = EXCLUDED.observacoes,
            status = EXCLUDED.status
        """,(num_nota,valor_posterior,observacoes,status))

        con.commit()

    # =========================
    # BUSCA
    # =========================
    busca = request.args.get("busca","").strip()

    # =========================
    # COUNT TOTAL (PAGINAÇÃO)
    # =========================
    count_sql = """
    SELECT COUNT(DISTINCT r.num_nota)
    FROM requisicoes r
    LEFT JOIN colaboradores c ON c.id = r.servidor_id
    WHERE r.num_nota IS NOT NULL
      AND r.num_nota <> ''
    """

    count_params = []

    if session.get("perfil") != "admin":
        count_sql += " AND r.servidor_id = %s"
        count_params.append(session.get("user_id"))

    if busca:
        count_sql += """
        AND (
            r.num_nota ILIKE %s
            OR c.nome ILIKE %s
        )
        """
        count_params.append(f"%{busca}%")
        count_params.append(f"%{busca}%")

    cur.execute(count_sql, count_params)
    total = cur.fetchone()["count"]

    # =========================
    # LISTAR NOTAS (COM LIMIT)
    # =========================
    sql = """
    SELECT
        r.num_nota,

        COUNT(r.id) AS qtd_requisicoes,

        SUM(r.valor_requisicao) AS valor_nota,

        na.valor_posterior,
        na.observacoes,
        na.status,

        STRING_AGG(DISTINCT c.nome, ', ') AS responsavel,
        
        CASE 
            WHEN na.valor_posterior IS NOT NULL 
            THEN SUM(r.valor_requisicao) - na.valor_posterior
            ELSE NULL
        END AS diferenca

    FROM requisicoes r

    LEFT JOIN colaboradores c
        ON c.id = r.servidor_id

    LEFT JOIN notas_auditoria na
        ON na.num_nota = r.num_nota

    WHERE r.num_nota IS NOT NULL
      AND r.num_nota <> ''
    """

    params = []

    if session.get("perfil") != "admin":
        sql += " AND r.servidor_id = %s"
        params.append(session.get("user_id"))

    if busca:
        sql += """
        AND (
            r.num_nota ILIKE %s
            OR c.nome ILIKE %s
        )
        """
        params.append(f"%{busca}%")
        params.append(f"%{busca}%")

    sql += """
    GROUP BY
        r.num_nota,
        na.valor_posterior,
        na.observacoes,
        na.status

    ORDER BY r.num_nota
    LIMIT %s OFFSET %s
    """

    params.extend([per_page, offset])

    cur.execute(sql, params)
    notas = cur.fetchall()

    con.close()

    # =========================
    # HTML
    # =========================
    html = """

    <h2>Notas de Auditoria</h2>

    <a href="/exportar-notas-auditoria" class="btn btn-success">
    Exportar Notas Auditoria
    </a>

    <br><br>

    <form method="get">
        <input type="text" name="busca"
        placeholder="Buscar por nota ou colaborador"
        value="{{request.args.get('busca','')}}">
        
        Exibir:
        <select name="per_page" onchange="this.form.submit()">
            {% for opt in [10,25,50,100] %}
            <option value="{{opt}}" {% if opt==per_page %}selected{% endif %}>
                {{opt}}
            </option>
            {% endfor %}
        </select>

        <button class="btn">Pesquisar</button>
    </form>

    <br>

    <table border="1" cellpadding="6">

    <tr>
        <th>Nota</th>
        <th>Responsável</th>
        <th>Qtd Requisições</th>
        <th>Valor Nota</th>
        <th>Valor Posterior</th>
        <th>Diferença</th>
        <th>Status</th>
        <th>Observações</th>
        <th>Ação</th>
        <th>Salvar</th>
    </tr>

    {% for n in notas %}

    <form method="post">

    <tr>

        <td>
        {{n.num_nota}}
        <input type="hidden" name="num_nota" value="{{n.num_nota}}">
        </td>

        <td>{{n.responsavel}}</td>

        <td>{{n.qtd_requisicoes}}</td>

        <td>{{fmt_br(n.valor_nota)}}</td>

        <td>
        <input name="valor_posterior"
        value="{{fmt_br(n.valor_posterior) if n.valor_posterior else ''}}">
        </td>
        <td>
            {% if n.diferenca is not none %}
                {{fmt_br(n.diferenca)}}
            {% endif %}
        </td>
        <td>
        <select name="status">
            <option value=""></option>
            <option value="MONITORADA"
            {% if n.status=="MONITORADA" %}selected{% endif %}>
            Monitorada
            </option>
        </select>
        </td>

        <td>
        <input name="observacoes"
        value="{{n.observacoes or ''}}">
        </td>

        <td>
        <a href="/notas-auditoria/{{n.num_nota}}">
        Ver nota
        </a>
        </td>

        <td>
        <button>Salvar</button>
        </td>

    </tr>

    </form>

    {% endfor %}

    </table>

    <div style="margin-top:10px;">
        {% if page > 1 %}
            <a href="?page={{page-1}}&per_page={{per_page}}&busca={{request.args.get('busca','')}}">Anterior</a>
        {% endif %}

        Página {{page}}

        {% if page * per_page < total %}
            <a href="?page={{page+1}}&per_page={{per_page}}&busca={{request.args.get('busca','')}}">Próxima</a>
        {% endif %}
    </div>
    """

    return render_template_string(
        BASE.replace("{% block content %}{% endblock %}", html),
        notas=notas,
        fmt_br=fmt_br,
        user=session['user'],
        perfil=session['perfil'],
        request=request,
        page=page,
        per_page=per_page,
        total=total
    )
    
@app.route("/notas-auditoria/<path:num_nota>")
def ver_nota(num_nota):

    if "user" not in session:
        return redirect("/")

    conn = get_db()
    cur = conn.cursor()

    if session["perfil"] == "admin":

        cur.execute("""
        SELECT
            r.chave,
            r.secretaria,
            r.tipo,
            r.criterio,
            r.edital,
            r.contrato,
            r.nome_fornecedor,
            r.valor_requisicao,
            c.nome AS colaborador
        FROM requisicoes r
        LEFT JOIN colaboradores c
            ON c.id = r.servidor_id
        WHERE r.num_nota = %s
        ORDER BY r.chave
        """,(num_nota,))

    else:

        cur.execute("""
        SELECT
            r.chave,
            r.secretaria,
            r.tipo,
            r.criterio,
            r.edital,
            r.contrato,
            r.nome_fornecedor,
            r.valor_requisicao,
            c.nome AS colaborador
        FROM requisicoes r
        LEFT JOIN colaboradores c
            ON c.id = r.servidor_id
        WHERE r.num_nota = %s
        AND r.servidor_id = %s
        ORDER BY r.chave
        """,(num_nota, session["user_id"]))

    reqs = cur.fetchall()
    conn.close()

    html = """

    <div class="container mt-4">

    <h3>Nota {{num_nota}}</h3>

    <table class="table table-bordered table-sm">

    <thead class="table-light">
    <tr>
        <th>Requisição</th>
        <th>Secretaria</th>
        <th>Tipo</th>
        <th>Critério</th>
        <th>Edital</th>
        <th>Contrato</th>
        <th>Fornecedor</th>
        <th>Responsável</th>
        <th style="text-align:right">Valor</th>
    </tr>
    </thead>

    <tbody>

    {% for r in reqs %}

    <tr>
        <td>{{r.chave}}</td>
        <td>{{r.secretaria}}</td>
        <td>{{r.tipo}}</td>
        <td>{{r.criterio}}</td>
         <td>{{r.edital or ''}}</td>
         <td>{{r.contrato or ''}}</td>
          <td>{{r.nome_fornecedor or ''}}</td>
        <td>{{r.colaborador}}</td>
        <td style="text-align:right">
            {{fmt_br(r.valor_requisicao)}}
        </td>
    </tr>

    {% endfor %}

    </tbody>
    </table>

    <a href="/notas-auditoria" class="btn btn-secondary">
    Voltar
    </a>

    </div>
    """

    return render_template_string(
        BASE.replace("{% block content %}{% endblock %}", html),
        reqs=reqs,
        num_nota=num_nota,
        fmt_br=fmt_br
    )

@app.route("/exportar-notas-auditoria")
def exportar_notas_auditoria():

    if "user" not in session:
        return redirect("/")

    con = get_db()
    cur = con.cursor()

    sql = """
    SELECT
        r.num_nota,
        r.chave,
        r.secretaria,
        r.tipo,
        r.criterio,
        r.edital,
        r.contrato,
        r.nome_fornecedor,
        r.valor_requisicao,
        c.nome AS colaborador,
        na.valor_posterior,
        na.observacoes,
        na.status,
        SUM(r.valor_requisicao) OVER (PARTITION BY r.num_nota) AS valor_nota
    FROM requisicoes r
    LEFT JOIN colaboradores c
        ON c.id = r.servidor_id
    LEFT JOIN notas_auditoria na
        ON na.num_nota = r.num_nota
    WHERE r.num_nota IS NOT NULL
      AND r.num_nota <> ''
    """

    params = []

    if session.get("perfil") != "admin":
        sql += " AND r.servidor_id = %s"
        params.append(session.get("user_id"))

    sql += " ORDER BY r.num_nota, r.chave"

    cur.execute(sql, params)
    dados = cur.fetchall()
    con.close()

    import csv
    import io
    from collections import defaultdict
    
    def gerar():
    
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_ALL)
    
        # BOM
        yield "\ufeff"
    
        header = [
            "num_nota",
            "qtd_requisicoes",
            "valor_nota",
            "valor_posterior",
            "diferenca",
            "status",
            "observacoes",
            "chaves",
            "valores_requisicoes"
        ]
    
        writer.writerow(header)
        yield output.getvalue()
        output.seek(0)
        output.truncate(0)
    
        agrupado = defaultdict(list)
    
        for d in dados:
            agrupado[d["num_nota"]].append(d)
    
        for nota, itens in agrupado.items():
    
            base = itens[0]
    
            valor_nota = float(base["valor_nota"] or 0)
    
            if base["valor_posterior"] is not None:
                valor_post = float(base["valor_posterior"])
                diferenca = valor_nota - valor_post
            else:
                valor_post = None
                diferenca = None
    
            chaves = []
            valores = []
    
            for i in itens:
                chaves.append(str(i["chave"]))
                valores.append(fmt_br(i["valor_requisicao"]))
    
            linha = [
                str(nota),
                str(len(itens)),
                fmt_br(valor_nota),
                fmt_br(valor_post) if valor_post is not None else "",
                fmt_br(diferenca) if diferenca is not None else "",
                str(base["status"] or ""),
                str(base["observacoes"] or ""),
                "\n".join(chaves),   # agora funciona
                "\n".join(valores)   # agora funciona
            ]
    
            writer.writerow(linha)
            yield output.getvalue()
            output.seek(0)
            output.truncate(0)

    return Response(
        gerar(),
        mimetype="text/csv; charset=utf-8",
        headers={
            "Content-Disposition":
            "attachment; filename=notas_auditoria.csv"
        }
    )

@app.route("/seed")
def seed():
    executar_seed()
    return "Seed executado com sucesso!"

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
    
